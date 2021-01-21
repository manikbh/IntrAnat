#! /usr/bin/env python
# -*- coding: utf-8 -*-
#
# Localisation graphique des electrodes
#
# (c) Inserm U836 2012-2014 - Manik Bhattacharjee
#
# License GNU GPL v3
#

# Pour un seul patient
#
# 1) Charger les modeles d'electrode
# 2) Recherche de sujet/classement par date des sujets
# 3) Obtenir la liste des données nécessaires
# 4) Rechercher les référentiels disponibles (MNI/Talairach/IRM/CA-CP/CA-CPnormalise)
#     et tous les referentiels "image" (pre/post/IRMf/PET ?)
# 5) Interface : Ajout/Sélection/Suppression d'électrodes, labels (attention au numéro des plots)
#      Plot 1 = extremite
# 6) Afficher les coordonnées de chacun des plots dans le referentiel choisi
# 7) Exporter (et reimporter) les coordonnées des electrodes


#
# TODO
# Liens BrainVisa (chargement des images, sauvegarde des implantations)
# Affichage des labels --> problème de référentiel ?
# En normalisant et exportant : si un fichier va être écrasé, demander confirmation !
# RemoveElectrode : remove also the label objects
# Exportation : classer les plots pour que l'exportation comporte A0 A1 A2 B1 B3 et pas B3 A1 B1 A0...
# AimsMIRegister -> utilisation des fichiers TRM ou utilisation du recalage SPM ?
# Valider les templates d'electrodes avant de les utiliser : un plot doit TOUJOURS se nommer "Plot152" et pas "Element 21" ou "Plot 152" -> modifier l'éditeur de template


# for DTI : ./AimsMeshDistance --help


import sys, os, pickle, glob, numpy, re, string, time, subprocess, json, copy, csv, gc  #gc.get_referrers pour trouver où est encore déclarer une variable (pour problem quand variable déclarer en python et en c++)
import openpyxl

from PyQt5 import QtGui, QtCore, uic, Qt
from PyQt5.QtWidgets import QVBoxLayout

from numpy import *
from math import sqrt, cos, sin
from collections import OrderedDict

from soma import aims
from brainvisa import axon

#from soma.aims.spmnormalizationreader import readSpmNormalization
from brainvisa import anatomist
from brainvisa.data import neuroHierarchy
from brainvisa import registration

from externalprocesses import *
from MicromedListener import MicromedListener as ML
from referentialconverter import ReferentialConverter
from checkSpmVersion import *
from readSulcusLabelTranslationFile import *
from readFreesurferLabelFile import *
from brainvisa.processes import defaultContext
from TimerMessageBox import *
from generate_contact_colors import *
from bipoleSEEGColors import bipoleSEEGColors
from DeetoMaison import DeetoMaison

import ImportTheoreticalImplentation
from scipy import ndimage
from brainvisa.processes import *
from PIL import Image

import pdb

#import objgraph #if not install in a terminal : pip install objgraph --prefix /brainvisa_4.50/

#name = 'Anatomist Show MarsAtlas Parcels Texture'
#roles = ('viewer',)
#userLevel = 0

#def validation():
#  anatomist.validation()

#signature = Signature(
#  'texture_marsAtlas_parcels', ReadDiskItem('hemisphere marsAtlas parcellation texture', 'aims Texture formats', requiredAttributes={ 'regularized': 'false' }),
#  'white_mesh',ReadDiskItem( 'Hemisphere White Mesh', 'aims mesh formats' ),
#)


########## SPM calls
# Convert SPM normalization _sn.mat to vector field
spm_SnToField8 = """try, addpath(genpath(%s));spm('defaults', 'FMRI');spm_jobman('initcfg');
clear matlabbatch;
FileNameSN = '%s';
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.matname{1}=FileNameSN;
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.vox=[NaN NaN NaN];
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.bb=NaN*ones(2,3);
matlabbatch{1}.spm.util.defs.comp{1}.inv.space{1}=['%s' ',1'];
matlabbatch{1}.spm.util.defs.ofname='%s';
matlabbatch{1}.spm.util.defs.fnames='';
matlabbatch{1}.spm.util.defs.savedir.saveusr{1}=spm_str_manip(FileNameSN,'h');
matlabbatch{1}.spm.util.defs.interp=1;
spm_jobman('run',matlabbatch);catch, disp 'AN ERROR OCCURED'; end;quit;""" # %(FileNameSN, FileSource, ofname) -> '_sn.mat' file and source image FileSource (normalized with the _sn). For the Database, we want y_<subject>_inverse.nii, so we need ofname = '<subject>_inverse' --> Maybe should provide also the output dir ? Right now, same as _sn.mat

# API changed in SPM12...
spm_SnToField12 = """try, addpath(genpath(%s));spm('defaults', 'FMRI');spm_jobman('initcfg');
clear matlabbatch;
FileNameSN = '%s';
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.matname{1}=FileNameSN;
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.vox=[NaN NaN NaN];
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.sn2def.bb=NaN*ones(2,3);
matlabbatch{1}.spm.util.defs.comp{1}.inv.space = {'%s'};
matlabbatch{1}.spm.util.defs.out{1}.savedef.ofname = '%s';
matlabbatch{1}.spm.util.defs.out{1}.savedef.savedir.saveusr{1}=spm_str_manip(FileNameSN,'h');
spm_jobman('run',matlabbatch);catch, disp 'AN ERROR OCCURED'; end;quit;"""

#spm_SnToField = spm_SnToField12

spm_inverse_y_field12 = """try,addpath(genpath(%s));spm('defaults', 'FMRI');spm_jobman('initcfg');
clear matlabbatch;
matlabbatch{1}.spm.util.defs.comp{1}.inv.comp{1}.def = {%s};
matlabbatch{1}.spm.util.defs.comp{1}.inv.space = {%s};
matlabbatch{1}.spm.util.defs.out{1}.savedef.ofname = %s;
matlabbatch{1}.spm.util.defs.out{1}.savedef.savedir.saveusr = {%s};
spm_jobman('run',matlabbatch);catch, disp 'AN ERROR OCCURED'; end;quit;"""

spm_inverse_y_field8 ="""try,addpath(genpath(%s));spm('defaults', 'FMRI');spm_jobman('initcfg');

spm_jobman('run',matlabbatch);catch, disp 'AN ERROR OCCURED'; end;quit;"""

# Read deformation field y_<subject>_inverse.nii, apply the vector field to scanner-based coordinates of electrodes
spm_normalizePoints = """
try, addpath(genpath(%s));P='%s';
P1=spm_vol([P ',1,1']);
P2=spm_vol([P ',1,2']);
P3=spm_vol([P ',1,3']);
[V1,XYZ]=spm_read_vols(P1);
V2=spm_read_vols(P2);
V3=spm_read_vols(P3);

%% Apply tranformation to electrodes
PosElectrode = dlmread('%s');
wPosElectrode=PosElectrode;
for i1=1:size(PosElectrode,1)
D=(XYZ(1,:)-PosElectrode(i1,1)).^2+(XYZ(2,:)-PosElectrode(i1,2)).^2+(XYZ(3,:)-PosElectrode(i1,3)).^2;
[tmp,order]=sort(D);
tmp=tmp(1:18);      %%  cubic neighborhood
order=order(1:18);
W=1./tmp;           %%  weight inverse to distance
if sum(isinf(W))>0
W=[1 zeros(1,length(W)-1)];
end
wPosElectrode(i1,:)=[sum(V1(order).*W)./sum(W) sum(V2(order).*W)./sum(W) sum(V3(order).*W)./sum(W)];
end
dlmwrite('%s',wPosElectrode,'precision',18);
catch, disp 'AN ERROR OCCURED'; end;quit;
"""


#coregister and reslice and segmentation (for resection estimation)
spm_resection_estimation = """try, addpath(genpath(%s)); spm('defaults', 'FMRI'); spm_jobman('initcfg');
clear matlabbatch;

final_directory = %s;
if isdir(final_directory) == 0
   mkdir(final_directory)
end
matlabbatch{1}.spm.spatial.coreg.estwrite.ref = %s;
matlabbatch{1}.spm.spatial.coreg.estwrite.source = %s;
matlabbatch{1}.spm.spatial.coreg.estwrite.other = {''};
matlabbatch{1}.spm.spatial.coreg.estwrite.eoptions.cost_fun = 'nmi';
matlabbatch{1}.spm.spatial.coreg.estwrite.eoptions.sep = [4 2];
matlabbatch{1}.spm.spatial.coreg.estwrite.eoptions.tol = [0.02 0.02 0.02 0.001 0.001 0.001 0.01 0.01 0.01 0.001 0.001 0.001];
matlabbatch{1}.spm.spatial.coreg.estwrite.eoptions.fwhm = [7 7];
matlabbatch{1}.spm.spatial.coreg.estwrite.roptions.interp = 4;
matlabbatch{1}.spm.spatial.coreg.estwrite.roptions.wrap = [0 0 0];
matlabbatch{1}.spm.spatial.coreg.estwrite.roptions.mask = 0;
matlabbatch{1}.spm.spatial.coreg.estwrite.roptions.prefix = 'r';
matlabbatch{2}.spm.spatial.preproc.channel.vols = %s;
matlabbatch{2}.spm.spatial.preproc.channel.biasreg = 0.001;
matlabbatch{2}.spm.spatial.preproc.channel.biasfwhm = 60;
matlabbatch{2}.spm.spatial.preproc.channel.write = [0 0];
matlabbatch{2}.spm.spatial.preproc.tissue(1).tpm = %s;
matlabbatch{2}.spm.spatial.preproc.tissue(1).ngaus = 2;
matlabbatch{2}.spm.spatial.preproc.tissue(1).native = [1 0];
matlabbatch{2}.spm.spatial.preproc.tissue(1).warped = [0 0];
matlabbatch{2}.spm.spatial.preproc.warp.mrf = 1;
matlabbatch{2}.spm.spatial.preproc.warp.cleanup = 1;
matlabbatch{2}.spm.spatial.preproc.warp.reg = [0 0.001 0.5 0.05 0.2];
matlabbatch{2}.spm.spatial.preproc.warp.affreg = 'mni';
matlabbatch{2}.spm.spatial.preproc.warp.fwhm = 0;
matlabbatch{2}.spm.spatial.preproc.warp.samp = 3;
matlabbatch{2}.spm.spatial.preproc.warp.write = [0 0];
matlabbatch{3}.spm.spatial.preproc.channel.vols = %s;
matlabbatch{3}.spm.spatial.preproc.channel.biasreg = 0.001;
matlabbatch{3}.spm.spatial.preproc.channel.biasfwhm = 60;
matlabbatch{3}.spm.spatial.preproc.channel.write = [0 0];
matlabbatch{3}.spm.spatial.preproc.tissue(1).tpm = %s;
matlabbatch{3}.spm.spatial.preproc.tissue(1).ngaus = 2;
matlabbatch{3}.spm.spatial.preproc.tissue(1).native = [1 0];
matlabbatch{3}.spm.spatial.preproc.tissue(1).warped = [0 0];
matlabbatch{3}.spm.spatial.preproc.warp.mrf = 1;
matlabbatch{3}.spm.spatial.preproc.warp.cleanup = 1;
matlabbatch{3}.spm.spatial.preproc.warp.reg = [0 0.001 0.5 0.05 0.2];
matlabbatch{3}.spm.spatial.preproc.warp.affreg = 'mni';
matlabbatch{3}.spm.spatial.preproc.warp.fwhm = 0;
matlabbatch{3}.spm.spatial.preproc.warp.samp = 3;
matlabbatch{3}.spm.spatial.preproc.warp.write = [0 0];
matlabbatch{4}.spm.util.imcalc.input = {
                                        %s
                                        %s
                                        };
matlabbatch{4}.spm.util.imcalc.output = %s;
matlabbatch{4}.spm.util.imcalc.outdir = %s;
matlabbatch{4}.spm.util.imcalc.expression = 'i1-i2';
matlabbatch{4}.spm.util.imcalc.var = struct('name', {}, 'value', {});
matlabbatch{4}.spm.util.imcalc.options.dmtx = 0;
matlabbatch{4}.spm.util.imcalc.options.mask = 0;
matlabbatch{4}.spm.util.imcalc.options.interp = 1;
matlabbatch{4}.spm.util.imcalc.options.dtype = 4;
spm_jobman('run',matlabbatch);catch, disp 'AN ERROR OCCURED'; end;quit;"""


region_grow="""try,addpath(genpath(%s));
imgresec = %s;
iSeed = %s;
reseccenter = %s;
f_orient = %s;
f_orient = reshape(f_orient,4,4)';
centermatF = eye(4);
centermatF(:,4) = f_orient*[reseccenter 1]';
VF = spm_vol(imgresec);
reseccenter_pix = inv(VF.mat)*centermatF(:,4);
mat_to_grow=spm_read_vols(VF);
new_image = regiongrowing(mat_to_grow,iSeed,reseccenter_pix(1:3));
VF.fname = %s;
VF.private.mat0 = VF.mat;
spm_write_vol(VF,new_image);
catch, disp 'AN ERROR OCCURED'; end;quit;"""

def viewFile(filepath):
	""" Launches an external app to display the provided file (windows/linux/mac-specific methods) """
	if sys.platform.startswith('darwin'):
			subprocess.call(('open', filepath))
	elif os.name == 'nt':
			os.startfile(filepath)
	elif os.name == 'posix':
			subprocess.call(('xdg-open', filepath))


# Functions to sort the contacts A1,A2...,A10 and not A1, A10, A2..
def atoi(text):
  return int(text) if text.isdigit() else text
def natural_keys(text):
  """alist.sort(key=natural_keys) sorts in human order"""
  return [ atoi(c) for c in re.split('(\d+)', text) ]

##################### Electrode functions (to be exported in another file FIXME needs natural_keys function ##################################

from electrode import ElectrodeModel

def moveElectrode(target, entry, referential, newRef, a, meshes):
  # a is anatomist object
  if entry is None or target is None:
    print("entry or target is None")
    return


  #print "entry : "+repr(entry)+"target : "+repr(target)

  if newRef is None:
    newRef = a.createReferential()
  transl = target#[-target[0], -target[1], -target[2]]
  i = array([[1,0,0]])
  z = array([target,]) - array([entry,])
  if z[0][1] < 0.001 and z[0][2] < 0.001:# si i est colinéaire avec z, changer i car on n'obtiendra pas un vecteur perpendiculaire aux deux
    i = array([[0,1,0]])
  if linalg.norm(z) == 0:
	  return
  z = -z / linalg.norm(z)
  #print "z = "+repr(z)
  y = cross (i,z)
  y = -y/linalg.norm(y)
  #print "y = "+repr(y)
  x = cross(y,z)
  #print "x = "+repr(x)
  #m = [x[0][0], x[0][1], x[0][2], y[0][0],y[0][1], y[0][2], z[0][0],  z[0][1], z[0][2]]
  m = [x[0][0], y[0][0], z[0][0], x[0][1],y[0][1], z[0][1], x[0][2],  y[0][2], z[0][2]]
  try:
    transf = a.createTransformation(transl+m, origin = newRef, destination = referential)
  except:
    print("problem transformation")
    return
    pdb.set_trace()
  a.assignReferential(newRef, meshes)
  #print " Electrode moved : target = "+repr(target) + ", entry = "+repr(entry) + " and Matrix "+repr(transl+m)
  #pdb.set_trace()
  return (newRef, transf)

def createElectrode(target, entry, referential, ana=None, windows=None, model = None, dispMode=None, dispParams=None):
  elecModel = ElectrodeModel(ana)
  elecModel.open(model,dispMode, dispParams)
  #elecModel.setDisplayReferential(newRef)
  meshes = elecModel.getAnatomistObjects()
  (newRef, transf) = moveElectrode(target, entry, referential, None, ana, meshes)
  #pdb.set_trace()
  if windows is not None:
    ana.addObjects(meshes, windows)
  return (newRef, transf, elecModel)

def createBipole(target, entry, referential, ana=None, windows=None, model = None, dispMode=None, dispParams=None):
  
  elecModel = ElectrodeModel(ana)
  elecModel.open(str(model),dispMode, dispParams)
  #elecModel.setDisplayReferential(newRef)
  
  meshes = elecModel.getAnatomistObjects()
  #mesh = self.a.toAObject(aims.SurfaceGenerator.cylinder(aims.Point3df(p[0], p[1], p[2]), aims.Point3df(pEnd), r, r, 24, True, True))
  (newRef, transf) = moveElectrode(target, entry, referential, None, ana, meshes)

  if windows is not None:
    ana.addObjects(meshes, windows)
  return (newRef, transf, elecModel)


# Récupération des plots dans l'ordre plot1 -> plot n
def getPlots(elecModel):
  cyls = elecModel.getCylinders()
  return dict((name,cyls[name]) for name in cyls if cyls[name]['type'] == 'Plot')

def getPlotsNames(elecModel):
  cyls = elecModel.getCylinders()
  plots = [n for n in cyls if cyls[n]['type'] == 'Plot']
  return sorted(plots, key=natural_keys)


# Récupération des coordonnées des centres des plots dans le référentiel électrode
def getPlotsCenters(elecModel):
  plots = getPlots(elecModel)
  return dict((p, plots[p]['center']) for p in plots)


############### 3D text annotations from anagraphannotate.py #############"


byvertex = False
# This intermediate class is only here because I cannot (yet) make SIP
# generate a non-abstract class for TextObject binding. One day, I'll find out!
class TObj ( anatomist.anatomist.cpp.TextObject ):
  def __init__( self, message='', pos=[0,0,0] ):
    anatomist.anatomist.cpp.TextObject.__init__( self, message, pos )

class Props( object ):
  def __init__( self ):
    self.lvert = []
    self.lpoly = []
    self.usespheres = True
    self.colorlabels = True
    self.center = aims.Point3df()

def makelabel( a, label, gc, pos, ref, color, props ):
  """ Create a text label designating gc at position pos, using anatomist instance a, a color  and stores the generated elements in props"""
  objects = []
  # create a text object
  to = TObj( label )
  to.setScale( 0.1 )
  to.setName( 'label: ' + label )
  a.releaseObject(to)

  # If we want to show a sphere at the target point
  if props.usespheres:
    sph = aims.SurfaceGenerator.icosphere( gc, 2, 50 )
    asph = a.toAObject( sph )
    asph.setMaterial( diffuse=color )
    asph.setName( 'gc: ' + label )
    a.registerObject( asph, False )
    a.releaseObject(asph) #registerObject le "dérelease"
    objects.append( asph )
  # Choose the color of the label text
  if props.colorlabels:
    to.GetMaterial().set( { 'diffuse': color } )
  # texto is the label (2D texture on a rectangle), but defined to always face the camera
  texto = anatomist.anatomist.cpp.TransformedObject( [ to ], False, True, pos )
  texto.setDynamicOffsetFromPoint( props.center )
  texto.setName( 'annot: ' + label )
  objects.append( texto )
  # Add to the polygons a line (to  link the label and the target position)
  props.lpoly.append( aims.AimsVector_U32_2( ( len( props.lvert ),
    len( props.lvert ) + 1 ) ) )
  props.lvert += [ gc, pos ]
  a.registerObject( texto, False )
  a.assignReferential(ref, objects)
  a.releaseObject(texto)
  
  return objects

############################### Useful functions

def createItemDirs(item):
  """ Create the directories containing the provided WriteDiskItem and insert them in the BrainVisa database """
  # Copied from brainvisa.processes, in ExecutionContext._processExecution()
  dirname = os.path.dirname( item.fullPath() )
  dir=dirname
  dirs = []
  while not os.path.exists( dir ):
	dirs.append(dir)
	dir=os.path.dirname(dir)
  if dirs:
	try:
	  os.makedirs( dirname )
	except OSError, e:
	  if not e.errno == errno.EEXIST:
		# filter out 'File exists' exception, if the same dir has
		# been created concurrently by another instance of BrainVisa
		# or another thread
		raise
	for d in dirs:
	  dirItem=neuroHierarchy.databases.createDiskItemFromFileName(d, None)



######################### FENETRE PRINCIPALE ############################

class LocateElectrodes(QtGui.QDialog):

  def __init__(self, app=None, loadAll = True):

    # UI init
    if loadAll == True:
       QtGui.QWidget.__init__(self)
       self.ui = uic.loadUi("epilepsie-electrodes.ui", self)
       self.setWindowTitle('Epilepsie - localisation des electrodes - NOT FOR MEDICAL USAGE')
       # Widget 0 (buttons panel) will be at minimum size (stretch factor 0), the windows will fill up the rest
       self.splitter_2.setStretchFactor(0,0)
       self.splitter_2.setStretchFactor(1,1)
       # Equal size for both views
       self.splitter.setStretchFactor(0,1)
       self.splitter.setStretchFactor(1,1)
    

    self.nameEdit.setText('A')

    # Load the list of protocols, patients and electrode models from BrainVisa
    if loadAll == True:
      self.modalities = ['Raw T1 MRI', 'T2 MRI', 'CT', 'PET', 'Electrode Implantation Coronal Image', 'Electrode Implantation Sagittal Image','fMRI-epile', 'Statistic-Data','FLAIR', 'resection', 'FreesurferAtlas', 'FGATIR','HippoFreesurferAtlas']
      # Electrode models
      self.elecModelList = []
      self.elecModelListByName = []
      self.loadFromBrainVisa()



    # Init of variables
    self.app = app
    self.dispObj = {} # All displayable objects "t1mri-pre", "t2"...
    self.objtokeep = {} #all object we must keep alive for anatomist but not in other variables
    self.diskItems = {} # For each dispObj, a list of dictionnaries {'refName', 'refObj', 'refId', 'transf'}
    # Coordinates displayed using referential : 'Natif' par defaut
    self.coordsDisplayRef = 'Natif'
    self.referentialCombo.clear()
    self.referentialCombo.addItems(['Natif',])
    self.dispMode = 'real'
    self.dispParams = None
    self.t1preMniFieldPath = None
    self.t1pre2ScannerBasedId = None
    self.electrodes = []# {Les objets electrodes avec les coordonnées, les meshes
    self.bipoles = [] #{Les objects bipoles}
    self.electrodeTemplateStubs = [] # Un objet electrode par template disponible dans la base de données (chargé à la demande par getElectrodeTemplates)
    self.contacts = [] # {name:n, number:2, electrode: e, mesh:m}
    self.transfs = [] # identity transforms (must be stored)
    self.currentWindowRef = None # Referential used by windows (because pyanatomist AWindow.getReferential is not implemented yet)
    self.linkedRefs = [] # Referentials linked by a identity transform
    self.transf2Mni = {} # Transformation from T1 pre referential to MNI referential
    self.threads = [] # List of running threads
    self.t1pre2ScannerBasedTransform = None #Transfo from T1pre native to scanner-based referential (Anatomist Transformation object)
    self.brainvisaPatientAttributes = None # Attributes of a BrainVisa ReadDiskItem MRI of the loaded patient
    self.spmpath = None

    #self.MicromedListener = ML()

    # list of objects to display in window for each scenario (MNI, pre, post, etc)
    self.windowContent = { 'IRM pre':['T1pre','electrodes',],\
                           'IRM pre T2':['T2pre','electrodes',],\
                           'IRM pre + hemisphere droit':['T1pre','T1pre-rightHemi','electrodes',],\
                           'IRM pre + MARS ATLAS droit':['T1pre','right MARS ATLAS BIDULE','electrodes',],\
                           'IRM pre + hemisphere gauche':['T1pre','T1pre-leftHemi','electrodes',],\
                           'IRM pre + MARS ATLAS gauche':['T1pre','left MARS ATLAS BIDULE','electrodes',],\
                           'IRM pre + hemispheres':['T1pre','T1pre-rightHemi','T1pre-leftHemi','electrodes',],\
                           'IRM pre + hemispheres + tete':['T1pre','T1pre-rightHemi','T1pre-leftHemi', 'T1pre-head','electrodes',],\
                           'IRM post':['T1post','electrodes',],\
                           'IRM post T2':['T2post','electrodes',],\
                           'CT post':['CTpost','electrodes',],\
                           'PET pre':['PETpre','electrodes',],\
                           'FLAIR pre':['FLAIRpre','electrodes',],\
                           'FGATIR pre':['FGATIRpre','electrodes',],\
                           'fMRI pre':['fMRIpre','electrodes'],\
                           'Statistic Data':['Statisticspre','electrodes'],\
                           'IRM post-op':['T1postOp','electrodes',],\
                           'Resection':['Resection','electrodes',],\
                           'FreeSurferAtlas':['FreesurferAtlaspre','electrodes',],\
                           'HippoFreeSurferAtlas':['HippoFreesurferAtlaspre','electrodes',],\
    }
    self.windowCombo1.clear()
    self.windowCombo1.addItems(sorted(self.windowContent.keys()))
    self.windowCombo2.clear()
    self.windowCombo2.addItems(sorted(self.windowContent.keys()))

    # Anatomist windows
    if loadAll == True:
      self.wins=[]
      self.a = anatomist.Anatomist('-b') #Batch mode (hide Anatomist window)
      self.a.onCursorNotifier.add(self.clickHandler)

      layoutAx = QtGui.QHBoxLayout( self.windowContainer1 )
      self.axWindow = self.a.createWindow( 'Axial' )#, no_decoration=True )
      self.axWindow.setParent(self.windowContainer1)
      layoutAx.addWidget( self.axWindow.getInternalRep() )

      layoutSag = QtGui.QHBoxLayout( self.windowContainer2 )
      self.sagWindow = self.a.createWindow( 'Axial' )#, no_decoration=True )
      self.sagWindow.setParent(self.windowContainer2)
      layoutSag.addWidget( self.sagWindow.getInternalRep() )

      self.wins = [self.axWindow, self.sagWindow]

    # Get Transformation Manager
    self.transfoManager = registration.getTransformationManager()
    # Get ReferentialConverter (for Talairach, AC-PC...)
    self.refConv = ReferentialConverter()

    if loadAll == True:
      # Linking UI elements to functions
      self.loadPatientButton.clicked.connect(self.loadPatient)
      self.changePatientButton.clicked.connect(self.changePatient)
      self.patientList.itemDoubleClicked.connect(lambda x:self.loadPatient())
      self.protocolCombo.currentIndexChanged[int].connect(self.updateBrainvisaProtocol)
      self.filterSiteCombo.currentIndexChanged[int].connect(self.filterSubjects)
      self.filterYearCombo.currentIndexChanged[int].connect(self.filterSubjects)
      self.addElectrodeButton.clicked.connect(self.addElectrode)
      self.removeElectrodeButton.clicked.connect(self.removeElectrode)
      self.nameEdit.editingFinished.connect(self.editElectrodeName)
      self.targetButton.clicked.connect(self.updateTarget)
      self.entryButton.clicked.connect(self.updateEntry)
      self.electrodeList.currentRowChanged.connect(self.electrodeSelect)
      self.electrodeList.itemDoubleClicked.connect(self.electrodeGo)
      self.contactList.itemClicked.connect(self.contactSelect)
      self.contactList.itemDoubleClicked.connect(self.contactGo)
      self.typeComboBox.currentIndexChanged[str].connect(self.updateElectrodeModel)
      # itemClicked.connect(QListWidgetItem*) , currentItemChanged ( QListWidgetItem * current, QListWidgetItem * previous ), currentRowChanged ( int currentRow )
      self.electrodeLoadButton.clicked.connect(self.loadElectrodes)
      self.electrodeSaveButton.clicked.connect(self.saveElectrodes)
      self.normalizeExportButton.clicked.connect(self.normalizeExportElectrodes)
      #self.marsatlasExportButton, QtCore.SIGNAL.connect('clicked()'),self.parcelsExportElectrodes)
      self.colorConfigButton.clicked.connect(self.configureColors)
      self.dispModeCombo.currentIndexChanged[int].connect(self.updateDispMode)
      self.windowCombo1.currentIndexChanged[str].connect(lambda s: self.windowUpdate(0,s))
      self.windowCombo2.currentIndexChanged[str].connect(lambda s: self.windowUpdate(1,s))
      self.referentialCombo.currentIndexChanged[str].connect(self.updateCoordsDisplay)
      self.electrodeRefCheck.stateChanged.connect(self.updateElectrodeView)
      self.electrodeRefRotationSlider.valueChanged.connect(self.updateElectrodeViewRotation)

      self.Clipping_checkbox.clicked.connect(self.clippingUpdate)
      self.makefusionButton.clicked.connect(self.makeFusion)
      self.generateResectionArray.clicked.connect(self.generateResection)
      self.validateROIresection.clicked.connect(self.ROIResectiontoNiftiResection)
      self.deleteMarsAtlasfiles.clicked.connect(self.DeleteMarsAtlasFiles)
      self.generateDictionariesComboBox.activated[str].connect(self.generateDictionaries)
      self.ImportTheoriticalImplantation.clicked.connect(self.importRosaImplantation)
      self.approximateButton.clicked.connect(self.approximateElectrode)

      prefpath_imageimport = os.path.join(os.path.expanduser('~'), '.imageimport')
      try:
        if (os.path.exists(prefpath_imageimport)):
          filein = open(prefpath_imageimport, 'rb')
          prefs_imageimport = pickle.load(filein)
          self.spmpath = prefs_imageimport['spm']
          self.fileNoDBpath = prefs_imageimport['FileNoDBPath']
          filein.close()
      except:
        print 'NO SPM path found, will be unable to export MNI position'
        pass

      self.warningMEDIC()

    # Reload options, check brainvisa and matlab/SPM

  def closeEvent(self, event):
    self.quit(event)

  def quit(self, event=None):
    reply = QtGui.QMessageBox.question(self, 'Message',
            "Quit the software without saving ?", QtGui.QMessageBox.Yes |
            QtGui.QMessageBox.No, QtGui.QMessageBox.No)
    #if self.brainvisaPatientAttributes is not None:
      #self.saveElectrodes()
    # Remove the vector field to MNI if it was computed
    #self.clearT1preMniTransform()
    if reply == QtGui.QMessageBox.Yes:
      axon.processes.cleanup()
      if event is None:
        self.app.quit()
      else:
        event.accept()
    else:
      event.ignore()


  def warningMEDIC(self):
     
     shortwarning = TimerMessageBox(5,self)
     shortwarning.exec_()
     
     #messagebox = QtGui.QMessageBox(self)
     #messagebox.setWindowTitle("NOT FOR MEDICAL USAGE")
     #messagebox.setText("NOT FOR MEDICAL USAGE\n (closing automatically in {0} secondes.)".format(3))
     #messagebox.setStandardButtons(messagebox.NoButton)
     #self.timer2 = QtCore.QTimer()
     #self.time_to_wait = 3
     #def close_messagebox(e):
        #e.accept()
        #self.timer2.stop()
        #self.time_to_wait = 10
     #def decompte():
        #messagebox.setText("NOT FOR MEDICAL USAGE\n (closing automatically in {0} secondes.)".format(self.time_to_wait))
        #if self.time_to_wait <= 0:
           #messagebox.closeEvent = close_messagebox
           #messagebox.close()
        #self.time_to_wait -= 1
     #self.timer2.timeout.connect(decompte)
     #self.timer2.start(1000)
     #messagebox.exec_()

  def loadFromBrainVisa(self):
    # Find available patients in BV database
    print "LOADING DATA FROM BRAINVISA"
    rdi = ReadDiskItem( 'Subject', 'Directory',requiredAttributes={'_ontology':'brainvisa-3.2.0'}  ) #, requiredAttributes={'center':'Epilepsy'} )
    subjects = list( rdi._findValues( {}, None, False ) )
    protocols = list(set([s.attributes()['center'] for s in subjects if 'center' in s.attributes()]))
    # Fill the combo
    self.protocolCombo.clear()
    self.protocolCombo.addItems(sorted(protocols))
    self.allSubjects = subjects
    self.updateBrainvisaProtocol()

  def updateBrainvisaProtocol(self, idx=None):
    """Updates the UI when the selected protocol changes"""
    self.currentProtocol = str(self.protocolCombo.currentText())
    self.subjects = [s.attributes()['subject'] for s in self.allSubjects if 'center' in s.attributes() and s.attributes()['center'] == self.currentProtocol]
    print 'all subjects:' + repr(self.subjects)
    self.patientList.clear()
    self.patientList.addItems(sorted(self.subjects))
    # Update the filters
    sites = ['*',] + sorted(set([s.split('_')[0] for s in self.subjects]))
    years = ['*',] + sorted(set([s.split('_')[1] for s in self.subjects if len(s.split('_')) > 1]))
    self.filterSiteCombo.clear()
    self.filterSiteCombo.addItems(sites)
    self.filterYearCombo.clear()
    self.filterYearCombo.addItems(years)
    # Loading electrode models
    rdiEM = ReadDiskItem('Electrode Model', 'Electrode Model format', requiredAttributes={'center':self.currentProtocol})
    self.elecModelList = list (rdiEM._findValues( {}, None, False ) )
    elecNames = [e.attributes()['model_name'] for e in self.elecModelList]
    self.elecModelListByName = dict((e.attributes()['model_name'], e) for e in self.elecModelList)
    self.typeComboBox.clear()
    self.typeComboBox.addItems(elecNames)

  def filterSubjects(self, value=None):
    """Filtering subject list"""
    subs = self.subjects
    if str(self.filterSiteCombo.currentText()) != '*':
      subs = [s for s in subs if s.split('_')[0] == str(self.filterSiteCombo.currentText())]
    if str(self.filterYearCombo.currentText()) != '*':
      subs = [s for s in subs if len(s.split('_')) > 1 and s.split('_')[1] == str(self.filterYearCombo.currentText())]
    self.patientList.clear()
    self.patientList.addItems(sorted(subs))

  def getT1preMniTransform(self):
    """Returns the path of the transformation to MNI (vector field) and compute it if necessary (from _sn.mat)"""
    #pdb.set_trace()
    if self.t1preMniFieldPath is not None:
      return self.t1preMniFieldPath
    # Find _sn.mat


    if 'T1pre' not in self.dispObj:
      print "No T1pre loaded : cannot get MNI transform from it"
      return None

    #look for a y_file_inverse first
    rdi_inv_read = ReadDiskItem('SPM normalization inverse deformation field','NIFTI-1 image')
    di_inv_read = rdi_inv_read.findValue(self.diskItems['T1pre'])

    if di_inv_read is None:
        print "No inverse deformation field found in database"
    else:
        print "inverse deformation field found and used"
        #pdb.set_trace()
        self.t1preMniFieldPath = di_inv_read.fileName()
        return self.t1preMniFieldPath

    spm_version = checkSpmVersion(self.spmpath)

    #look for a y_file second
    rdi_y = ReadDiskItem('SPM normalization deformation field','NIFTI-1 image')
    di_y = rdi_y.findValue(self.diskItems['T1pre'])
    #pdb.set_trace()
    if di_y is None:
        print "No deformation field found in database"
    else:
        print "deformation field found and used"
        wdi_inverse = WriteDiskItem('SPM normalization inverse deformation field','NIFTI-1 image')
        dir_yinv_split = str(di_y.fileName()).split('/')
        name_yinverse = dir_yinv_split.pop()[2:]
        #name_yinverse.replace('.nii','_inverse.nii')
        dir_yinverse = "/".join(dir_yinv_split)
        di_inverse = wdi_inverse.findValue(di_y)
        #on fait l'inversion de la deformation
        #pdb.set_trace()
        #pour le moment ce bout de code ne marche qu'avec spm12
        if spm_version == '(SPM12)':
           print 'SPM12 used'
           matlabRun(spm_inverse_y_field12%("'"+self.spmpath+"'","'"+str(di_y.fileName())+"'","'"+self.dispObj['T1pre'].fileName()+"'","'"+name_yinverse.replace('.nii','_inverse.nii')+"'","'"+dir_yinverse+"'"))
        if spm_version == '(SPM8)':
           print 'SPM8 used'
           matlabRun(spm_inverse_y_field8%("'"+self.spmpath+"'","'"+str(di_y.fileName())+"'","'"+self.dispObj['T1pre'].fileName()+"'","'"+name_yinverse.replace('.nii','_inverse.nii')+"'","'"+dir_yinverse+"'"))

        self.t1preMniFieldPath = di_inverse.fileName()
        neuroHierarchy.databases.insertDiskItem( di_inverse, update=True )
        return self.t1preMniFieldPath

    #look for a _sn.mat if no y_file
    rdi = ReadDiskItem( 'SPM2 normalization matrix', 'Matlab file' )
    di = rdi.findValue(self.diskItems['T1pre'])
    #pdb.set_trace()
    if di is None:
      print "SPM deformation _sn.mat not found in database"
      return None

    # Convert to field
    #pdb.set_trace()
    wdi = WriteDiskItem( 'SPM normalization inverse deformation field', 'NIFTI-1 image' )
    diField = wdi.findValue(di)
    if diField is None:
      print "Cannot find path to save MNI vector field in the DB"
      return None
    #For a file /database/y_SubjectName_inverse.nii, get SubjectName_inverse
    ofname = os.path.basename(diField.fullPath()).lstrip('y_').rsplit('.',1)[0]

    #pdb.set_trace()

    if spm_version == '(SPM12)':
       print 'SPM12 used'
       matlabRun(spm_SnToField12%("'"+self.spmpath+"'",str(di.fullPath()), str(self.diskItems['T1pre'].fullPath()),  ofname) )
    if spm_version == '(SPM8)':
       print 'SPM8 used'
       matlabRun(spm_SnToField8%("'"+self.spmpath+"'",str(di.fullPath()), str(self.diskItems['T1pre'].fullPath()),  ofname) )

    if os.path.exists(diField.fullPath()):
      self.t1preMniFieldPath = diField.fullPath()
      return self.t1preMniFieldPath
    else:
      print "Matlab did not convert the MNI transform to vector field !"
      return None


  def clearT1preMniTransform(self):
    """Reset MNI transform field if it was generated"""
    if self.t1preMniFieldPath is not None:
      try:
        os.remove(self.t1preMniFieldPath) #to change with: self.removeDiskItems(di,eraseFiles=True)
      except:
        pass
      self.t1preMniFieldPath = None

  def changePatient(self):
      
      self.loadPatientButton.setEnabled(True)
      self.patientList.setEnabled(True)
      self.a.removeObjects(self.a.getObjects(),self.wins[0])
      self.a.removeObjects(self.a.getObjects(),self.wins[1])
      self.a.config()[ 'linkedCursor' ] = 0
      referentials=self.a.getReferentials()
      for element in referentials:
          if element.getInfos().get('name') not in ('Talairach-MNI template-SPM', 'Talairach-AC/PC-Anatomist'):
             self.a.deleteElements(element)
             
      #for element in self.electrodes:
          #for elecKeys in element.keys():
             #del element[elecKeys]
      #variablesG=globals()       
      #variablesL=locals()
      self.electrodeList.clear()
      self.contactList.clear()
      self.currentWindowRef = None
      listEl=[]
      #for el , value in self.dispObj.items():
      #      listEl.append(value)
      #self.__init__(self,loadAll=False)      
      #for el in listEl:
      #    if type(el)==list:
      #        listElec=el
      #    else:        
      #        self.a.deleteObjects(el)
      
      #del self.electrodes
      self.electrodes = []
      #del self.dispObj
      #parcourir les objets, détruire les fusions avant
      #for obj in self.dispObj: 
      self.currentElectrodes = []
      self.currentContacts = []
       
      #todelete = []
      #for name,obj in self.dispObj.items():
      #    if isinstance(obj.internalRep,anatomist.anatomist.cpp.MObject):
      #        todelete.append(name)
      #for name in todelete:
      #    del self.dispObj[name]
          
      self.dispObj={}
      
      #if hasattr(self,"objtokeep"):
      self.objtokeep = {}
      
      self.__init__(loadAll = False)
    
  
  def loadPatient(self, patient=None):
        
    if patient is None:
      patient = str(self.patientList.selectedItems()[0].text())
      
    volumes = []
    self.t1pre2ScannerBasedTransform = None
    self.clearT1preMniTransform()

    pre_select_1 = self.windowCombo1.currentText()
    pre_select_2 = self.windowCombo2.currentText()



    for moda in self.modalities:
      rdi2 = ReadDiskItem( moda, 'aims readable volume formats', requiredAttributes={'subject':patient, 'center':self.currentProtocol} )
      volumes.extend(list( rdi2._findValues( {}, None, False ) ))


    dictionnaire_list_images = {'IRM pre':['T1pre','electrodes',],\
                           'IRM pre + hemisphere droit':['T1pre','T1pre-rightHemi','electrodes',],\
                           'IRM pre + hemisphere gauche':['T1pre','T1pre-leftHemi','electrodes',],\
                           'IRM pre + hemispheres':['T1pre','T1pre-rightHemi','T1pre-leftHemi','electrodes',],\
                           'IRM pre + hemispheres + tete':['T1pre','T1pre-rightHemi','T1pre-leftHemi', 'T1pre-head','electrodes']}


    #pdb.set_trace()

    for t in volumes:
      if "skull_stripped" in t.fullName():
          continue
      self.brainvisaPatientAttributes = t.attributes()
      if (t.attributes()['modality'] == 't2mri') and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'IRM pre T2':['T2pre','electrodes']})
      elif (t.attributes()['modality'] == 't2mri') and ('post' in t.attributes()['acquisition']):
        dictionnaire_list_images.update({'IRM post T2':['T2post','electrodes']})
      elif (t.attributes()['modality'] == 't1mri') and ('post' in t.attributes()['acquisition']) and not ('postOp' in t.attributes()['acquisition']):
        dictionnaire_list_images.update({'IRM post':['T1post','electrodes']})
      elif (t.attributes()['modality'] == 'ct') and ('post' in t.attributes()['acquisition']) and not ('postOp' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'CT post':['CTpost','electrodes']})
      elif (t.attributes()['modality'] == 'ct') and ('postOp' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'CT post-op':['CTpostOp','electrodes']})   
      elif (t.attributes()['modality'] == 'pet') and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'PET pre':['PETpre','electrodes']})
      elif (t.attributes()['modality'] == 'flair') and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'FLAIR pre':['FLAIRpre','electrodes']})
      elif (t.attributes()['modality'] == 'fgatir') and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'FGATIR pre':['FGATIRpre','electrodes']})
      elif (t.attributes()['modality'] == 'fmri_epile') and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'fMRI pre'+ ' - ' + t.attributes()['subacquisition']:['fMRIpre','electrodes']})   #mettre le nom de la subacquisition
      elif t.attributes()['modality'] == 'statistic_data' and ('pre' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'Statistic Data' + ' - ' + t.attributes()['subacquisition']:['Statisticspre'+t.attributes()['subacquisition'],'electrodes']}) #mettre le nom de la subacquisition
      elif t.attributes()['modality'] == 'statistic_data' and ('post' in t.attributes()['acquisition']) and not ('postOp' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'Statistic Data' + ' - ' + t.attributes()['subacquisition']:['Statisticspost'+t.attributes()['subacquisition'],'electrodes']}) #mettre le nom de la subacquisition
      elif (t.attributes()['modality'] == 't1mri') and ('postOp' in t.attributes()['acquisition']):
         dictionnaire_list_images.update({'IRM post-op':['T1postOp','electrodes']})
      elif (t.attributes()['modality'] == 'resection'):
         dictionnaire_list_images.update({'Resection':['Resection','electrodes']})
      elif (t.attributes()['modality'] == 'freesurfer_atlas'):
          dictionnaire_list_images.update({'FreeSurferAtlas pre':['FreesurferAtlaspre','electrodes']})
      elif (t.attributes()['modality'] == 'hippofreesurfer_atlas'):
          dictionnaire_list_images.update({'HippoFreeSurferAtlas pre':['HippoFreesurferAtlaspre','electrodes']})


      try:
        nameAcq = t.attributes()['acquisition']
        #print "Loading %s as %s"%(t.fileName(), nameAcq)
        #print repr(t.attributes())
        # We try to get the acquisition name without the date (if there is one) : T1pre_2000-01-01 -> T1pre
        if 'Statistics' in nameAcq:
           na = nameAcq.split('_')[0] + t.attributes()['subacquisition']
        else:
           na = nameAcq.split('_')[0]
      except:
        if moda == 'Electrode Implantation Coronal Image':
          na = 'ImplantationCoro'
        elif moda == 'Electrode Implantation Sagittal Image':
          na = 'ImplantationSag'
        else:
          print "CANNOT find a nameAcq for ",repr(t)
          na = 'unknown'

      self.loadAndDisplayObject(t, na)
      if na == 'T1pre':
        # Load standard transformations (AC-PC, T1pre Scanner-based, BrainVisa Talairach)
        try:
          self.refConv.loadACPC(t)
        except Exception, e:
          print "Cannot load AC-PC referential from T1 pre MRI : "+repr(e)
        try:
          self.refConv.loadTalairach(t)
        except Exception, e:
          print "Cannot load Talairach referential from T1 pre MRI : "+repr(e)
        try:
          tr2sb = self.t1pre2ScannerBased()
          if tr2sb is not None:
            self.refConv.setAnatomistTransform("Scanner-based", tr2sb, toRef=True)
            # Add the AC-centered Scanner-Based (for PTS importation using AC-centered Olivier David method
            if self.refConv.isRefAvailable('AC-PC'):
              acInScannerBased = self.refConv.anyRef2AnyRef([0.0,0.0,0.0],'AC-PC', 'Scanner-based')
              inf = tr2sb.getInfos()
              rot = inf['rotation_matrix']
              trans = [inf['translation'][0] - acInScannerBased[0], inf['translation'][1] - acInScannerBased[1], inf['translation'][2] - acInScannerBased[2]]
              m = aims.Motion(rot[:3]+[trans[0]]+rot[3:6]+[trans[1]]+rot[6:]+[trans[2]]+[0,0,0,1])
              self.refConv.setTransformMatrix('AC-centered Scanner-Based', m.inverse(), m)
        except Exception, e:
          print "Cannot load Scanner-based referential from T1 pre MRI : "+repr(e)
        # Get the hemisphere meshes for the acquisition : name = na + filename base : for example, if the acquisition is T1pre_2000-01-01 and the file head.gii, we want T1pre-head
        rdi3 = ReadDiskItem( 'Hemisphere Mesh', 'Anatomist mesh formats', requiredAttributes={'subject':patient, 'acquisition':nameAcq, 'center':self.currentProtocol} )
        hemis = list(rdi3._findValues( {}, None, False ) )


        for hh in hemis:
          #pdb.set_trace()
          self.loadAndDisplayObject(hh, na + '-' + hh.attributes()['side'] + 'Hemi', color=[0.8,0.7,0.4,0.7])
          print "Found hemisphere "+ str(na + '-' + hh.attributes()['side'] + 'Hemi')

        atlas_di = ReadDiskItem('hemisphere marsAtlas parcellation texture', 'aims Texture formats', requiredAttributes={ 'regularized': 'false','subject':patient, 'center':self.currentProtocol, 'acquisition':nameAcq })
        atlas_di_list = list(atlas_di._findValues({}, None, False ))
        #probleme
        wm_di = ReadDiskItem( 'Hemisphere White Mesh', 'aims mesh formats',requiredAttributes={'subject':patient, 'center':self.currentProtocol })


        if len(atlas_di_list) > 0:
          for atl in atlas_di_list:
            wm_side = wm_di.findValue(atl)
            self.loadAndDisplayObject(wm_side, na + '-' + atl.attributes()['side'] + 'MARSATLAS', texture_item = atl, palette = 'MarsAtlas', color=[0.8,0.7,0.4,0.7])
            print "Found hemisphere "+ str(na + '-' + atl.attributes()['side'] + 'MARSATLAS')
            dictionnaire_list_images.update({'IRM pre + MARS ATLAS ' + atl.attributes()['side']:['T1pre','T1pre-'+ atl.attributes()['side'] + 'MARSATLAS','electrodes']})
            #pdb.set_trace()

        # Get head mesh for the acquisition
        #probleme
        rdi3 = ReadDiskItem( 'Head Mesh', 'Anatomist mesh formats', requiredAttributes={'subject':patient, 'acquisition':nameAcq, 'center':self.currentProtocol} )
        head = list(rdi3._findValues( {}, None, False ) )
        if len(head) > 0: # Only if there is one !
          self.loadAndDisplayObject(head[0], na + '-' + 'head', color=[0.0,0.0,0.8,0.3])

    self.windowContent = dictionnaire_list_images;
    self.windowCombo1.clear()
    self.windowCombo1.addItems(sorted(dictionnaire_list_images.keys()))
    self.windowCombo2.clear()
    self.windowCombo2.addItems(sorted(dictionnaire_list_images.keys()))

    self.windowCombo1.setCurrentIndex(max(self.windowCombo1.findText(pre_select_1),0))
    self.windowCombo2.setCurrentIndex(max(self.windowCombo2.findText(pre_select_2),0))

    # Display referential informations
    self.setWindowsReferential()
    self.loadElectrodes(self.brainvisaPatientAttributes)
    self.refreshAvailableDisplayReferentials()
    # Display all
    self.allWindowsUpdate()
    # Disable the button because no cleanup is attempted when loading a patient when one is already loaded -> there may be a mixup
    self.loadPatientButton.setEnabled(False)
    self.patientList.setEnabled(False)



  # Chargement d'un objet (IRM, mesh...) dans Anatomist et mise à jour de l'affichage
  def loadAndDisplayObject(self, diskitem, name = None, color=None, palette=None, texture_item = None):

      if name is None:
        return

      #Already exists ! Remove it.
      if name in self.dispObj:
        self.a.removeObjects([self.dispObj[name],], self.wins) # Remove from windows
        self.a.deleteObjects(self.dispObj[name]) # CURRENT
        del self.dispObj[name]
        del self.diskItems[name]

      print "loading "+repr(diskitem)+ ' as '+name

      obj = self.a.loadObject(diskitem)
      # Read all available transforms in the image header (for example the coregistration by SPM to the T1 Pre MRI)
      #obj.loadReferentialFromHeader()
      if 'ColorPalette' in diskitem.attributes():
          obj.setPalette(palette = diskitem.attributes()['ColorPalette'])
      elif palette is not None and texture_item is None:
          obj.setPalette(palette = palette)

      if texture_item is not None:
         texture = self.a.loadObject(texture_item)
         if palette is not None:
            texture.setPalette(palette = palette)
         #pdb.set_trace()
         textured_mesh = self.a.fusionObjects((obj,texture),method = 'FusionTexSurfMethod')
         #we need to keep the texture and the mesh as well as the fusion of both
         self.objtokeep[name + '_mesh'] = obj
         self.objtokeep[name + '_texture'] = texture
         obj = textured_mesh


      # Store the object
      self.dispObj[name] = obj
      self.diskItems[name] = diskitem
      #print "Referential of object %s : %s"%(repr(name), repr(obj.getReferential().uuid()))
      # If this is a volume, smooth it :
      try:
        self.a.execute('TexturingParams', objects=[obj], filtering='linear')
      except:
        pass
      #pdb.set_trace()
      if color is not None:
	self.a.setMaterial(obj, diffuse=color)
      return obj




  def setWindowsReferential(self, ref=None):
    """ Get all available referentials from anatomist and tries to match identical referentials from SPM"""
    # If the T1pre image is already loaded
    if ref is None:
      if self.preReferential():
        # Assign the T1pre native referential to the windows
        self.currentWindowRef = self.preReferential()
        self.a.assignReferential(self.currentWindowRef, self.wins)

    else:
        self.currentWindowRef = ref
        self.a.assignReferential(ref, self.wins)


  # Get the click events
  def clickHandler(self, eventName, params):
        #pos=params['position']
        #currentwin=params['window']
        coords = [0.0,0.0,0.0]
        if 'T1pre' in self.dispObj:
          pT1Pre = self.a.linkCursorLastClickedPosition(self.dispObj['T1pre'].getReferential()).items()
          #pdb.set_trace()
          if self.coordsDisplayRef == 'Natif':
            coords = pT1Pre
          elif self.coordsDisplayRef == 'Scanner-based':
              infos = self.t1pre2ScannerBased().getInfos()
              rot = infos['rotation_matrix']
              trans = infos['translation']
              m = aims.Motion(rot[:3]+[trans[0]]+rot[3:6]+[trans[1]]+rot[6:]+[trans[2]]+[0,0,0,1])
              coords = m.transform(pT1Pre)
          else:
            try:
              coords = self.refConv.real2AnyRef(pT1Pre, self.coordsDisplayRef)
            except:
              coords = [0.0,0.0,0.0]
        #print type(coords)
        if coords is None:
          coords = [0.0,0.0,0.0]
        self.positionLabel.setText("%.2f, %.2f, %.2f" % tuple(coords))


  def preReferential(self):
    if 'T1pre' in self.dispObj:
      return self.dispObj['T1pre'].getReferential()
    else:
      return None

  def positionPreRef(self):
    return list(self.a.linkCursorLastClickedPosition(self.preReferential()).items())

  def t1pre2ScannerBased(self):
    """ Returns a Transformation object that transforms T1pre referential to T1pre Scanner-Based referential """
    if self.t1pre2ScannerBasedTransform is not None:
       if "dead" not in self.t1pre2ScannerBasedTransform.getInfos():
          return self.t1pre2ScannerBasedTransform
    rdi = ReadDiskItem('Transformation to Scanner Based Referential', 'Transformation matrix', exactType=True,\
      requiredAttributes={'modality':'t1mri', 'subject':self.brainvisaPatientAttributes['subject'], 'center':self.brainvisaPatientAttributes['center']})
    allTransf = list (rdi._findValues( {}, None, False ) )
    for trsf in allTransf:
      if trsf.attributes()['acquisition'].startswith(u'T1pre'):
        print repr(trsf.attributes())
        srcrDiskItem = self.transfoManager.referential( trsf.attributes()['source_referential'] )
        srcr = self.a.createReferential(srcrDiskItem)
        dstrDiskItem = self.transfoManager.referential(trsf.attributes()['destination_referential'])
        self.t1pre2ScannerBasedId = trsf.attributes()['destination_referential']
        dstr = self.a.createReferential(dstrDiskItem)
        self.t1pre2ScannerBasedTransform = self.a.loadTransformation(trsf.fullPath(), srcr, dstr)
        return self.t1pre2ScannerBasedTransform
    return None


  def mniReferentialId(self):
    return aims.StandardReferentials.mniTemplateReferentialID()

  def mniReferential(self):
    return self.a.mniTemplateRef

  def refreshAvailableDisplayReferentials(self):
    curr = str(self.referentialCombo.currentText())
    self.referentialCombo.clear()
    refs = self.refConv.availableReferentials().keys() + ['Natif',]
    print "Available referentials from refConv : "+repr(refs)
    self.referentialCombo.addItems(sorted(refs))
    #if curr in refs:
      #self.set

  def updateCoordsDisplay(self, text):
    self.coordsDisplayRef = str(text)
    self.clickHandler(None, None)

  def updateDispMode(self, index):
	""" Update the display mode of all electrodes """
	mode = 'sphere'
	params = {}
	isbipole = False
	if index == 0:
	  mode = 'real'
	  self.colorConfigButton.setEnabled(False)
	elif index == 1:
	  params = {'diameter':1.0}
	  self.colorConfigButton.setEnabled(False)
	elif index == 2:
	  params = {'diameter':2.0}
	  self.colorConfigButton.setEnabled(False)
	elif index == 3:
	  params = {'diameter':5.0}
	  self.colorConfigButton.setEnabled(False)
	elif index == 4:
          mode = 'off'
          self.colorConfigButton.setEnabled(False)
        elif index == 5:
          mode = 'bipole'
          isbipole = True
          #is there a json file about the seeg results in the database.
          rdi_seeglabel = ReadDiskItem('Electrodes SEEG Labels','Electrode sEEG Label Format',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.brainvisaPatientAttributes['center']})
          di_seeglabel = list(rdi_seeglabel.findValues({},None,False))
          #if not, ask for one
          if len(di_seeglabel) == 0:
             load_new_file = True               
          else:
             #ask if you want to replace the loaded data
             rep = QtGui.QMessageBox.warning(self, u'Use database or Import?', u"Use sEEG stim result from the database ? (if no, it will ask for a file)", QtGui.QMessageBox.Yes | QtGui.QMessageBox.No | QtGui.QMessageBox.Cancel, QtGui.QMessageBox.Cancel)
             if rep == QtGui.QMessageBox.Yes:
                load_new_file = False 
             elif rep == QtGui.QMessageBox.No:    
                load_new_file = True
             elif rep == QtGui.QMessageBox.Cancel:
                return
          
          if load_new_file:
            wdi_seeg = WriteDiskItem('Electrodes SEEG Labels','Electrode sEEG Label Format')
            di_seeg = wdi_seeg.findValue({'subject':self.brainvisaPatientAttributes['subject'], 'center':self.brainvisaPatientAttributes['center']} )
            
            fichierseegLabel =  QtGui.QFileDialog.getOpenFileName(self, "Select a file containing seeg labels: ", "", "(*.xlsx *.csv *.json)")
            if os.path.basename(str(fichierseegLabel)).split('.')[-1] == 'json':
                fin = open(str(fichierseegLabel),'rb')
                new_label = json.loads(fin.read())
                fin.close()
                
                try:
                  os.mkdir(os.path.dirname(str(di_seeg)))
                except:
                  pass
                fout = open(str(di_seeg),'w')
                fout.write(json.dumps({'title':new_label['title'],'contacts':new_label['contacts']}))
                fout.close()                
                
                neuroHierarchy.databases.insertDiskItem(di_seeg, update=True )
                
            elif os.path.basename(str(fichierseegLabel)).split('.')[-1] == 'xlsx':
                
                contact_label_class = generate_contact_colors()
                
                inter_label = contact_label_class.from_excel_files(str(fichierseegLabel))
                
                #write the json and include it in the database
                try:
                  os.mkdir(os.path.dirname(str(di_seeg)))
                except:
                  pass
                fout = open(str(di_seeg),'w')
                fout.write(json.dumps({'title':inter_label[0],'contacts':inter_label[1]}))
                fout.close() 
                
                neuroHierarchy.databases.insertDiskItem(di_seeg, update=True )
                new_label = {'title':inter_label[0],'contacts':inter_label[1]}
          
          else:
               fin = open(str(di_seeglabel[0]),'rb')
               new_label = json.loads(fin.read())
               fin.close()

          bipole_label_sorted = sorted(new_label['contacts'].keys(),key=natural_keys)
          plotsT1preRef = self.getAllPlotsCentersT1preRef()
          info_plotsT1Ref= []
          for k,v in plotsT1preRef.iteritems():
            plot_name_split = k.split('-$&_&$-')
            info_plotsT1Ref.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v.list()))
          #plots_label[k]=(label,label_name)

          plotsT1Ref_sorted = dict(sorted(info_plotsT1Ref, key=lambda plot_number: plot_number[0]))
          
          #remplir self.bipole
          for i_bipole in bipole_label_sorted:
              #get the name of both contact from the bipole name
              try:
                pos_bipol = (numpy.array(plotsT1Ref_sorted[i_bipole.split()[0].title()]) + numpy.array(plotsT1Ref_sorted[i_bipole.split()[2].title()]))/2
              except:
                  print("problem plotsT1Ref")
                  pdb.set_trace()
              entry_bipole = numpy.array(plotsT1Ref_sorted[i_bipole.split()[0].title()])
              #orient_vector_bip = (numpy.array(plotsT1Ref_sorted[i_bipole.split()[0]]) - numpy.array(plotsT1Ref_sorted[i_bipole.split()[2]]))/linalg.norm((numpy.array(plotsT1Ref_sorted[i_bipole.split()[0]]) - numpy.array(plotsT1Ref_sorted[i_bipole.split()[2]])))
              #il faut un orient vector
              self.addBipole(i_bipole,pos_bipol,self.preReferential().uuid(),entry_bipole) #on rajoute pour finir un vecteur


          self.colorConfigButton.setEnabled(True)
          #show bipole instead of contacts ?
          #remove all contacts (mesh and in the list), and set the bipoles instead
          
          contact_labels_seeg = {}
          params.update({'contact_colors':contact_labels_seeg})
          
          self.bipoleLabels = new_label
          self.configureColors()
          
          #self.dispObj['electrodes'][0].__getattr__('name')

	#pdb.set_trace()
	if mode == self.dispMode and params == self.dispParams:
	  print "already using the right display mode"
	  return
	self.updateElectrodeMeshes(clear=True)
	for elec in self.electrodes:
	  elec['elecModel'].setDisplayMode(mode, params)
	  elec['elecModel'].updateDisplay()
	  elec['elecModel'].setDisplayReferential(elec['ref'])
	self.dispMode = mode
	self.dispParams = params
	self.updateElectrodeMeshes(bipole = isbipole)
	# Update the contact list meshes of the current electrode (so they can be selected)
	self.electrodeSelect(self.electrodeList.currentRow())
	

  def updateElectrodeMeshes(self, clear=False, bipole=False):
		if clear:
			print "UPDATE CLEAR -> remove electrodes from windows and delete them"
			#traceback.print_stack(limit=4)
			self.a.removeObjects(self.dispObj['electrodes'], self.wins)
			#self.a.deleteObjects(self.dispObj['electrodes'])# Should be destroyed in electrode.py CURRENT
			self.dispObj['electrodes'] = []
			return
                if not bipole:    
		  self.dispObj['electrodes'] = [mesh for elec in self.electrodes for mesh in elec['elecModel'].getAnatomistObjects() if mesh is not None]
		  self.setElectrodeMeshesNames()		  
		elif bipole:
                  self.dispObj['electrodes'] = [mesh for elec in self.bipoles for mesh in elec['elecModel'].getAnatomistObjects() if mesh is not None]
                  self.setBipoleMeshesNames()
		
		#self.dispObj['electrodes'][0].__getattr__('name')
		self.allWindowsUpdate()

  # Add an electrode from a template
  def addElectrode(self, name=None, model=None, target=[0,0,0], entry=[0,0,-1], refId = None):
    if name is None:
      name = str(self.nameEdit.text())
    if model is None:
      model = str(self.typeComboBox.currentText())
    if refId is not None:
       if self.preReferential().uuid() != refId:
         print "ERROR : the electrode is not defined in reference to the T1 pre image (%s) but in reference to %s !\nGoing on anyway..."%(self.preReferential().uuid(), refId)
    for el in self.electrodes:
      if name == el['name']:
        name = self.findFreeElectrodeName()
    (newRef, transf, elecModel) = createElectrode(target, entry, self.preReferential(), self.a,\
             model = self.elecModelListByName[str(model)].fullPath(), dispMode = self.dispMode, dispParams = self.dispParams)
    self.electrodes.append({'name': name, 'ref':newRef, 'transf':transf, 'elecModel':elecModel,\
                                        	'target':target, 'entry':entry, 'model':model})
    self.electrodeList.addItem(name)
    #index = self.elecCombo.findText("Electrode "+str(self.elecname))
    self.electrodeList.setCurrentRow(self.electrodeList.count() - 1)
    self.addElectrodeLabel(name, [0,0,-10], newRef, len(self.electrodes) - 1)
    self.updateElectrodeMeshes()

  def addBipole(self, name=None, positionbip=[0,0,0], refId = None,entry_bipole = None):
    if name is None:
      print("error, bipole must have a name")
      return
      pdb.set_trace()
    if refId is None:
      print("error, bipole has to be assigned to a referential")
      return
      pdb.set_trace()
    if refId is not None:
       if self.preReferential().uuid() != refId:
         print "ERROR : the electrode is not defined in reference to the T1 pre image (%s) but in reference to %s !\nGoing on anyway..."%(self.preReferential().uuid(), refId)     
    #if refId is not None:
       #if self.preReferential().uuid() != refId:
         #print "ERROR : the electrode is not defined in reference to the T1 pre image (%s) but in reference to %s !\nGoing on anyway..."%(self.preReferential().uuid(), refId)
    #check is name is already taken
    if len(self.bipoles)>0:
      for bip in self.bipoles:
        if name == bip['name']:
          print("error, bipole name already taken")
         
    #je met un objet elecmodel ici ou juste le mesh ? un objet elecmodel de 1 contact ?
    rdiEM = ReadDiskItem('Electrode Model', 'Electrode Model format', requiredAttributes={'center':self.currentProtocol})
    listEM = list(rdiEM.findValues({},None,False))
    matches = filter((lambda x: u"bipole" in str(x)), listEM)
        
    (newRef, transf, elecModel) = createBipole(positionbip.tolist(), entry_bipole.tolist(), self.preReferential(), self.a, model = matches[0], dispMode = 'bipole', dispParams = None)
    self.bipoles.append({'name': name, 'ref':newRef, 'transf':transf, 'target':positionbip.tolist(), 'entry': entry_bipole.tolist(), 'elecModel':elecModel}) #to change target and entry
    #on ne touche pas à electrodeList
    #self.electrodeList.addItem(name)
    #index = self.elecCombo.findText("Electrode "+str(self.elecname))
    #self.electrodeList.setCurrentRow(self.electrodeList.count() - 1)
    
    self.addElectrodeLabel(name, [0,0,-10], newRef, len(self.bipoles) - 1, True)
    #self.updateElectrodeMeshes(bipole=True)


  # Setting names on meshes to get a nice tooltip for each mesh
  def setElectrodeMeshesNames(self, electrode = None):
    if electrode is None:
      electrodes = self.electrodes
    else:
      electrodes = [electrode,]
    for el in electrodes:
      for name, element in el['elecModel'].getDisplayed().iteritems():
        if element['mesh'] is not None:
          if element['type'] == 'Plot':
            element['mesh'].setName(name.replace('Plot', el['name']))
          else:
            element['mesh'].setName(el['name'])
          #element['mesh'].setChanged()
          #element['mesh'].notifyObservers() #Should be needed to refresh display, but it works without it
          
  def setBipoleMeshesNames(self, bipole = None):
      if bipole is None:
          bipoles = self.bipoles
      else:
          bipoles = [bipole,]
      for bp in bipoles:
          for name, element in bp['elecModel'].getDisplayed().iteritems():
            element['mesh'].setName(bp['name'])  

  def findFreeElectrodeName(self):
    if len(self.electrodes) == 0:
      return 'A'
    n = [el['name'] for el in self.electrodes]
    newName = n[-1]
    firstletter = newName[::-1][-1]
    if firstletter in string.ascii_uppercase:
			newName = string.uppercase[(string.uppercase.find(firstletter)+1)%len(string.uppercase)] + newName[1:]
    if newName in n:
			while newName in n:
				newName = newName+'_'
    return newName

  def currentElectrode(self):
    idx = self.electrodeList.currentRow()
    print "Electrode %i is selected"%idx
    if idx < 0:
      return None
    return self.electrodes[idx]

  def removeElectrode(self):
    """Remove an electrode (and all contacts)"""
    elec = self.currentElectrode()
    idx = self.electrodes.index(elec)
    if elec is None:
      return
    # Remove meshes from Anatomist
    self.updateElectrodeMeshes(clear=True)
    elec['elecModel'].clearDisplay()
    item = self.electrodeList.takeItem(idx)
    del item
    del self.electrodes[idx]
    self.updateElectrodeMeshes()

  def updateElectrodeModel(self, model):
    elec = self.currentElectrode()
    if elec is None:
      return
    if str(elec['model']) == str(model):
      return
    #import pdb; pdb.set_trace()
    self.updateElectrodeMeshes(clear=True)
    elec['elecModel'].clearDisplay()
    del elec['elecModel']
    (newRef, transf, elecModel) = createElectrode(elec['target'], elec['entry'], self.preReferential(), self.a,\
             model = self.elecModelListByName[str(model)].fullPath(), dispMode = self.dispMode, dispParams = self.dispParams)
    elec['elecModel'] = elecModel
    elec['model']=str(model)
    self.electrodeSelect(self.electrodeList.currentRow())
    self.updateElectrodeMeshes()


  def updateEntry(self, e=None):
    """ Updates the current electrode entry point from the cursor position"""
    el = self.currentElectrode()
    pos = self.positionPreRef()
    # Just ignore if new entry is identical to target
    if pos == el['target']:
      return

    (newRef, transf) = moveElectrode(el['target'], pos, self.preReferential(),\
                                     el['ref'], self.a, el['elecModel'].getAnatomistObjects())
    el['entry'] = pos
    el['transf'] = transf

  def updateTarget(self, t=None):
    """ Updates the current electrode target point from the cursor position"""
    el = self.currentElectrode()
    if el is None:
      return
    pos = self.positionPreRef()
    # Just ignore if new target is identical to entry
    if pos == el['entry']:
      return

    (newRef, transf) = moveElectrode(pos, el['entry'], self.preReferential(),\
                                     el['ref'], self.a, el['elecModel'].getAnatomistObjects())
    el['target'] = pos
    el['transf'] = transf
    print repr(self.a.getReferentials())


  def editElectrodeName(self):
    """Update the electrode name of the selected contact"""
    name = str(self.nameEdit.text())
    idx = self.electrodeList.currentRow()
    sameNameItems = self.electrodeList.findItems(name, QtCore.Qt.MatchFixedString)
    if len(sameNameItems) != 0:
			if sameNameItems[0] == self.electrodeList.item(idx): # Name was not changed
				return
			else:
				QtGui.QMessageBox.warning(self, u'Error', u"The name %s is already used by another electrode. Choose an other one !"%name)
				self.nameEdit.setText(self.electrodeList.item(idx).text())
    self.electrodes[idx]['name'] = name
    self.setElectrodeMeshesNames(self.electrodes[idx])
    self.electrodeList.currentItem().setText(name)
    self.electrodeSelect(idx) # updates the contacts' display


  def electrodeSelect(self, idx):
    """Electrode/contact selection changed"""
    el = self.electrodes[idx]
    self.nameEdit.setText(el['name'])
    self.typeComboBox.setCurrentIndex(self.typeComboBox.findText(el['model']))
    #import pdb;pdb.set_trace()
    # Select the contacts in anatomist
    g = self.a.getDefaultWindowsGroup()
    g.setSelection(el['elecModel'].plotMeshes())
    self.currentContacts = dict((name.replace('Plot', el['name']), element['mesh']) for name, element in el['elecModel'].getDisplayed().iteritems() if element['type'] == 'Plot' and element['mesh'] is not None)
    self.currentElectrodes = [el,]
    #Update contact list
    self.contactList.clear()
    self.contactList.addItems(sorted(self.currentContacts.keys(), key=natural_keys))
    self.updateElectrodeView()

  def electrodeGo(self, idx = None, electrode = None):
      if idx is not None:
          if type(idx) == int:
            electrode = self.electrodes[idx]
          else:
              electrode = self.currentElectrodes[0]
      elif electrode is None:
          return
      self.contactGo(el = electrode)

  def contactSelect(self, item):
    # Select the contacts in anatomist
    print "Click on the contact " + str(item.text())
    g = self.a.getDefaultWindowsGroup()
    #print "CurrentContacts = "+repr(self.currentContacts.keys())
    g.setSelection([self.currentContacts[str(item.text())],])

  def contactGo(self, item = None, el=None):
    # Put the cursor on the contact in anatomist
    #print "Double clic sur le contact " + str(item.text())
    try:
      if el is None:
        el = self.currentElectrodes[0]
      if item is not None:
        name = str(item.text()).replace(el['name'], 'Plot')
      else:
          name = 'Plot1'
      xyz = getPlotsCenters(el['elecModel'])[name]
      # setLinkedCursor uses window referential : must apply transform before setting the position
      if self.currentWindowRef == self.preReferential():
          xyz = el['transf'].transform(xyz)
          self.wins[0].moveLinkedCursor(xyz)
          #self.a.execute('LinkedCursor', window=self.wins[0], position=xyz)
      elif  self.currentWindowRef == el['ref']:
          self.wins[0].moveLinkedCursor(xyz)
      else:
          print "Warning : Current window referential is not T1pre referential ! Cannot set the linkedCursor on the selected electrode contact..."
    except:
      print "Error moving the cursor to the contact"


  def addElectrodeLabel(self, label, position, ref, elecId,bipole=False):
    props = Props()
    props.usespheres = False
    props.center = [0, 0, 0]
    gc = aims.Point3df(position)
    if not bipole:
      self.electrodes[elecId]['labelObjects'] = makelabel( self.a, label, gc, gc+aims.Point3df([5,0,0]), ref, (0, 0, 0, 0), props )
      self.electrodes[elecId]['props'] = props
    elif bipole:
      self.bipoles[elecId]['labelObjects'] = makelabel( self.a, label, gc, gc+aims.Point3df([5,0,0]), ref, (0, 0, 0, 0), props )
      self.bipoles[elecId]['props'] = props      

  def getElectrodeTemplates(self):
    """Returns a list of Electrode objects, one for each available electrode template model available in the DB"""
    if self.electrodeTemplateStubs == []:
      self.electrodeTemplateStubs = dict([(n,ElectrodeModel(modelPath=model.fullPath(), dispMode='off')) for n, model in self.elecModelListByName.iteritems()])
    return self.electrodeTemplateStubs

  def fitElectrodeModelToLength(self, target, entry):
    """ Tries to find a match between the length of the electrode and a model, but prefering uniform electrodes (no variable spacing between contacts"""
    length = linalg.norm(array(entry)-array(target))

    models = self.getElectrodeTemplates()
    uniformModelsLength = {}
    for n,model in models.iteritems():
      interPlotsM=[]
      plotsM = sorted([[int(''.join(e for e in namePlot if e.isdigit())), ] + content['center'] for namePlot, content in model.getPlots().iteritems()], key=lambda p:p[0])

      for p in range(len(plotsM)-1):
        interPlotsM.append(math.sqrt((plotsM[p][1]-plotsM[p+1][1])**2 + (plotsM[p][2]-plotsM[p+1][2])**2 + (plotsM[p][3]-plotsM[p+1][3])**2))
      lengthM = sum(interPlotsM)
      if len(set(interPlotsM)) == 1: # All intervals are identical
        uniformModelsLength[n] = lengthM
    if length >= max(uniformModels.values()):
      return [m for m,l in uniformModels.iteritems() if l == max(uniformModels.values())][0]
    else:
      largerModels = dict([(m,l) for m,l in uniformModels.iteritems() if l >= length])
      return [m for m,l in largerModels.iteritems() if l == min(largerModels.values())][0]


  def fitElectrodeModelToPlots(self, plots):
    """ Tries to find a match between a list of plots [[numPlot, x, y, z],[...], ...] and available electrode templates.
      Return None if it fail, [ModelName, [targetX, targetY, targetZ], [entryX, entryY, entryZ]] if it works
    """

    plots = sorted(plots, key=lambda p:int(p[0]))
    # Compute inter-contact distances
    interPlots=[]
    for p in range(len(plots)-1):
      interPlots.append(math.sqrt((plots[p][1]-plots[p+1][1])**2 + (plots[p][2]-plots[p+1][2])**2 + (plots[p][3]-plots[p+1][3])**2))
    # Find electrode models with a similar number of plots
    models = dict([(n,tpl) for n,tpl in self.getElectrodeTemplates().iteritems() if tpl.countPlots() == len(plots)]) # identical number
    if len(models) == 0:
      print "Cannot find a template with the right number of contacts (%i), trying longer ones"%len(plots)
      nbPlotsModels = dict([(n, tpl.countPlots()) for n,tpl in self.getElectrodeTemplates().iteritems() if tpl.countPlots() > len(plots)])
      if len(nbPlotsModels) == 0:
        print "Cannot find a template with enough contacts for this electrode ! (available : %s)\nWill match with the largest available electrodes !\n THIS WILL LOSE SOME PLOTS"%repr(nbPlotsModels.values())
        nbPlotsModels = dict([(n, tpl.countPlots()) for n,tpl in self.getElectrodeTemplates().iteritems()])
        models = dict([(n,tpl) for n,tpl in self.getElectrodeTemplates().iteritems() if tpl.countPlots() == max(nbPlotsModels.values())])
      else:
        models = dict([(n,tpl) for n,tpl in self.getElectrodeTemplates().iteritems() if tpl.countPlots() == min(nbPlotsModels.values())])

    # Now, check the interPlots distance in the model and compare with the template
    distanceFromModel = {}
    for n,model in models.iteritems():
      interPlotsM=[]
      plotsM = sorted([[int(''.join(e for e in namePlot if e.isdigit())), ] + content['center'] for namePlot, content in model.getPlots().iteritems()], key=lambda p:p[0])

      for p in range(len(plotsM)-1):
        interPlotsM.append(math.sqrt((plotsM[p][1]-plotsM[p+1][1])**2 + (plotsM[p][2]-plotsM[p+1][2])**2 + (plotsM[p][3]-plotsM[p+1][3])**2))
      distanceFromModel[n] = sum([(interPlotsM[i]-interPlots[i])**2 for i in range(min([len(interPlotsM), len(interPlots)]))])

    # Choose the model with the smallest distance, or the first one if a few are equally good, reject if sum is more than 1mm per contact
    minDist = min(distanceFromModel.values())
    if minDist > 1.0*len(plots):
      print "Cannot match a template : minimum distance found is %f mm, should be %f or less !"%(minDist, 1.0*len(plots))
      return None
    # We have a winner !
    goodModel = [m for m in distanceFromModel if distanceFromModel[m] == minDist][0]
    print "Found model %s, match distance = %f"%(goodModel, minDist)
    #import pdb; pdb.set_trace()
    entry = plots[-1][1:]
    # Target is not the center of the first contact, but its end ! -> coords - lengthOfPlot/2*vector(plot0->plotLast)
    target = array(plots[0][1:]) - (array(plotsarr[-1][1:])-array(plots[0][1:]))/linalg.norm(array(plots[-1][1:])-array(plots[0][1:])) * 0.5 * self.getElectrodeTemplates()[goodModel].getCylinder('Plot1')['length']
    return [goodModel, target.tolist(), entry]


  def comboMessageBox(self, text, choices):
    """ Displays a message box with a choice (combobox), Ok and Cancel button
        Returns the selected value or None if it was cancelled"""
    if choices is None or choices == []:
      return None
    msgBox = QtGui.QMessageBox()
    msgBox.setText(text)
    combo = QtGui.QComboBox()
    combo.addItems(choices)
    msgBox.layout().addWidget(combo,1,0)
    msgBox.addButton(QtGui.QPushButton('Ok'), QtGui.QMessageBox.AcceptRole)
    msgBox.addButton(QtGui.QPushButton('Annuler'), QtGui.QMessageBox.RejectRole)
    ret = msgBox.exec_();
    if ret == QtGui.QMessageBox.Cancel:
      return None
    return str(combo.currentText())

  def loadPTSbasic(self, path):
    """Load a PTS file, creating an 'electrode model' for each contact (VERY slow with > 50 contacts) """
    refId = self.preReferential().uuid()
    els = []
    elecs = {}
    try:
      f = open(path, 'r')
      lines = f.readlines()
      f.close()
      lines.reverse()
      if not lines.pop().startswith('ptsfile'):
        print 'This is not a valid PTS file !'
        return (refId, [])
      lines.pop() # Useless line 1 1 1
      nb = int(lines.pop()) # Number of contacts
      for i in range(nb):
        l = lines.pop().rstrip().split("\t") # name x y z 0 0 0 2 2.0 (last ones may be contact length/diameter ?)
        name = ''.join(e for e in l[0] if not e.isdigit())
        if name not in elecs:
          elecs[name]=[]
        plot = int(''.join(e for e in l[0] if e.isdigit()))
        coords = list(self.refConv.anyRef2Real([float(l[1]), float(l[2]),float(l[3])], 'Scanner-based'))
        nameplot = l[0]
        elecs[name].append ( [plot,] + coords)
        els.append( {'name':nameplot, 'model':'plot2mmD1mmCentered', 'target': coords, 'entry':coords[0:2]+[coords[2]+1.0,]} )
    except:
      print "Error reading PTS file %s"%path
    return (refId, els)

  def loadPTS(self, path):
    """Load a PTS file (tries to find a suitable electrode model in the available templates)  """
    refId = self.preReferential().uuid()
    els = []
    elecs = {}
    #try:
    f = open(path, 'r')
    lines = f.readlines()
    f.close()

    # The coordinates in the PTS are expressed in which referential ?
    refOfPts = self.comboMessageBox(u'Importation d\'un fichier PTS. Choisissez le référentiel utilisé (Scanner-based...)', sorted(self.refConv.availableReferentials().keys()))
    if refOfPts is None:
      print "User cancelled PTS importation, or no valid referential found"
      return
    print "PTS import referential : %s"%repr(refOfPts)
    lines.reverse()
    if not lines.pop().startswith('ptsfile'):
      print 'This is not a valid PTS file !'
      return (refId, [])
    lines.pop() # Useless line 1 1 1
    nb = int(lines.pop()) # Number of contacts
    for i in range(nb): # All lines (= all plots)
      l = lines.pop().rstrip().split() # name x y z 0 0 0 2 2.0 (last ones may be contact length/diameter ?)
      name = ''.join(e for e in l[0] if not e.isdigit())

      print l[0]
      plot = ''.join(e for e in l[0] if e.isdigit())
      if plot == '':
        continue
      plot = int(plot)

      if name not in elecs:
        elecs[name]=[]

      coords = list(self.refConv.anyRef2Real([float(l[1]), float(l[2]),float(l[3])], refOfPts))
      nameplot = l[0]
      elecs[name].append ( [plot,] + coords)
      #els.append( {'name':nameplot, 'model':'plot2mmD1mmCentered', 'target': coords, 'entry':coords[0:2]+[coords[2]+1.0,]} )
    # Iterate over all electrodes
    for k,l in elecs.iteritems():
      # Try to get a real model from l [[numPlot, x,y,z], [...], ...]
      res = self.fitElectrodeModelToPlots(l)
      if res is not None:
        els.append( {'name':k.lower(), 'model':res[0], 'target': res[1], 'entry':res[2]} )
      else:
        print "Could not find model matching electrode %s, adding individual contacts."%k
        for pl in l:
          coords = pl[1:]
          els.append( {'name':k+str(pl[0]), 'model':'plot2mmD1mmCentered', 'target': coords, 'entry':coords[0:2]+[coords[2]+1.0,]} )

    #except:
    #  print "Error reading PTS file %s"%path
    return (refId, els)

  def loadElectrodeTXT(self, path):
    """Load an electrode.txt file with triplets of lines : nameOfElectrode\n X1 Y1 Z1\n X2 Y2 Z2 such as used by ImaGIN"""
    els=[]
    f = open(path, 'r')
    lines = f.readlines()
    f.close()
    lines.reverse()

    # The coordinates in the TXT are expressed in which referential ?
    refOfTxt = self.comboMessageBox(u'Importation d\'un fichier electrode TXT. Choisissez le référentiel utilisé (Scanner-based a priori)', sorted(self.refConv.availableReferentials().keys()))
    if refOfTxt is None:
      print "User cancelled electrode TXT importation, or no valid referential found"
      return
    print "TXT electrode import referential : %s"%repr(refOfTxt)

    while (len(lines) >= 3):
      try:
        name = lines.pop()
        targ = map (float,lines.pop().replace(',','.').split())
        entr = map (float,lines.pop().replace(',','.').split())
        if len(targ) == 3 and len(entr) == 3 and len(name)>0:
          targ = list(self.refConv.anyRef2Real(targ, refOfTxt))
          entr = list(self.refConv.anyRef2Real(entr, refOfTxt))
          els.append( {'name':name.lower(), 'model':fitElectrodeModelToLength(targ, entr), 'target': targ, 'entry':entr} )
        else:
          print "Invalid lines in electrode txt file : %s, %s, %s"%(repr(name), repr(targ), repr(entr))
      except:
        pass

    refId = self.preReferential().uuid()
    return (refId, els)


  def loadElectrodes(self, patient=None):
    """Load electrode implantation (if already entered) from BrainVisa or from a file"""
    path = None
    #import pdb;pdb.set_trace()
    if patient is not None:
      rdi = ReadDiskItem( 'Electrode implantation', 'Electrode Implantation format', requiredAttributes={'subject':patient['subject'], 'center':patient['center']} )
      elecs = list(rdi._findValues( {}, None, False ) )
      if len(elecs) == 1:
	path = elecs[0].fileName()
      elif len(elecs) > 1:
	print "CAREFUL : more than one electrode implantation are available, strange -> load the first found one" # TODO Dialogue de choix
	path = elecs[0].fileName()
      else: # No implantation available
        print "no electrode implantation found"
        return
    if not path:
      path = str(QtGui.QFileDialog.getOpenFileName(self, "Open electrode implantation", "", "All implantations(*.elecimplant *.pts *.txt);;Electrode implantation (*.elecimplant);;PTS file (*.pts);;Electrode.txt (*.txt)"))

    if not path:
      return
    # Check if we have a PTS/TXT/elecimplant file
    extension = os.path.splitext(path)[1].lower()
    els = []
    refId = None
    if extension == '.elecimplant':
      filein = open(path, 'rb')
      try:
        dic = json.loads(filein.read())
      except:
        filein.close()
        filein = open(path, 'rb')
        dic = pickle.load(filein)

      filein.close()
      els = dic['electrodes']
      refId = dic['ReferentialUuid']

    elif extension == '.txt':
      (refId, els) = self.loadElectrodeTXT(path) # Verifier que c'est un electrode.txt, si c'est un electrode_Pos ou electrode_Name on peut zapper ?
    elif extension == '.pts':
      (refId, els) = self.loadPTS(path)
      # Demander le référentiel, ou deviner si le nom contient _MNI ?
    else:
      print "file format unknown : %s !"%extension
      QtGui.QMessageBox.warning(self, u'Erreur', u"the file format has not been recognized : %s"%extension)

    if refId != self.preReferential().uuid():
      print "CAREFUL: electrodes load are defined in an other referential that the one of the T1 pre, problem possible !"
    #indices = [3,7,11, 0,1,2, 4,5,6, 8,9,10]
    #if dic['2mni'] is not None and 'T1pre' not in self.transf2Mni:
      #trsf = dic['2mni']
      #self.transf2Mni['T1pre'] = self.a.createTransformation([trsf[i] for i in indices], self.preReferential(), self.mniReferential())
    for e in els:
      self.addElectrode(e['name'], e['model'], e['target'], e['entry'], refId)


  def saveElectrodes(self):
    """Save electrode implantation in BrainVisa Database"""
    # Saving : electrode model, name, target and entry point
    els = [dict((k,el[k]) for k in ['name', 'target', 'entry', 'model']) for el in self.electrodes]
    # Save Referential UID which is the base of electrode coordinates
    refId = self.preReferential().uuid()
    print 'Sauvegarde electrodes'+repr(els)+"\nReferential UUID : "+refId
    path = None
    di = None
    if self.brainvisaPatientAttributes is not None:
      wdi = WriteDiskItem( 'Electrode implantation', 'Electrode Implantation format' )
      di = wdi.findValue({'subject':self.brainvisaPatientAttributes['subject'], 'center':self.brainvisaPatientAttributes['center']} )
      if not di:
	print "Cannot find a valid path to store the electrode implantation for the current patient !"
	QtGui.QMessageBox.warning(self, u'Erreur', u"Cannot find a valid path to store the electrode implantation for the current patient !")
      else:
	path = di.fileName()
	createItemDirs(di)
    if not path:
      path = str(QtGui.QFileDialog.getSaveFileName(self, "Save electrode implantation", "", "Electrode implantation (*.elecimplant)"))
    if not path:
      return
    # If there is no extension, add the standard one !

    if os.path.splitext(path)[1] == '':
      path = path+'.elecimplant'

    plotsT1preRef = self.getAllPlotsCentersT1preRef()
    info_plotsT1Ref= []
    for k,v in plotsT1preRef.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plotsT1Ref.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v.list()))
          #plots_label[k]=(label,label_name)

    plotsT1Ref_sorted = sorted(info_plotsT1Ref, key=lambda plot_number: plot_number[0])
    #previous_data.update({'plotsMNI':info_plotMNI})
    
    fileout = open(path, 'wb')
    fileout.write(json.dumps({'electrodes':els, 'ReferentialUuid':refId, '2mni':None, 'timestamp':time.time(),'plotsT1Nat':plotsT1Ref_sorted}))
    #pickle.dump({'electrodes':els, 'ReferentialUuid':refId, '2mni':None, 'timestamp':time.time()}, fileout)
    fileout.close()
    if di is not None:
      neuroHierarchy.databases.insertDiskItem( di, update=True )
    QtGui.QMessageBox.information(self, u'Implantation saved', u"Implantation has been saved in database. Careful, it deleted MNI information if there was any. you'll have to do the normalization again")

  def convertT1ScannerBasedToMni(self, points):
    """Converts an array of points [x,y,z] in scanner-based coords to MNI coords if deformation field is available"""
    field = self.getT1preMniTransform()
    if field is None:
      print "MNI deformation field not found"
      return None
    #pdb.set_trace()
    tmpOutput = getTmpFilePath('csv')
    arr = numpy.asarray(points)#([ [1,2,3], [4,5,6], [7,8,9] ])
    numpy.savetxt(tmpOutput, arr, delimiter=",")
    matlabRun(spm_normalizePoints % ("'"+self.spmpath+"'",field, tmpOutput, tmpOutput))
    out = numpy.loadtxt(tmpOutput, delimiter=",")
    #pdb.set_trace()
    os.remove(tmpOutput)

    if numpy.array_equal(out, arr):
      print "Points to MNI : Error, result read is identical to input"
      return None
    if out.shape != arr.shape:
      print "Points to MNI : Error, result (%s) has not the same number of elements as input (%s)"%(repr(out),repr(arr))
      return None
    return out.tolist()

  # Non blocking wrapper of getAllPlotsCentersMNIRef
  def getAllPlotsCentersMNIRefNB(self, plots, toKeep=None):
		""" Computes the MNI coordinates of plots from t1 pre scanner-based coordinaotes in a thread (non-blocking)
		    then call finishAllPlotsCentersMNIRef with the toKeep elements : these will determine how the data is saved

		"""
		pe = PythonExecutor(lambda myself=self, thePlots=plots: myself.getAllPlotsCentersMNIRef(thePlots), toKeep=toKeep)
		self.threads.append(pe)
		# Remove it from the list of threads when finished
		pe.finished.connect(lambda th=pe:self.finishAllPlotsCentersMNIRef(th))
		pe.start()


  def finishAllPlotsCentersMNIRef(self, thread):
    """ Callback function when MNI coordinates of plots are computed. Will save to basepath
        or to provided diskItems (in 'kept' objects of the thread - which must be a PythonExecutor)
     """
    plotsMNI = thread.output()
    kept = thread.kept()
    try:
      self.threads.remove(thread)
    except:
      pass
    if plotsMNI is None:
        QtGui.QMessageBox.information(self, u'Implantation sauvegardée', u"Implantation has been exported in PTS/TXT file formats in database\nMNI normalization failed !.")
        return

    if 'basepath' in kept:# custom path
      self.saveTXT(path=kept['basepath']+'_MNI.txt', referential='MNI')
      self.savePTS(path=kept['basepath']+'_MNI.pts', referential='MNI')
    elif all (k in kept for k in ('wdiptsmni','timestamp','wditxtmnipos', 'wditxtmniname')): # find and save in DB
      try:
        self.savePTS(path = kept['wdiptsmni'].fullPath(), contacts = plotsMNI)
        kept['wdiptsmni'].setMinf('referential', self.mniReferentialId())
        kept['wdiptsmni'].setMinf('timestamp', kept['timestamp'])
        neuroHierarchy.databases.insertDiskItem(kept['wdiptsmni'], update=True )
        self.saveTXT(pathPos=kept['wditxtmnipos'].fullPath(), pathName=kept['wditxtmniname'].fullPath(), contacts = plotsMNI)
        neuroHierarchy.databases.insertDiskItem(kept['wditxtmniname'], update=True )
        kept['wditxtmnipos'].setMinf('referential', self.mniReferentialId())
        kept['wditxtmnipos'].setMinf('timestamp', kept['timestamp'])
        neuroHierarchy.databases.insertDiskItem(kept['wditxtmnipos'], update=True )
      except:
        QtGui.QMessageBox.information(self, u'Implantation sauvegardée', u"Implantation has been exported in PTS/TXT file formats in database\nThe saving of the MNI normalization failed.")
        return
    else:
        print "Erreur lors de l'exportation PTS/TXT dans le référentiel MNI"
        QtGui.QMessageBox.information(self, u'Implantation sauvegardée', u"Implantation has been exported in PTS/TXT file formats in database\nMNI normalization failed.")
        return

    QtGui.QMessageBox.information(self, u'Implantation sauvegardée', u"Implantation has been exported in PTS/TXT file formats in database\nMNI normalization has been applied")

    #add to the elecimplant file
    rdi = ReadDiskItem( 'Electrode implantation', 'Electrode Implantation format' ,requiredAttributes={'center':self.brainvisaPatientAttributes['center'],'subject':self.brainvisaPatientAttributes['subject']})
    di=rdi.findValues({},None,False)
    ldi = list(rdi.findValues({},None,False))

    if len(ldi) >0:
       if (os.path.exists(str(ldi[0]))):
	        filein = open(str(ldi[0]), 'rb')
	        try:
	          previous_data = json.loads(filein.read())
	        except:
			  filein.close()
			  filein = open(str(ldi[0]), 'rb')
			  previous_data = pickle.load(filein)

	        filein.close()

    info_plotMNI= []
    for k,v in plotsMNI.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plotMNI.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          #plots_label[k]=(label,label_name)

    plotMNI_sorted = sorted(info_plotMNI, key=lambda plot_number: plot_number[0])
    previous_data.update({'plotsMNI':info_plotMNI})

    #resave as json file
    fout = open(str(ldi[0]),'w')
    fout.write(json.dumps(previous_data))
    fout.close()
    #pdb.set_trace()
    neuroHierarchy.databases.insertDiskItem([x for x in di][0], update=True )
    print ".elecimplant done with MNI"

    print("start generate dictionaries")
    self.parcelsExportElectrodes(Callback=lambda:[self.marsatlasExportResection(),self.exportCSVdictionaries(),self.generateMappingContactCortex(),self.generateTemplateBipoleStimulationFile(),self.screenshot(),self.makeMP4(),self.calculParcel()]) 

  # Normalize the coordinates and export them in PTS format and TXT format
  def normalizeExportElectrodes(self, saveInDB=True):
    timestamp = time.time()
    plots = self.getAllPlotsCentersT1preScannerBasedRef()
    if not saveInDB: # Choose the output file name
      basepath = str(QtGui.QFileDialog.getSaveFileName(self, 'Choose a file name for exported files', "", "Electrode implantation TXT files (*.txt)"))
      if not basepath:
        return
      basepath = os.path.splitext(basepath)[0]
      self.saveTXT(path=basepath+'.txt', contacts=plots)
      self.savePTS(path=basepath+'.pts', contacts=plots)
      plotsMNI = self.getAllPlotsCentersMNIRefNB(plots, toKeep={'basepath':basepath})
      return
    else: # Find the right places in the Database
      # T1-pre scanner-based referential
      wdipts = WriteDiskItem('Electrode Implantation PTS', 'PTS format', requiredAttributes={'no_ref_name':'True'}).findValue(self.diskItems['T1pre'])#,_debug= sys.stdout requiredAttributes={'ref_name':'default'}
      wditxtpos = WriteDiskItem('Electrode Implantation Position TXT', 'Text file', requiredAttributes={'no_ref_name':'True'}).findValue(self.diskItems['T1pre'])
      wditxtname = WriteDiskItem('Electrode Implantation Name TXT', 'Text file', requiredAttributes={'no_ref_name':'True'}).findValue(self.diskItems['T1pre'])
      if wdipts is None or wditxtpos is None or wditxtname is None:
        print "Cannot find a path to save PTS or TXT files in the database !"
        return
      if plots is None:
        print "Cannot get the scanner-based plots coords"
        return
      self.savePTS(path = wdipts.fullPath(), contacts=plots)
      wdipts.setMinf('referential', self.t1pre2ScannerBasedId)
      wdipts.setMinf('timestamp', timestamp)
      neuroHierarchy.databases.insertDiskItem(wdipts, update=True )
      self.saveTXT(pathPos=wditxtpos.fullPath(), pathName=wditxtname.fullPath(), contacts=plots)
      neuroHierarchy.databases.insertDiskItem(wditxtname, update=True )
      wditxtpos.setMinf('referential', self.t1pre2ScannerBasedId)
      wditxtpos.setMinf('timestamp', timestamp)
      neuroHierarchy.databases.insertDiskItem(wditxtpos, update=True )

      # MNI referential
      if self.getT1preMniTransform() is None:
        QtGui.QMessageBox.information(self, u'Implantation sauvegardée', u"Implantation has been exported in PTS/TXT file formats in database\nMNI normlization is not available (not found in database).")
        return
      wdiptsmni = WriteDiskItem('Electrode Implantation PTS', 'PTS format', requiredAttributes={'ref_name':'MNI'}).findValue(self.diskItems['T1pre'])
      wditxtmnipos = WriteDiskItem('Electrode Implantation Position TXT', 'Text file', requiredAttributes={'ref_name':'MNI'}).findValue(self.diskItems['T1pre'])
      wditxtmniname = WriteDiskItem('Electrode Implantation Name TXT', 'Text file', requiredAttributes={'ref_name':'MNI'}).findValue(self.diskItems['T1pre'])
      if wdiptsmni is None or wditxtmnipos is None or wditxtmniname is None:
        print "Cannot find a path to save MNI PTS or TXT files in the database !"
      else:
        plotsMNI = self.getAllPlotsCentersMNIRefNB(plots, toKeep={'wdiptsmni':wdiptsmni, 'timestamp':timestamp, 'wditxtmnipos':wditxtmnipos, 'wditxtmniname':wditxtmniname})

      # AC-PC referential
      if self.refConv.isRefAvailable('AC-PC'):
        wdiptsacpc = WriteDiskItem('Electrode Implantation PTS', 'PTS format', requiredAttributes={'ref_name':'ACPC'}).findValue(self.diskItems['T1pre'])
        if wdiptsacpc is None:
          print "Cannot find a path to save AC-PC PTS file in the database !"
        else:
           plotsACPC = self.getAllPlotsCentersAnyReferential('AC-PC')
           self.savePTS(path = wdiptsacpc.fullPath(), contacts=plotsACPC)
           wdiptsacpc.setMinf('referential', 'AC-PC')
           wdiptsacpc.setMinf('timestamp', timestamp)
           neuroHierarchy.databases.insertDiskItem(wdiptsacpc, update=True )

      #return

      #print("start generate dictionaries")
      #self.marsatlasExportElectrodes(Callback=lambda:[self.marsatlasExportResection(),self.exportCSVdictionaries(),self.generateMappingContactCortex()])

  def generateDictionaries(self):

      if self.generateDictionariesComboBox.currentIndex()==1:
          self.parcelsExportElectrodes()
      elif self.generateDictionariesComboBox.currentIndex()==0:
          #Pour que l'exportResection et l'exportCSVdictionaries attendent que export electrodes soit fini
          self.parcelsExportElectrodes(Callback=lambda:[self.marsatlasExportResection(),self.exportCSVdictionaries(),self.generateMappingContactCortex(),self.generateTemplateBipoleStimulationFile(),self.screenshot(),self.makeMP4(),self.calculParcel()])
          #self.marsatlasExportResection()
          #self.exportCSVdictionaries()
      elif self.generateDictionariesComboBox.currentIndex()==2:
          self.marsatlasExportResection()
      elif self.generateDictionariesComboBox.currentIndex()==3:
          self.exportCSVdictionaries()
      elif self.generateDictionariesComboBox.currentIndex()==4:
          self.generateMappingContactCortex()
      elif self.generateDictionariesComboBox.currentIndex()==5:
          self.generateTemplateBipoleStimulationFile()
      elif self.generateDictionariesComboBox.currentIndex()==6:
          self.screenshot()
      elif self.generateDictionariesComboBox.currentIndex()==7:
          self.makeMP4()    
      elif self.generateDictionariesComboBox.currentIndex()==8:
          self.calculParcel()
      elif self.generateDictionariesComboBox.currentIndex()==9:
          self.generateFiberContactDistance()

  def screenshot(self):
      #Check if all the volumes are available.
      #self.verifParcel allows us to know if the verification has already been made, in that case we won't do it again
      Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskleft = Mask_left.findValue(self.diskItems['T1pre'])
  
      Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskright = Mask_right.findValue(self.diskItems['T1pre'])

      MaskGW_right = ReadDiskItem('Right Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_right = MaskGW_right.findValue(self.diskItems['T1pre'])

      MaskGW_left = ReadDiskItem('Left Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_left = MaskGW_left.findValue(self.diskItems['T1pre'])
      
      if diMaskleft is None or diMaskright is None or diMaskGW_right is None or diMaskGW_left is None :
          #if a parcellation is missing we won't be able to make the screenshot, so we exit the function
          print "No parcellisation found."  
          return
      else:
          #we take out the cursor
          self.a.config()[ 'linkedCursor' ] = 0
          #and replace the electrodes realistic model with 2mm spheres, which are more visible
          self.updateDispMode(2)
          
          #Right hemisphere
          #opens an Anatomist window and loads the parcellation and the electrodes
          w2=self.a.createWindow("Sagittal")
          contWin=self.windowContent['IRM pre + MARS ATLAS right']
          content=[x for x in contWin if "rightMARSATLAS" in x]
          ele=[x for x in contWin if "electrode" in x]
          elec=self.dispObj[ele[0]] 
          for el in elec:
              w2.addObjects(el)
          image = self.dispObj[content[0]]
          w2.addObjects(image)
          time.sleep(1)
          w2.windowConfig(snapshot="/tmp/"+self.brainvisaPatientAttributes['subject']+"RightHemiSagittal.png")
          #rotates the view by 180°
          w2.camera(view_quaternion=[-0.5, 0.5, 0.5, -0.5])     
          #waits a second so that the window has the tame to be updated before the next screenshot, they are saved in the tmp repertory
          time.sleep(1)
          w2.windowConfig(snapshot="/tmp/"+self.brainvisaPatientAttributes['subject']+"RightHemiSagittalRot.png")
          
          w2.close()
          
          #Left hemisphere          
          w21=self.a.createWindow("Sagittal")
          contWin=self.windowContent['IRM pre + MARS ATLAS left']
          content=[x for x in contWin if "leftMARSATLAS" in x]
          ele=[x for x in contWin if "electrode" in x]
          elec=self.dispObj[ele[0]] 
          for el in elec:
              w21.addObjects(el)
          image = self.dispObj[content[0]]
          w21.addObjects(image)
          time.sleep(1)
          w21.windowConfig(snapshot="/tmp/"+self.brainvisaPatientAttributes['subject']+"LeftHemiSagittal.png")
          w21.camera(view_quaternion=[-0.5, 0.5, 0.5, -0.5])          
          time.sleep(1)
          w21.windowConfig(snapshot="/tmp/"+self.brainvisaPatientAttributes['subject']+"LeftHemiSagittalRot.png")
          w21.close()

          #puts the cursor back
          self.a.config()[ 'linkedCursor' ] = 1
          #opens the pictures of the left and right brain which were taken previously
          images1 = map(Image.open, ["/tmp/"+self.brainvisaPatientAttributes['subject']+"RightHemiSagittal.png", "/tmp/"+self.brainvisaPatientAttributes['subject']+"RightHemiSagittalRot.png"])
          images2 = map(Image.open, ["/tmp/"+self.brainvisaPatientAttributes['subject']+"LeftHemiSagittal.png","/tmp/"+self.brainvisaPatientAttributes['subject']+"LeftHemiSagittalRot.png"])
          
          #calculates the width and height of the new picture we are going to make
          widths1, heights1 = zip(*(i.size for i in images1))
          widths2, heights2 = zip(*(i.size for i in images2))
          
          total_width1 = sum(widths1)
          max_height1 = max(heights1)
          total_width2 = sum(widths2)
          max_height2 = max(heights2)
          
          #opens new empty images of the size calculated before
          new_im1 = Image.new('RGB', (total_width1, max_height1))
          new_im2 = Image.new('RGB', (total_width2, max_height2))
        
          #we are then going to paste the two images from right and left hemispheres together by the x axis
          x_offset = 0
          for im in images1:
              new_im1.paste(im, (x_offset,0))
              x_offset += im.size[0]
              
          new_im1.save('/tmp/associated1.png')
          
          x_offset = 0
          for im in images2:
              new_im2.paste(im, (x_offset,0))
              x_offset += im.size[0]   
          
          new_im2.save('/tmp/associated2.png')
          
          #pastes the two images obtained previously by the y axis
          images = map(Image.open, ["/tmp/associated1.png","/tmp/associated2.png"])    
          new_im=Image.new('RGB', (total_width1, max_height1*2))
          
          y_offset = 0
          for im in images:
              new_im.paste(im, (0,y_offset))
              y_offset += im.size[1]  
          
          
          new_im.save('/tmp/associated.png')
          #verification that the path is creatable
          wdi = WriteDiskItem('Screenshot of Mars Atlas','PNG image')
          di=wdi.findValue(self.diskItems['T1pre'])

          if di is None:
              print "Can't generate file"
              return
          else:
              try:
                  os.mkdir(os.path.dirname(di.fullPath()))                                                          
                  #line0 = runCmd(cmd0) 
              except:
                  line0=1
              cmd1 = ['mv', '/tmp/associated.png', str(di.fullPath())]
              line1 = runCmd(cmd1)
              #updates the database with the image of the 2 views of the 2 parcellation
              neuroHierarchy.databases.insertDiskItem(di, update=True )
      #puts back the realistic electrode model        
      self.updateDispMode(0)
          
  def makeMP4(self):
        

        from brainvisa import quaternion
        
        #takes the voxel size of the T1
        volume = aims.read(str(self.diskItems['T1pre']))
        sizeT1=volume.getVoxelSize()
        
        #verification that a CT or MRI is present, if both are, we take the CT
        T1post=None
        diCT = ReadDiskItem( 'CT', 'BrainVISA volume formats', requiredAttributes={'center':self.brainvisaPatientAttributes['center'], 'subject':self.brainvisaPatientAttributes['subject'] } )
        path = list(diCT.findValues({}, None, False ))
        #check if it is a CT post

        idCTpost = [x for x in range(len(path)) if 'CTpost' in str(path[x]) and not 'CTpostOp' in str(path[x])]

        try:
            path=path[idCTpost[0]].fullPath()
            volCT = aims.read(path)
            npCT = volCT.arraydata()
            PresCT = True
        except: 
            diMRI = ReadDiskItem( 'Raw T1 MRI', 'BrainVISA volume formats', requiredAttributes={'center':self.brainvisaPatientAttributes['center'], 'subject':self.brainvisaPatientAttributes['subject'] } )
            pathMRI = list(diMRI.findValues({}, None, False ))         
            id_post = [x for x in range(len(pathMRI)) if 'post' in str(pathMRI[x]) and not 'postOp' in str(pathMRI[x])]
            pathMRI=pathMRI[0].fullPath()
            volCT = aims.read(pathMRI)
            npCT = volCT.arraydata()
            PresCT=False
            if id_post!=[]:    
                T1post = pathMRI[id_post[0]]
        
        #if both CT and MRI are absent, we won't do the computation

        if T1post is None and idCTpost==[]:
            warning = QtGui.QMessageBox(self)
            warning.setText("No CT post or MRI post found")
            warning.setWindowTitle("Warning")
            return
        
        #finding the brainMask if there is one
        try:
            brainMask = ReadDiskItem('Brain Mask', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
            diBrain = list(brainMask.findValues({}, None, False ))
            diBrain0=diBrain[0].fullPath()
            volBrainMask = aims.read(diBrain0)
            brainMaskArray = volBrainMask.arraydata() 
        except:
            brainMaskArray=None 
        
        #if the brainMask is present, the gif will be made taking only the brain tissues
        if brainMaskArray is not None:
            z=0
            while z<brainMaskArray.shape[1]:
                maxi=brainMaskArray[0,z,:,:].max()
                if maxi != 0:
                    i=round(z*sizeT1[2])
                    z=brainMaskArray.shape[1]
                z+=1
            z-=2
            while z>0:
                maxi=brainMaskArray[0,z,:,:].max()
                if maxi != 0:
                    limit=round(z*sizeT1[2])
                    z=0
                z-=1 
        else:
            i=1  
            limit=npCT.shape[1]
            
        #Erase all previous files
        os.system('find /tmp/gif*.png -type f -exec rm -f {} \;')
        
        #Take out the cursor
        self.a.config()[ 'linkedCursor' ] = 0
        #put the realistic electrode model back
        self.updateDispMode(2)
        
        #Electrodes GIF 
        #load the window
        w21=self.a.createWindow("Axial")
        w21.windowConfig(fullscreen=1)
        #if it is a CT
        if PresCT==True:
            contWin=self.windowContent['CT post']
            content=[x for x in contWin if "CTpost" in x]
            ele=[x for x in contWin if "electrode" in x]
        #if it is an MRI    
        else:
            contWin=self.windowContent['IRM post']
            content=[x for x in contWin if "T1post" in x]
            ele=[x for x in contWin if "electrode" in x]
        #add the image and electrodes    
        image = self.dispObj[content[0]]
        elec=self.dispObj[ele[0]] 
        w21.addObjects(image)
        for el in elec:
            w21.addObjects(el)
        #get all referentials    
        refs=self.a.getReferentials()
        #choose the T1 pre referential...
        for el in refs:
            info=str(el.getInfos())
            if 'T1pre' in info and 'native' in info:
                r1=el    
        #...and assign it        
        w21.assignReferential(r1)
        #activate the clipping
        self.a.execute('WindowConfig',windows = [w21],clipping=2,clip_distance=5.)
        if brainMaskArray is None:
            limit=npCT.shape[1]
        #take pictures of slices following the z axis    
        snapShots=[]
        while i<limit:
            self.a.execute( 'LinkedCursor', window=w21, position=( 0, 0, i ) )   
            w21.windowConfig(snapshot="/tmp/gifCT%03i.png"%(i))
            snapShots.append("/tmp/gifCT%03i.png"%(i))
            i+=1    
        w21.close()
        brainvisaContext = defaultContext()
        brainvisaContext.runProcess('mpegEncode_avconv', animation="/tmp/animationElec.mp4",images = snapShots[:-1],encoding='h264',quality=75,passes=1)

        wdi = WriteDiskItem('MP4 of Electrodes','MP4 film')
        if PresCT==True:
            di=wdi.findValue(self.diskItems['CTpost'])
        else:
            di=wdi.findValue(self.diskItems['T1post'])
              
        if di is None:
            print "Can't generate file"
            return
        else:
            try:
               os.mkdir(os.path.dirname(di.fullPath()))
            except:
               pass
               
            cmd1 = ['mv', "/tmp/animationElec.mp4" , str(di.fullPath())]
            line1 = runCmd(cmd1)
            neuroHierarchy.databases.insertDiskItem(di, update=True )
            
        #mpegConfig.mpegFormats[3]    
        #Mars Atlas GIF
        #Check if all the volumes are available.
        Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diMaskleft = Mask_left.findValue(self.diskItems['T1pre'])

        Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diMaskright = Mask_right.findValue(self.diskItems['T1pre'])

        MaskGW_right = ReadDiskItem('Right Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diMaskGW_right = MaskGW_right.findValue(self.diskItems['T1pre'])

        MaskGW_left = ReadDiskItem('Left Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diMaskGW_left = MaskGW_left.findValue(self.diskItems['T1pre'])
            
        
        if diMaskleft is None or diMaskright is None or diMaskGW_right is None or diMaskGW_left is None :
            print "No parcellisation found."          
        else:
            #Right hemisphere
            w2=self.a.createWindow("Sagittal")
            w2.windowConfig(fullscreen=1)
            contWin=self.windowContent['IRM pre + MARS ATLAS right']
            #contWin = self.windowContent['IRM pre + hemispheres']
            content=[x for x in contWin if "rightMARSATLAS" in x]
            #content = [x for x in contWin if "rightHemi" in x]
            ele=[x for x in contWin if "electrode" in x]
            elec=self.dispObj[ele[0]] 
            for el in elec:
                w2.addObjects(el)
            image = self.dispObj[content[0]]
            w2.addObjects(image)
            time.sleep(1)
            w2.windowConfig(snapshot="/tmp/gifR000.png")
            i=1
            v0=[0.5, 0.5, 0.5, 0.5]
            v1=[-0.5,0.5,0.5,-0.5] 
            incr=float(1)/174
            while i<176:
                quat=quaternion.Quaternion((v0[0]*(1-incr*i)+v1[0]*incr*i,v0[1]*(1-incr*i)+v1[1]*incr*i,v0[2]*(1-incr*i)+v1[2]*incr*i,v0[3]*(1-incr*i)+v1[3]*incr*i)).normalized().vector()
                w2.camera(view_quaternion=quat)
                w2.windowConfig(snapshot="/tmp/gifR%03i.png"%(i))
                i+=1  
            i=0    
            v1=[0.5, 0.5, 0.5, 0.5]
            v0=[-0.5, 0.5, 0.5, -0.5] 
            while i<175:
                quat=quaternion.Quaternion((v0[0]*(1-incr*i)-v1[0]*incr*i,v0[1]*(1-incr*i)-v1[1]*incr*i,v0[2]*(1-incr*i)-v1[2]*incr*i,v0[3]*(1-incr*i)-v1[3]*incr*i)).normalized().vector()
                w2.camera(view_quaternion=quat)                                                               
                w2.windowConfig(snapshot="/tmp/gifR%03i.png"%(i+175))
                i+=1    
            w2.close()  
        
            #Left hemisphere
            w2=self.a.createWindow("Sagittal")
            w2.windowConfig(fullscreen=1)
            contWin=self.windowContent['IRM pre + MARS ATLAS left']
            content=[x for x in contWin if "leftMARSATLAS" in x]
            ele=[x for x in contWin if "electrode" in x]
            elec=self.dispObj[ele[0]] 
            for el in elec:
                w2.addObjects(el)
            image = self.dispObj[content[0]]
            w2.addObjects(image)
            time.sleep(1)
            w2.windowConfig(snapshot="/tmp/gifL000.png")
            i=1
            v0=[0.5, 0.5, 0.5, 0.5]
            v1=[-0.5,0.5,0.5,-0.5] 
            incr=float(1)/174
            while i<175:
                quat=quaternion.Quaternion((v0[0]*(1-incr*i)+v1[0]*incr*i,v0[1]*(1-incr*i)+v1[1]*incr*i,v0[2]*(1-incr*i)+v1[2]*incr*i,v0[3]*(1-incr*i)+v1[3]*incr*i)).normalized().vector()
                w2.camera(view_quaternion=quat)    
                w2.windowConfig(snapshot="/tmp/gifL%03i.png"%(i))
                i+=1  
            i=0    
            v1=[0.5, 0.5, 0.5, 0.5]
            v0=[-0.5,0.5,0.5,-0.5] 
            while i<175:
                quat=quaternion.Quaternion((v0[0]*(1-incr*i)-v1[0]*incr*i,v0[1]*(1-incr*i)-v1[1]*incr*i,v0[2]*(1-incr*i)-v1[2]*incr*i,v0[3]*(1-incr*i)-v1[3]*incr*i)).normalized().vector()
                w2.camera(view_quaternion=quat)                                                               
                w2.windowConfig(snapshot="/tmp/gifL%03i.png"%(i+175))
                i+=1    
            w2.close()  
        
            new_list_image = []
            j=0
            while j<350:
                #print 'j', j
                images = map(Image.open,["/tmp/gifL%03i.png"%(j),"/tmp/gifR%03i.png"%(j)])
                
                widths, heights = zip(*(i.size for i in images))
          
                total_width = sum(widths)
                max_height = max(heights)
          
                new_im = Image.new('RGB', (total_width, max_height))

                x_offset = 0
                for im in images:
                    new_im.paste(im, (x_offset,0))
                    x_offset += im.size[0]
                
                new_im.save('/tmp/gifF%03i.jpg'%(j),quality=70)
                new_list_image.append('/tmp/gifF%03i.jpg'%(j))
                
                j+=1    
            #cmd1 = ['convert','-delay', '50', '-loop','0','/tmp/gif*.png','/home/gin11_stage/animation.gif']
            #os.system('convert -delay 50 -loop 0 /tmp/gif*.png /tmp/animation.gif')
            #line1 = runCmd(cmd1)
            
            brainvisaContext = defaultContext()
            brainvisaContext.runProcess('mpegEncode_avconv', animation='/tmp/animationMA.mp4',images = new_list_image,encoding='h264',quality=50,passes=1)
            wdi = WriteDiskItem('MP4 of Mars Atlas','MP4 film')
            di=wdi.findValue(self.diskItems['T1pre'])

            if di is None:
                print "Can't generate file"
                return
            else:
                try:
                  os.mkdir(os.path.dirname(di.fullPath()))
                except:
                  pass

                cmd1 = ['mv', '/tmp/animationMA.mp4', str(di.fullPath())]
                line1 = runCmd(cmd1)
                neuroHierarchy.databases.insertDiskItem(di, update=True )
                
        self.updateDispMode(0)
        self.a.config()[ 'linkedCursor' ] = 1     
        self.a.execute('WindowConfig',windows = [self.wins[0]],clipping=0)
        
        print("MP4 done")


  def fitvolumebyellipse(self,volumeposition):

        #cut the hippocampus in two:
        # from https://github.com/minillinim/ellipsoid/blob/master/ellipsoid.py
        tolerance = 0.025
        if len(volumeposition)>16000:
            volumeposition = [volumeposition[x] for x in range(len(volumeposition)) if x & 1 ==0]

        volumeposition=numpy.array(volumeposition)
        (N, dd) = numpy.shape(volumeposition)
        dd = float(dd)

        # Q will be our working array
        Q = numpy.vstack([numpy.copy(volumeposition.T), numpy.ones(N)])
        QT = Q.T

        # initializations
        err = 1.0 + tolerance
        u = (1.0 / N) * numpy.ones(N)

        # Khachiyan Algorithm
        while err > tolerance:
            V = numpy.dot(Q, numpy.dot(numpy.diag(u), QT))
            M = numpy.diag(numpy.dot(QT , numpy.dot(numpy.linalg.inv(V), Q)))    # M the diagonal vector of an NxN matrix
            j = numpy.argmax(M)
            maximum = M[j]
            step_size = (maximum - dd - 1.0) / ((dd + 1.0) * (maximum - 1.0))
            new_u = (1.0 - step_size) * u
            new_u[j] += step_size
            err = numpy.linalg.norm(new_u - u)
            u = new_u

        # center of the ellipse
        center = numpy.dot(volumeposition.T, u)

        # the A matrix for the ellipse
        AA = numpy.linalg.inv(numpy.dot(volumeposition.T, numpy.dot(numpy.diag(u), volumeposition)) - numpy.array([[a * b for b in center] for a in center])) / dd

        # Get the values we'd like to return
        UU, ss, rotation = numpy.linalg.svd(AA)
        radii = 1.0/numpy.sqrt(ss)

        return radii   #UU, ss, rotation, center, 
   
          
  def calculParcel(self):
      import copy
      
      infos={}
      parcels={}
      tailleG=0
      
      #take the parcellations
      Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskleft = Mask_left.findValue(self.diskItems['T1pre'])

      Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskright = Mask_right.findValue(self.diskItems['T1pre'])
      
      #take the parcels names in one list
      parcels_names = readSulcusLabelTranslationFile('parcels_label_name.txt')
      
      try:
          #transform the images to matrices
          Maskleft = aims.read(diMaskleft.fullPath())
          npleftMarsAtlas = Maskleft.arraydata()
      
          Maskright = aims.read(diMaskright.fullPath())
          nprightMarsAtlas = Maskright.arraydata()
      except:
          print "No parcellation found"
          return
      #take voxel size of T1 pre
      volume = aims.read(str(self.diskItems['T1pre']))
      sizeT1=volume.getVoxelSize()
      
      print 'verif done'
      #pdb.set_trace()
      #calculate the total volume of the parcels
      for key, value in parcels_names.items():
          if value[0].lower()=='l':
              parcel1 = numpy.where(npleftMarsAtlas == key)
          else:
              parcel1 = numpy.where(nprightMarsAtlas == key)
          coord=zip(parcel1[1],parcel1[2],parcel1[3])
          tailleG+=len(coord)
              
      print 'tailleG Done'     

      t=1
      #we do the computations for each parcel
      for key, value in parcels_names.items():
          print 'doing parcel ' ,value ,t
          t+=1
          par=0
          #we locate the parcel in the right or left hemispheres (if they begin with 'L':left, 'R':right)
          if value[0].lower()=='l':
              parcel1 = numpy.where(npleftMarsAtlas == key)
              try:
                  #we take only the parcel 
                  parcel=copy.deepcopy(npleftMarsAtlas[0,parcel1[1].min():parcel1[1].max()+1,parcel1[2].min():parcel1[2].max()+1,parcel1[3].min():parcel1[3].max()+1])
              except:
                  #if the parcel is absent we update it with the value 'not found'
                  print 'no parcel found'
                  parcels.update({value:"not found"})
                  par=None
          else:
              parcel1 = numpy.where(nprightMarsAtlas == key)
              try:
                  parcel=copy.deepcopy(nprightMarsAtlas[0,parcel1[1].min():parcel1[1].max()+1,parcel1[2].min():parcel1[2].max()+1,parcel1[3].min():parcel1[3].max()+1])
              except:
                  print 'no parcel found'
                  parcels.update({value:"not found"})
                  par=None
          #if the parcel has been computed        
          if par is not None:       
              #take the coordinates of the points
              coord=zip(parcel1[1],parcel1[2],parcel1[3])
              #volume of the parcel
              taille=len(coord)
              infos.update({"volume":int(taille)})
              parcelPositionN=[]
              parcelPositionNT=[]
              parcelPosition = [[round(parcel1[1][i]*sizeT1[2]),round(parcel1[2][i]*sizeT1[1]),round(parcel1[3][i]*sizeT1[0])] for i in range(len(parcel1[1]))]
              i=0
              #we take one over 3 values (downsampling by 3)
              while i<len(parcelPosition):
                  parcelPositionN.append(parcelPosition[i])
                  i+=3
              #pdb.set_trace()    
              d=0
              if len(parcelPositionN)>10000:
                  downSamp=int(math.ceil(float(len(parcelPositionN))/10000))
                  while d<len(parcelPositionN):
                      parcelPositionNT.append(parcelPositionN[d])
                      d+=downSamp
                  print "Downsampling more"  
                  print 'downSamp',downSamp
                  parcelPositionN=parcelPositionNT
              #pdb.set_trace()    
              #approximation by an ellips which will give the length width and height 
              fit=self.fitvolumebyellipse(parcelPositionN)
              print fit
              width=fit[0]
              height=fit[1]
              length=fit[2]
              infos.update({"width":width})
              infos.update({"length":length})
              infos.update({"height":height})
              parcel[parcel!=key]=0
              #center of mass of the parcel
              cM=ndimage.measurements.center_of_mass(parcel) 
              centreMasse=(cM[0]+parcel1[1].min()*sizeT1[2],cM[1]+parcel1[2].min()*sizeT1[1],cM[2]+parcel1[3].min()*sizeT1[0])
              infos.update({"center of mass":centreMasse})
              rapport=float(taille)/tailleG
              infos.update({"ratio":rapport})
              parcels.update({value:infos})
              infos={} 
      
      #transformation to the MNI referential
      points=[]
      for element in parcels:
          points.append(parcels[element]['center of mass'])
      
      field = self.getT1preMniTransform()
      if field is None:
          print "MNI deformation field not found"
          return None
      tmpOutput = '/tmp/test.csv' #pour tester
      arr = numpy.asarray(points) #tous tes centres de masses pour toutes les parcels tel quel ([ [1,2,3], [4,5,6], [7,8,9] ])
      numpy.savetxt(tmpOutput, arr, delimiter=",")
      matlabRun(spm_normalizePoints % ("'"+self.spmpath+"'",field, tmpOutput, tmpOutput))
      out = numpy.loadtxt(tmpOutput, delimiter=",")
      os.remove(tmpOutput)
      for element in out:
          for parcel in parcels:
              if all([points[[list(x) for x in tuple(out)].index(list(element))][i] == parcels[parcel]['center of mass'][i] for i in range(len(parcels[parcel]['center of mass']))]):
              #if parcels[parcel]['center of mass']==points[[list(x) for x in tuple(out)].index(list(element))]:
                  parcels[parcel]['center of mass']=tuple(element)
      valPatient={"Mars Atlas":parcels} 
      wdi = WriteDiskItem('Parcels Infos','Atlas metrics')
      di=wdi.findValue(self.diskItems['T1pre'])     
      
      #writes the file
      fout = open(di.fullPath(),'w')
      fout.write(json.dumps(valPatient))
      fout.close()

      #updates the database
      neuroHierarchy.databases.insertDiskItem(di, update=True )
      
      print "parcel metrics done"
  
  def generateTemplateBipoleStimulationFile(self):
      
      return
      excel_data = openpyxl.load_workbook('template_clincalbipoleLabel.xlsx')
      worksheet_names = excel_data.get_sheet_names()
      datasheet1 = excel_data[worksheet_names[0]]
      
      plotsSB = self.getAllPlotsCentersT1preScannerBasedRef()
      info_plotSB = []
      for k,v in plotsSB.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plotSB.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          
      plotSB_sorted = sorted(info_plotSB, key=lambda plot_number:plot_number[0])
      
      #montage bipolaire
      info_plotSB_bipolaire = []
      for pindex in range(1,len(plotSB_sorted)):
          previous_contact = "".join([i for i in plotSB_sorted[pindex-1][0] if not i.isdigit()])
          current_contact = "".join([i for i in plotSB_sorted[pindex][0] if not i.isdigit()])
          if previous_contact == current_contact:
               info_plotSB_bipolaire.append((plotSB_sorted[pindex][0]+' - '+ plotSB_sorted[pindex-1][0],(numpy.array(plotSB_sorted[pindex][1])+numpy.array(plotSB_sorted[pindex-1][1]))/2 ))
               
      plotSB_sorted=dict(plotSB_sorted)
      #info_plotSB_bipolaire=dict(info_plotSB_bipolaire)
      
      for bip_index in range(len(info_plotSB_bipolaire)):
          actual_cell = datasheet1.cell(row=bip_index*2+2, column = 1)
          actual_cell.set_explicit_value(info_plotSB_bipolaire[bip_index][0])
          actual_cell.style.font.size = 11
          actual_cell.style.font.bold = True
          actual_cell = datasheet1.cell(row=bip_index*2+2, column = 2)
          actual_cell.set_explicit_value('50 Hz')
          actual_cell.style.font.size = 11
          actual_cell.style.font.bold = True
          actual_cell = datasheet1.cell(row=bip_index*2+3, column = 1)
          actual_cell.set_explicit_value(info_plotSB_bipolaire[bip_index][0])
          actual_cell.style.font.size = 11
          actual_cell.style.font.bold = True
          actual_cell = datasheet1.cell(row=bip_index*2+3, column = 2)
          actual_cell.set_explicit_value('1 Hz')
          actual_cell.style.font.size = 11
          actual_cell.style.font.bold = True
      
      path_to_save = self.fileNoDBpath + os.path.sep + self.brainvisaPatientAttributes['subject'] + '_bipole' + '.xlsx'
      
      excel_data.save(path_to_save)
  
  
  def generateMappingContactCortex(self):
      
      print("not finished")
      return
      SizeHorizon = 10

      #check if the meshes has been generated
      rdi_mesh_right = ReadDiskItem( 'Right Hemisphere White Mesh', 'Anatomist mesh formats', requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol} )
      hemis = list(rdi_mesh_right._findValues( {}, None, False ) )
      hemispre_right = [hemis[i] for i in range(len(hemis)) if "T1pre" in hemis[i].attributes()['acquisition']]

      rdi_mesh_left = ReadDiskItem( 'Left Hemisphere White Mesh', 'Anatomist mesh formats', requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol} )
      hemis = list(rdi_mesh_left._findValues( {}, None, False ) )
      hemispre_left = [hemis[i] for i in range(len(hemis)) if "T1pre" in hemis[i].attributes()['acquisition']]

      MaskGW_right = ReadDiskItem('Right Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_right = MaskGW_right.findValue(self.diskItems['T1pre'])

      MaskGW_left = ReadDiskItem('Left Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_left = MaskGW_left.findValue(self.diskItems['T1pre'])
      
      FreesurferThickness = ReadDiskItem('FreesurferThicknessType','Texture',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'],'_ontology':'freesurfer'})
      diAllFSThick = list(FreesurferThickness.findValues({},None,False))
      indexThickRight = [i for i in range(len(diAllFSThick)) if 'rh.thickness' in diAllFSThick[i].fullPath()]
      indexThickLeft = [i for i in range(len(diAllFSThick)) if 'lh.thickness' in diAllFSThick[i].fullPath()]
      
      FreesurferWhite = ReadDiskItem('White','Anatomist mesh formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'],'_ontology':'freesurfer'})
      diFSWhite = list(FreesurferWhite.findValues({},None,False))
      indexFSWhiteRight = [i for i in range(len(diFSWhite)) if 'rh.white' in diFSWhite[i].fullPath()]
      indexFSWhiteLeft = [i for i in range(len(diFSWhite)) if 'lh.white' in diFSWhite[i].fullPath()]      


      if len(hemispre_right) <= 0 or len(hemispre_left) <= 0:
          print "no mesh found"
          return

      meshRight = aims.read(str(hemispre_right[0]))
      meshLeft = aims.read(str(hemispre_left[0]))
      
      if len(diFSWhite)>0:
        meshFSRight = aims.read(str(diFSWhite[indexFSWhiteRight[0]]))
        meshFSLeft = aims.read(str(diFSWhite[indexFSWhiteLeft[0]]))

      #get plots information
      plots = self.getAllPlotsCentersT1preRef() #coordonnées en mm à convertir en voxel
      if len(plots)==0:
          print("no contact found")
          return

      print "start matching contacts with vertexes"

      info_plot= []
      for k,v in plots.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plot.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          #plots_label[k]=(label,label_name)

      plot_sorted = sorted(info_plot, key=lambda plot_number: plot_number[0])

      #montage bipolaire
      info_plot_bipolaire= []
      for pindex in range(1,len(plot_sorted)):
          previous_contact = "".join([i for i in plot_sorted[pindex-1][0] if not i.isdigit()])
          current_contact = "".join([i for i in plot_sorted[pindex][0] if not i.isdigit()])
          if previous_contact == current_contact:
               info_plot_bipolaire.append((plot_sorted[pindex][0]+' - '+ plot_sorted[pindex-1][0],(plot_sorted[pindex][1]+plot_sorted[pindex-1][1])/2 ))

      #here I have to calculate distances: vertex to each contacts.
      meshRight_vertex_np = numpy.array(meshRight.vertex())
      meshLeft_vertex_np = numpy.array(meshLeft.vertex())
      
      meshFSRight_vertex_np = numpy.array(meshFSRight.vertex())
      meshFSLeft_vertex_np = numpy.array(meshFSLeft.vertex())

      dist_contact_meshRight = {}
      dist_contact_meshLeft = {}

      #coordonnées longitude et latitude marsAtlas
      wdi_longitude = ReadDiskItem( 'Longitude coordinate texture', 'aims Texture formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol} )
      di_longitude = list(wdi_longitude.findValues({},None,False))
      wdi_latitude = ReadDiskItem( 'Latitude coordinate texture', 'aims Texture formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol} )
      di_latitude = list(wdi_latitude.findValues({},None,False))
         
      indexLongRight = [i for i in range(len(di_longitude)) if 'pre' in di_longitude[i].fullPath() and 'Rwhite' in di_longitude[i].fullPath()]
      indexLongLeft = [i for i in range(len(di_longitude)) if 'pre' in di_longitude[i].fullPath() and 'Lwhite' in di_longitude[i].fullPath()]
       
      indexLatRight = [i for i in range(len(di_latitude)) if 'pre' in di_latitude[i].fullPath() and 'Rwhite' in di_latitude[i].fullPath()]
      indexLatLeft = [i for i in range(len(di_latitude)) if 'pre' in di_latitude[i].fullPath() and 'Lwhite' in di_latitude[i].fullPath()]
      
      MatLongLeft = aims.read(di_longitude[indexLongLeft[0]].fullPath())
      MatLatLeft = aims.read(di_latitude[indexLatLeft[0]].fullPath())
      
      MatLongRight = aims.read(di_longitude[indexLongRight[0]].fullPath())
      MatLatRight = aims.read(di_latitude[indexLatRight[0]].fullPath())

      if len(diFSWhite)>0:
        FSThickRight = aims.read(diAllFSThick[indexThickRight[0]].fullPath())
        FSTHickLeft = aims.read(diAllFSThick[indexThickLeft[0]].fullPath())
      
      import copy
      for ii in range(len(plot_sorted)):

         print "{} on {}".format(ii+1,len(plot_sorted))

         np_plot = numpy.array(plot_sorted[ii][1])
         diffRight = meshRight_vertex_np - np_plot  #la je sais pas si il y en a pas un en voxel et un en mm.
         diffLeft = meshLeft_vertex_np - np_plot
         if len(diFSWhite)>0:
           diffFSRight = meshFSRight_vertex_np - np_plot
           diffFSLeft = meshFSLeft_vertex_np - np_plot

         
         distRight = [numpy.linalg.norm(diffRight[i]) for i in range(len(diffRight))]
         distLeft = [numpy.linalg.norm(diffLeft[i]) for i in range(len(diffLeft))]
         if len(diFSWhite)>0:
           distFSRight = [numpy.linalg.norm(diffFSRight[i]) for i in range(len(diffFSRight))]
           distFSLeft = [numpy.linalg.norm(diffFSLeft[i]) for i in range(len(diffFSLeft))]
         

         #it's too big, I should keep only the indexes of the one nearest than 2 times SizeHorizon
         indexvertexLeft = numpy.where(numpy.array(distLeft)<SizeHorizon)
         indexvertexRight = numpy.where(numpy.array(distRight)<SizeHorizon)
         
         indexSelectedLeft = dict([(indexvertexLeft[0][i],distLeft[indexvertexLeft[0][i]]) for i in range(len(indexvertexLeft[0]))])
         indexSelectedRight = dict([(indexvertexRight[0][i],distRight[indexvertexRight[0][i]]) for i in range(len(indexvertexRight[0]))])

         dist_contact_meshRight.update({plot_sorted[ii][0]:copy.deepcopy(indexSelectedRight)})
         dist_contact_meshLeft.update({plot_sorted[ii][1]:copy.deepcopy(indexSelectedLeft)})
         
         numpy.array(distRight).min()
         numpy.array(distLeft).min()
         distRight.index(numpy.array(distRight).min())
         distLeft.index(numpy.array(distLeft).min())

         if numpy.array(distRight).min() > numpy.array(distLeft).min():
             MarsAtlasCoordinate = ('L',MatLatLeft[0][distLeft.index(numpy.array(distLeft).min())],MatLongLeft[0][distLeft.index(numpy.array(distLeft).min())])
             pdb.set_trace()
             if len(diFSWhite)>0:
               ValueThickness = FSTHickLeft[0][distFSLeft.index(numpy.array(distFSLeft).min())]
               ValueThickness = (ValueThickness,)
             else:
               ValueThickness = (0,)  
             MarsAtlasCoordinate = MarsAtlasCoordinate + ValueThickness               
         elif numpy.array(distRight).min() < numpy.array(distLeft).min():
             MarsAtlasCoordinate = ('R',MatLatRight[0][distRight.index(numpy.array(distRight).min())],MatLongRight[0][distRight.index(numpy.array(distRight).min())])
             if len(diFSWhite)>0:
               ValueThickness = FSThickRight[0][distFSRight.index(numpy.array(distFSRight).min())]
               ValueThickness = (ValueThickness,)
             else:
               ValueThickness = (0,) 
             MarsAtlasCoordinate = MarsAtlasCoordinate + ValueThickness
             
         pdb.set_trace()


      pdb.set_trace()
      #test2 = [i for i in range(len(dist_contact_mesh1[1])) if dist_contact_mesh1[1][i] <= SizeHorizon]


      #generate the texture
      #textureContacts = aims.TimeTexture()
      print "contacts vertexes matching done"






  def exportCSVdictionaries(self):

    #test si eleclabel et resection sont générés
    #il faut au moins eleclabel, si pas resection pas grave

    #eleclabel
    rdi_eleclabel=ReadDiskItem('Electrodes Labels','Electrode Label Format',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
    wdi_eleclabel=list(rdi_eleclabel.findValues({},None,False))

    #reseclabel
    rdi_reseclabel=ReadDiskItem('Resection Description','Resection json',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
    wdi_reseclabel=list(rdi_reseclabel.findValues({},None,False))

    wdi = WriteDiskItem('Final Export Dictionaries','CSV file')
    di=wdi.findValue(self.diskItems['T1pre'])

    if len(wdi_eleclabel)==0:
      print('no ...')
    else:
      di_eleclabel=rdi_eleclabel.findValue(self.diskItems['T1pre'])

      fin = open(di_eleclabel.fullPath(),'r')
      info_label_elec = json.loads(fin.read())
      fin.close()

      plots = self.getAllPlotsCentersT1preScannerBasedRef()

      #look for mni position, are there already calculated ?
      rdi_elecimplant = ReadDiskItem( 'Electrode implantation', 'Electrode Implantation format',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
      impl = list(rdi_elecimplant.findValues({},None,False))

      if len(impl) == 0:
        print "Cannot find implantation"
        plotsMNI = self.getAllPlotsCentersMNIRef(plots)
        info_plotMNI= []
        for k,v in plotsMNI.iteritems():
           plot_name_split = k.split('-$&_&$-')
           info_plotMNI.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))

      elif os.path.exists(str(impl[0])):
        filein = open(str(impl[0]), 'rb')
        try:
          dic_impl = json.loads(filein.read())
        except:
          filein.close()
          filein = open(str(impl[0]), 'rb')
          dic_impl = pickle.load(filein)

          filein.close()

        if 'plotsMNI' in dic_impl.keys():
           print "MNI position already estimated, ok"
           plot_dict_MNIinter = dict(dic_impl['plotsMNI'])
           info_plotMNI = [(k, v) for k, v in plot_dict_MNIinter.iteritems()]

        else:
           plotsMNI = self.getAllPlotsCentersMNIRef(plots)

           info_plotMNI= []
           for k,v in plotsMNI.iteritems():
              plot_name_split = k.split('-$&_&$-')
              info_plotMNI.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          #plots_label[k]=(label,label_name)

      plotMNI_sorted = sorted(info_plotMNI, key=lambda plot_number: plot_number[0])

      #montage bipolaire
      info_plotMNI_bipolaire= []
      for pindex in range(1,len(plotMNI_sorted)):
          previous_contact = "".join([i for i in plotMNI_sorted[pindex-1][0] if not i.isdigit()])
          current_contact = "".join([i for i in plotMNI_sorted[pindex][0] if not i.isdigit()])
          if previous_contact == current_contact:
               info_plotMNI_bipolaire.append((plotMNI_sorted[pindex][0]+' - '+ plotMNI_sorted[pindex-1][0],(numpy.array(plotMNI_sorted[pindex][1])+numpy.array(plotMNI_sorted[pindex-1][1]))/2 ))

      plotMNI_sorted=dict(plotMNI_sorted)
      info_plotMNI_bipolaire=dict(info_plotMNI_bipolaire)

      
      plotsSB = self.getAllPlotsCentersT1preScannerBasedRef()
      info_plotSB = []
      for k,v in plotsSB.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plotSB.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          
      plotSB_sorted = sorted(info_plotSB, key=lambda plot_number:plot_number[0])
      
      #montage bipolaire
      info_plotSB_bipolaire = []
      for pindex in range(1,len(plotSB_sorted)):
          previous_contact = "".join([i for i in plotSB_sorted[pindex-1][0] if not i.isdigit()])
          current_contact = "".join([i for i in plotSB_sorted[pindex][0] if not i.isdigit()])
          if previous_contact == current_contact:
               info_plotSB_bipolaire.append((plotSB_sorted[pindex][0]+' - '+ plotSB_sorted[pindex-1][0],(numpy.array(plotSB_sorted[pindex][1])+numpy.array(plotSB_sorted[pindex-1][1]))/2 ))
               
      plotSB_sorted=dict(plotSB_sorted)
      info_plotSB_bipolaire=dict(info_plotSB_bipolaire)
      
      
      with open(di.fullPath(), 'w') as csvfile:
         #fieldnames=['MarsAtlas','GreyWhite','Resection']
         writer = csv.writer(csvfile, delimiter='\t')
         writer.writerow([u'Contacts Positions'])
         writer.writerow([u'Use of MNI Template','MarsAtlas',info_label_elec['Template']['MarsAtlas'],'Freesurfer',info_label_elec['Template']['Freesurfer'],'HippoSubfieldFreesurfer',info_label_elec['Template']['HippocampalSubfield Freesurfer']])

         #add a row with "MNI or Patient for MarsAtlas and Freesurfer
         list_to_write = set(info_label_elec['plots_label'][info_label_elec['plots_label'].keys()[0]].keys())
         list_by_default = set([u'contact','MarsAtlas','MarsAtlasFull', 'Freesurfer', 'Hippocampal Subfield','GreyWhite','AAL', 'AALDilate', 'Broadmann', 'BroadmannDilate', 'Hammers', 'Resection', 'MNI','T1pre Scanner Based'])
         diff_list = list(list_to_write.difference(list_by_default))
         full_list = [u'contact','MarsAtlas','MarsAtlasFull', 'Freesurfer', 'Hippocampal Subfield','GreyWhite', 'ALL', 'AALDilate', 'Broadmann','BroadmannDilate', 'Hammers', 'Resection', 'MNI','T1pre Scanner Based']
         full_list.extend(diff_list)
         writer.writerow(full_list)

         #pdb.set_trace()
         dict_sorted_tmp = OrderedDict(sorted(info_label_elec['plots_label'].items()))

         for kk,vv in dict_sorted_tmp.iteritems():
            listwrite = [kk]
            listwrite.append(vv['MarsAtlas'][1])
            listwrite.append(vv['MarsAtlasFull'])
            listwrite.append(vv['Freesurfer'][1])
            listwrite.append(vv['Hippocampal Subfield'][1])
            listwrite.append(vv['GreyWhite'][1])
            listwrite.append(vv['AAL'][1])
            listwrite.append(vv['AALDilate'][1])
            listwrite.append(vv['Broadmann'][1])
            listwrite.append(vv['BroadmannDilate'][1])            
            listwrite.append(vv['Hammers'][1])
            listwrite.append(vv['Resection'][1])
            #[listwrite.append(x[1]) for x in vv.values()]
            listwrite.append([float(format(plotMNI_sorted[kk][i],'.3f')) for i in range(3)])
            listwrite.append([float(format(plotSB_sorted[kk][i],'.3f')) for i in range(3)])
            if len(full_list)>12:
                for i_supp in range(len(full_list)-14):
                    listwrite.append(vv[full_list[14+i_supp]])
            writer.writerow(listwrite)

         writer.writerow([])
         writer.writerow([])

         dict_sorted_tmp = OrderedDict(sorted(info_label_elec['plots_label_bipolar'].items()))

         for kk,vv in dict_sorted_tmp.iteritems():
            #pdb.set_trace()
            listwrite = [kk]
            listwrite.append(vv['MarsAtlas'][1])
            listwrite.append(vv['MarsAtlasFull'])
            listwrite.append(vv['Freesurfer'][1])
            listwrite.append(vv['Hippocampal Subfield'][1])
            listwrite.append(vv['GreyWhite'][1])
            listwrite.append(vv['AAL'][1])
            listwrite.append(vv['AALDilate'][1])
            listwrite.append(vv['Broadmann'][1])
            listwrite.append(vv['BroadmannDilate'][1])             
            listwrite.append(vv['Hammers'][1])
            listwrite.append(vv['Resection'][1])
            #[listwrite.append(x[1]) for x in vv.values()]
            listwrite.append([float(format(info_plotMNI_bipolaire[kk][i],'.3f')) for i in range(3)])
            listwrite.append([float(format(info_plotSB_bipolaire[kk][i],'.3f')) for i in range(3)])
            if len(full_list)>12:
                for i_supp in range(len(full_list)-14):
                    listwrite.append(vv[full_list[14+i_supp]])
            writer.writerow(listwrite)

         writer.writerow([])
         writer.writerow([])

    if len(wdi_reseclabel)==0:
      print('no ...')
    else:
      di_reseclabel = rdi_reseclabel.findValue(self.diskItems['T1pre'])
      fin = open(di_reseclabel.fullPath(),'r')
      info_label_resec = json.loads(fin.read())
      fin.close()

      with open(di.fullPath(), 'a') as csvfile:
         writer = csv.writer(csvfile, delimiter='\t')
         writer.writerow([u'Resection Information'])
         #pdb.set_trace()
         for kk,vv in info_label_resec.iteritems():
            #writer.writerow([kk])
            if type(vv) == type(float()):
                listwrite = [kk,format(vv,'.1f')]
                writer.writerow(listwrite)
            else:
               writer.writerow([kk])
               for ll,bb in vv.iteritems():
                 listwrite = [ll, format(float(bb[1]),'.1f')]
                 writer.writerow(listwrite)

    neuroHierarchy.databases.insertDiskItem(di, update=True )
    print "export csv done"


    #wdi = WriteDiskItem('PatientInfoTemplate','Patient Template format')
    #di = wdi.findValue(self.diskItems['T1pre'])

  def marsatlasExportResection(self):

      wdi_resec = ReadDiskItem('Resection', 'NIFTI-1 image', requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
      di_resec = list(wdi_resec.findValues({}, None, False ))

      if len(di_resec)==0:
        print('no resection image found')
        return

      Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskleft = Mask_left.findValue(self.diskItems['T1pre'])

      Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskright = Mask_right.findValue(self.diskItems['T1pre'])

      FreesurferMask = ReadDiskItem('FreesurferAtlas', 'BrainVISA volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diFreesurferMask = list(FreesurferMask.findValues({}, None, False ))

      if diMaskleft is None:
        print('left gyri volume failed, perform export mars atlas export contact first')
        #return
      else:
        vol_left = aims.read(diMaskleft.fileName())

      if diMaskright is None:
        print('right gyri volume failed, perform export mars atlas export contact first')
        #return
      else:
        vol_right = aims.read(diMaskright.fileName())

      Vol_resec = aims.read(di_resec[0].fileName())


      Vox_size_resection = Vol_resec.getVoxelSize().arraydata()
      Vol_resection_mm = Vol_resec.arraydata().sum()*Vox_size_resection.prod() #choppe les trois premiers

      #intersection avec mars atlas label
      if diMaskleft is not None and diMaskright is not None:
        Vol_mask_tot = vol_left.arraydata()+vol_right.arraydata()

        inter_resec_mars_atlas = numpy.multiply(Vol_resec.arraydata(),Vol_mask_tot)
        label_resec_mars_atlas = numpy.histogram(inter_resec_mars_atlas,bins = 255, range = (0,255))
        total_label_mars_atlas = numpy.histogram(Vol_mask_tot,bins = 255)
        percent_resec_mars_atlas = numpy.divide(label_resec_mars_atlas[0],total_label_mars_atlas[0],dtype=float)*100
        non_zero_inter = numpy.add(numpy.nonzero(label_resec_mars_atlas[0][1:-1]),1)

        parcels_names = readSulcusLabelTranslationFile('parcels_label_name.txt')

        #pdb.set_trace()
        list1 = [parcels_names[label_resec_mars_atlas[1][x]] for x in non_zero_inter.tolist()[0]]
        list2 = [(label_resec_mars_atlas[1][x],percent_resec_mars_atlas[x]) for x in non_zero_inter.tolist()[0]]
        resection_mars_atlas_info = dict(zip(list1,list2)) #{parcels_names[label_resec_mars_atlas[1][x]]:(label_resec_mars_atlas[1][x],percent_resec_mars_atlas[x]) for x in non_zero_inter.tolist()[0]}

      else:
          resection_mars_atlas_info = {"MarsAtlas not calculated":[0,0]}

      if len(diFreesurferMask)>0:
        vol_FS = aims.read(diFreesurferMask[0].fileName())
        inter_resec_FS = numpy.multiply(Vol_resec.arraydata(),vol_FS.arraydata())
        label_resec_FS = numpy.histogram(inter_resec_FS,bins = 20000, range = (0,20000))
        total_label_FS = numpy.histogram(vol_FS.arraydata(),bins = 20000, range = (0,20000))
        percent_resec_FS = numpy.divide(label_resec_FS[0],total_label_FS[0],dtype=float)*100
        #we keep only value between 1 and 100 to not display the thousands freesurfer parcels...
        interesting_percent_resec_FS = numpy.where((percent_resec_FS >1) & (percent_resec_FS < 100))
        
        parcels_names_FS = readFreesurferLabelFile('freesurfer_label.txt')
        
        list1 = [parcels_names_FS[unicode(x)][0] for x in interesting_percent_resec_FS[0]]
        list2 = [(float(x),percent_resec_FS[x]) for x in interesting_percent_resec_FS[0]]
        
        resection_freesurfer_info = dict(zip(list1,list2))

      else:
         resection_freesurfer_info = {"Freesurfer not calculated": [0,0]}

      #write into database
      wdi = WriteDiskItem('Resection Description','Resection json')
      di = wdi.findValue(self.diskItems['Resection'])
      if di is None:
         print('Can t generate files')
         return

      fout = open(di.fullPath(),'w')
      fout.write(json.dumps({'mars_atlas':resection_mars_atlas_info,'Volume resection (mm3): ':Vol_resection_mm,'Freesurfer':resection_freesurfer_info})) #ici il faut ajouter l'ajout de la key volume mais je sais pas pourquoi des fois ca fait planter l'export en csv
      fout.close()

      neuroHierarchy.databases.insertDiskItem(di, update=True )

      print "export resection info done"



  def parcelsExportElectrodes(self, saveInDBmarsatlas = True, Callback = None):

      #check presence of marsatlas data

      LeftGyri = ReadDiskItem('hemisphere parcellation texture','Aims texture formats',requiredAttributes={ 'side': 'left' ,'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol, 'parcellation_type':'marsAtlas' })
      #LeftGyri = list(LeftGyri.findValues({}, None, False ))
      LeftGyri = LeftGyri.findValue(self.diskItems['T1pre'])

      #check the presence of freesurfer data
      freesurferdi = ReadDiskItem('FreesurferAtlas', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
      rdi_freesurfer = list(freesurferdi.findValues({}, None, False ))

      #check the presence of Hippo subfield freesurfer data
      HippoSubfielddi = ReadDiskItem('HippoFreesurferAtlas', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
      rdi_HippoSubfield = list(HippoSubfielddi.findValues({},None,False))

      #if either marsatlas and freesurfer are not found. stop (for now but later has to go throw the MNI)
      if LeftGyri is None:
      	print('no hemisphere parcellation texture found')
      	TemplateMarsAtlas = True
      	#return
      else:
        TemplateMarsAtlas = False

      if len(rdi_freesurfer) == 0:
        print('no freesurfer atlas found')
        TemplateFreeSurfer = True
      else:
        TemplateFreeSurfer = False

      if len(rdi_HippoSubfield)  == 0:
          print('no hippo subfield atlas found')
          TemplateHippoSubfieldFreesurfer = True
      else:
          TemplateHippoSubfieldFreesurfer = False


      #if TemplateMarsAtlas or TemplateFreeSurfer or TemplateHippoSubfieldFreesurfer: #plus besoin de if vu qu'il faut maintenant forcément passer par le mni pour les atlas mni
      rdi = ReadDiskItem( 'Electrode implantation', 'Electrode Implantation format',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
      impl = list(rdi.findValues({},None,False))

      if len(impl) == 0:
        print "Cannot find implantation"
        QtGui.QMessageBox.warning(self,"Error","Can't find implantation, you have to click on save and after on normalize/export to generate MNI position first")
        return
      elif os.path.exists(str(impl[0])):
        filein = open(str(impl[0]), 'rb')
        try:
          dic_impl = json.loads(filein.read())
        except:
          filein.close()
          filein = open(str(impl[0]), 'rb')
          dic_impl = pickle.load(filein)

          filein.close()

        if 'plotsMNI' in dic_impl.keys():
          print "MNI position already estimated, ok"
          plot_dict_MNIinter = dict(dic_impl['plotsMNI'])
        else:
          print "you have to click on normalize/export to generate MNI position first"
          QtGui.QMessageBox.warning(self,"Error","A template will have to be used, you have to click on normalize/export to generate MNI position first")
          return
      #else:
        #plot_dict_MNIinter=None
      
      #pdb.set_trace()
      brainvisaContext = defaultContext()

      #pour que lorsque le thread Brainvisa ça appelle parcellationdone ET les autres export si il y a besoin
      if Callback is not None:
        Callback2= lambda x=None,plotMNI = plot_dict_MNIinter, templateHPSub = TemplateHippoSubfieldFreesurfer, templateMA = TemplateMarsAtlas,templateFS=TemplateFreeSurfer:[self.parcellationDone(useTemplateMarsAtlas = templateMA,useTemplateFreeSurfer=templateFS,useTemplateHippoSubFreesurfer=templateHPSub,plot_dict_MNI=plotMNI),Callback()]
      else:
        Callback2 = lambda x=None,plotMNI = plot_dict_MNIinter, templateHPSub = TemplateHippoSubfieldFreesurfer, templateMA = TemplateMarsAtlas,templateFS=TemplateFreeSurfer:self.parcellationDone(useTemplateMarsAtlas = templateMA,useTemplateFreeSurfer=templateFS,useTemplateHippoSubFreesurfer=templateHPSub,plot_dict_MNI=plotMNI)

      try:
        brainvisaContext.runInteractiveProcess(Callback2,'2D Parcellation to 3D parcellation', Side = "Both", left_gyri = LeftGyri)  #, sulcus_identification ='label')
      except:
        Callback2()


  def parcellationDone(self,useTemplateMarsAtlas = False, useTemplateFreeSurfer = False, useTemplateHippoSubFreesurfer = False,plot_dict_MNI=None):

      #pdb.set_trace()
      print "export electrode start"
      timestamp = time.time()
      Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskleft = Mask_left.findValue(self.diskItems['T1pre'])

      Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskright = Mask_right.findValue(self.diskItems['T1pre'])

      MaskGW_right = ReadDiskItem('Right Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_right = MaskGW_right.findValue(self.diskItems['T1pre'])

      MaskGW_left = ReadDiskItem('Left Grey White Mask','Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diMaskGW_left = MaskGW_left.findValue(self.diskItems['T1pre'])
      

      #here put if on use template or no
      if diMaskleft is None:
        print('left gyri conversion surface to volume failed')
        useTemplateMarsAtlas = True
        #return

      if diMaskright is None:
        print('right gyri conversion surface to volume failed')
        useTemplateMarsAtlas = True
        #return

      GWAtlas = True

      if diMaskGW_left is None:
        print('not found left grey/white label')
        GWAtlas = False
      else:
        volGW_left = aims.read(diMaskGW_left.fileName())

      if diMaskGW_right is None:
        print('not found right grey/white label')
        GWAtlas = False
      else:
        volGW_right = aims.read(diMaskGW_right.fileName())


      #on s'intéresse à la résection
      wdi_resec = ReadDiskItem('Resection', 'NIFTI-1 image', requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
      di_resec = list(wdi_resec.findValues({}, None, False ))
      #careful, if we use templates, the resection has to be deformed in the mni template

      if len(di_resec)==0:
        print('no resection image found')
        DoResection = False
      else:
        DoResection = True
        vol_resec = aims.read(di_resec[0].fileName())

      #is there any statistic-Data
      wdi_stat = ReadDiskItem('Statistic-Data','NIFTI-1 image',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol})
      di_stat = list(wdi_stat.findValues({}, None, False ))

      #number of different subacquisition #not used for now, subacquisition has to be all different
      subacq_existing = [di_stat[i].attributes()['subacquisition'] for i in range(len(di_stat))]

      if len(subacq_existing) >0:
          subacq_stat=[]
          for i_subacq in range(len(subacq_existing)):
              subacq_stat.append(aims.read(di_stat[i_subacq].fileName()))

      plots = self.getAllPlotsCentersT1preRef() #coordonnées en mm à convertir en voxel
      if len(plots)==0:
          print("no contact found")
          return

      if not useTemplateMarsAtlas:
        vol_left = aims.read(diMaskleft.fileName())
        vol_right = aims.read(diMaskright.fileName())
      else:
        vol_left = aims.read('MNI_Brainvisa/t1mri/T1pre_1900-1-3/default_analysis/segmentation/mesh/surface_analysis/Gre_2016_MNI1_L_gyriVolume.nii.gz')
        vol_right = aims.read('MNI_Brainvisa/t1mri/T1pre_1900-1-3/default_analysis/segmentation/mesh/surface_analysis/Gre_2016_MNI1_R_gyriVolume.nii.gz')

      if not useTemplateFreeSurfer:
        freesurferdi = ReadDiskItem('FreesurferAtlas', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
        rdi_freesurfer = list(freesurferdi.findValues({}, None, False ))
        vol_freesurfer = aims.read(str(rdi_freesurfer[0]))
        freesurfHippoAntPostleft = ReadDiskItem('leftHippocampusNII', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
        freesurfHippoAntPostright = ReadDiskItem('rightHippocampusNII', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
        difreesurfHippoAntPostleft = list(freesurfHippoAntPostleft.findValues({}, None, False ))
        difreesurfHippoAntPostright = list(freesurfHippoAntPostright.findValues({}, None, False ))

        if len(difreesurfHippoAntPostright)>0:
            vol_hippoanteropostright = aims.read(str(difreesurfHippoAntPostright[0]))
            if len(difreesurfHippoAntPostleft)>0:
                vol_hippoanteropostleft = aims.read(str(difreesurfHippoAntPostleft[0]))
                vol_hippoanteropost = vol_hippoanteropostright + vol_hippoanteropostleft
            else:
                vol_hippoanteropost = vol_hippoanteropostright
        else:
            if len(difreesurfHippoAntPostleft)>0:
              vol_hippoanteropostleft = aims.read(str(difreesurfHippoAntPostleft[0]))
              vol_hippoanteropost = vol_hippoanteropostleft

        if len(difreesurfHippoAntPostleft) == 0 and len(difreesurfHippoAntPostright) == 0:
            vol_hippoanteropost = False

      else:
        vol_hippoanteropostleft = aims.read('MNI_Brainvisa/t1mri/T1pre_1900-1-3/default_analysis/segmentation/leftHippocampusGre_2016_MNI1.nii')
        vol_hippoanteropostright = aims.read('MNI_Brainvisa/t1mri/T1pre_1900-1-3/default_analysis/segmentation/rightHippocampusGre_2016_MNI1.nii')
        vol_hippoanteropost = vol_hippoanteropostright + vol_hippoanteropostleft
        vol_freesurfer = aims.read('MNI_Freesurfer/mri/freesurfer_parcelisation_mni2.nii')

      if not useTemplateHippoSubFreesurfer:
         HippoSubfielddi = ReadDiskItem('HippoFreesurferAtlas', 'BrainVISA volume formats', requiredAttributes={'center':self.currentProtocol, 'subject':self.brainvisaPatientAttributes['subject'] })
         rdi_HippoSubfield = list(HippoSubfielddi.findValues({},None,False))
         vol_hipposubfieldFS = aims.read(str(rdi_HippoSubfield[0]))
      else:
        vol_hipposubfieldFS = aims.read('MNI_Freesurfer/mri/bhHippoMNI.nii')

      #chargement des atlas dans le MNI (broadman, aal etc ...)
      vol_AAL = aims.read('MNI_Atlases/rAALSEEG12.nii')
      vol_AALDilate = aims.read('MNI_Atlases/rAALSEEG12Dilate.nii')
      vol_BroadmannDilate = aims.read('MNI_Atlases/rBrodmannSEEG3spm12.nii')
      vol_Broadmann = aims.read('MNI_Atlases/rbrodmann.nii')
      vol_Hammers = aims.read('MNI_Atlases/rHammersSEEG12.nii')
      
      info_image = self.diskItems['T1pre'].attributes() #['voxel_size'] #ca devrait etre les meme infos pour gauche et droite "probem when freesurfer is indi and mars atlas is template

      info_plot = []
      for k,v in plots.iteritems():
          plot_name_split = k.split('-$&_&$-')
          info_plot.append((plot_name_split[0]+plot_name_split[1][4:].zfill(2),v))
          #plots_label[k]=(label,label_name)

      plot_sorted = sorted(info_plot, key=lambda plot_number: plot_number[0])

      matrix_MNI_Nativ = numpy.matrix([[  -1.,    0.,    0.,   90.],[0.,   -1.,    0.,   91.],[0.,    0.,   -1.,  109.],[0.,    0.,    0.,    1.]])
      plot_dict_MNI_Native = {}
      for vv,kk in plot_dict_MNI.iteritems():
          inter_pos = [kk[0], kk[1], kk[2], 1]
          inter_pos = numpy.matrix(inter_pos).reshape([4,1])
          result_pos = numpy.dot(matrix_MNI_Nativ,inter_pos)
          plot_dict_MNI_Native.update({vv:[result_pos.tolist()[0][0],result_pos.tolist()[1][0],result_pos.tolist()[2][0]]})

      #montage bipolaire
      info_plot_bipolaire= []
      for pindex in range(1,len(plot_sorted)):
          previous_contact = "".join([i for i in plot_sorted[pindex-1][0] if not i.isdigit()])
          current_contact = "".join([i for i in plot_sorted[pindex][0] if not i.isdigit()])
          if previous_contact == current_contact:
               info_plot_bipolaire.append((plot_sorted[pindex][0]+' - '+ plot_sorted[pindex-1][0],(plot_sorted[pindex][1]+plot_sorted[pindex-1][1])/2 ))

      #if useTemplateMarsAtlas or useTemplateFreeSurfer or useTemplateHippoSubFreesurfer: #on fait le mni dans tous les cas à cause des atlas mni
      info_plot_bipolaire_MNI = {}
      for pindex in range(1,len(plot_sorted)):
         previous_contact = "".join([i for i in plot_sorted[pindex-1][0] if not i.isdigit()])
         current_contact = "".join([i for i in plot_sorted[pindex][0] if not i.isdigit()])
         if previous_contact == current_contact:
            info_plot_bipolaire_MNI.update({plot_sorted[pindex][0]+' - '+ plot_sorted[pindex-1][0]:(numpy.array(plot_dict_MNI_Native[plot_sorted[pindex][0]])+numpy.array(plot_dict_MNI_Native[plot_sorted[pindex-1][0]]))/2})


      parcels_names = readSulcusLabelTranslationFile('parcels_label_name.txt')
      freesurfer_parcel_names = readFreesurferLabelFile('freesurfer_label.txt')
      Hammers_parcels_names = readSulcusLabelTranslationFile('parcels_label_name_Hammers.txt')
      AAL_parcels_names = readSulcusLabelTranslationFile('parcels_label_name_AAL.txt')
      AALDilate_parcels_names = readSulcusLabelTranslationFile('parcels_label_name_AALDilate.txt')
      #parcels_AAL_names = 
      #parcels_Broadmann_names = 
      #parcels_Hammers_names = 

      #conversion x mm en nombre de voxel:
      sphere_size_stat = 10
      nb_voxel_sphere_stat = [int(round(sphere_size_stat/info_image['voxel_size'][i])) for i in range(3)]

      sphere_size = 3 #en mm
      nb_voxel_sphere = [int(round(sphere_size/info_image['voxel_size'][i])) for i in range(3)]
      #if useTemplateMarsAtlas or useTemplateFreeSurfer or useTemplateHippoSubFreesurfer:
      nb_voxel_sphere_MNI = [sphere_size, sphere_size, sphere_size] #because mni has a 1 mm isotropic resolution #int(round(sphere_size/info_image['voxel_size'][i])) for i in range(0,3)]

      print("start contact estimation")
      plots_label = {}
      for pindex in range(len(plot_sorted)):

         print(plot_sorted[pindex][0])
         #pdb.set_trace()
         plot_pos_pix_indi = [round(plot_sorted[pindex][1][i]/info_image['voxel_size'][i]) for i in range(3)]

         plot_pos_pix_MNI = [round(plot_dict_MNI_Native[plot_sorted[pindex][0]][i]) for i in range(3)]
         
         #mars atlas:
         if not useTemplateMarsAtlas:
             plot_pos_pixMA = plot_pos_pix_indi
             nb_voxel_sphereMA = nb_voxel_sphere
         elif useTemplateMarsAtlas:
             plot_pos_pixMA = plot_pos_pix_MNI #because MNI has a 1 mm istropic resolution #/info_image['voxel_size'][i]) for i in range(3)]
             nb_voxel_sphereMA = nb_voxel_sphere_MNI

         #on regarde si une sphère de x mm de rayon touche une parcel
         voxel_within_sphere_left = [vol_left.value(plot_pos_pixMA[0]+vox_i,plot_pos_pixMA[1]+vox_j,plot_pos_pixMA[2]+vox_k) for vox_k in range(-nb_voxel_sphereMA[2],nb_voxel_sphereMA[2]+1) for vox_j in range(-nb_voxel_sphereMA[1],nb_voxel_sphereMA[1]+1) for vox_i in range(-nb_voxel_sphereMA[0], nb_voxel_sphereMA[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_within_sphere_right = [vol_right.value(plot_pos_pixMA[0]+vox_i,plot_pos_pixMA[1]+vox_j,plot_pos_pixMA[2]+vox_k) for vox_k in range(-nb_voxel_sphereMA[2],nb_voxel_sphereMA[2]+1) for vox_j in range(-nb_voxel_sphereMA[1],nb_voxel_sphereMA[1]+1) for vox_i in range(-nb_voxel_sphereMA[0],nb_voxel_sphereMA[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]

         voxel_to_keep = [x for x in voxel_within_sphere_left+voxel_within_sphere_right if x != 0 and x !=255 and x != 100]
         
         if GWAtlas:
           voxelGW_within_sphere_left = [volGW_left.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
           voxelGW_within_sphere_right = [volGW_right.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]

           voxelGW_to_keep = [x for x in voxelGW_within_sphere_left+voxelGW_within_sphere_right if x !=255 and x !=0]
         else:
           GW_label = 255
         
         #freesurfer:
         if not useTemplateFreeSurfer:
             plot_pos_pixFS = plot_pos_pix_indi
             nb_voxel_sphereFS = nb_voxel_sphere
         elif useTemplateFreeSurfer:
             plot_pos_pixFS = plot_pos_pix_MNI  #I have to apply the transfo Scanner-Based to Native #because MNI has a 1 mm istropic resolution #/info_image['voxel_size'][i]) for i in range(3)]
             nb_voxel_sphereFS = nb_voxel_sphere_MNI

         voxel_within_sphere_FS = [vol_freesurfer.value(plot_pos_pixFS[0]+vox_i,plot_pos_pixFS[1]+vox_j,plot_pos_pixFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereFS[2],nb_voxel_sphereFS[2]+1) for vox_j in range(-nb_voxel_sphereFS[1],nb_voxel_sphereFS[1]+1) for vox_i in range(-nb_voxel_sphereFS[0],nb_voxel_sphereFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keep_FS = [x for x in voxel_within_sphere_FS if x != 0 and x != 2 and x != 41] #et 2 et 41 ? left and right white cerebral matter

         #HippoSubfield
         if not useTemplateHippoSubFreesurfer:
             plot_pos_pixHippoFS = plot_pos_pix_indi
             nb_voxel_sphereHippoFS = nb_voxel_sphere
         elif useTemplateHippoSubFreesurfer:
             plot_pos_pixHippoFS = plot_pos_pix_MNI
             nb_voxel_sphereHippoFS = nb_voxel_sphere_MNI

         #pdb.set_trace()
         voxel_within_sphere_HippoFS = [vol_hipposubfieldFS.value(plot_pos_pixHippoFS[0]+vox_i,plot_pos_pixHippoFS[1]+vox_j,plot_pos_pixHippoFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereHippoFS[2],nb_voxel_sphereHippoFS[2]+1) for vox_j in range(-nb_voxel_sphereHippoFS[1],nb_voxel_sphereHippoFS[1]+1) for vox_i in range(-nb_voxel_sphereHippoFS[0],nb_voxel_sphereHippoFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keep_HippoFS = [x for x in voxel_within_sphere_HippoFS if x != 0 and x != 2 and x != 41] #et 2 et 41 ? left and right white cerebral matter

         #MNI Atlases
         #AAL
         voxel_within_sphere_AAL = [round(vol_AAL.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepAAL = [x for x in voxel_within_sphere_AAL if x != 0 and not math.isnan(x)]
         
         #AALDilate
         voxel_within_sphere_AALdilate = [round(vol_AALDilate.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepAALDilate = [x for x in voxel_within_sphere_AALdilate if x != 0 and not math.isnan(x)]
         
         #Broadmann
         voxel_within_sphere_Broadmann = [round(vol_Broadmann.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepBroadmann = [x for x in voxel_within_sphere_Broadmann if x != 0 and not math.isnan(x)]
         
         #Brodmann dilate
         voxel_within_sphere_Broadmanndilate = [round(vol_BroadmannDilate.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepBroadmannDilate = [x for x in voxel_within_sphere_Broadmanndilate if x != 0 and not math.isnan(x)]
         
         #Hammers
         voxel_within_sphere_Hammers = [round(vol_Hammers.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepHammers = [x for x in voxel_within_sphere_Hammers if x != 0 and not math.isnan(x)]
         
         if DoResection:
            voxel_resec = [vol_resec.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]

         if len(subacq_existing)>0:
             voxel_substat=[]
             for i_substat in range(len(subacq_stat)):
                voxel_substat_all = [subacq_stat[i_substat].value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere_stat[2],nb_voxel_sphere_stat[2]+1) for vox_j in range(-nb_voxel_sphere_stat[1],nb_voxel_sphere_stat[1]+1) for vox_i in range(-nb_voxel_sphere_stat[0],nb_voxel_sphere_stat[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_stat]
                voxel_substat.append(numpy.mean([x for x in voxel_substat_all if abs(x) >= 0.1]))

         #prendre le label qui revient le plus en dehors de zero (au cas où il y en ait plusieurs)
         from collections import Counter

         if not voxel_to_keep:
            label_name = 'not in a mars atlas parcel'
            label = max(vol_left.value(plot_pos_pixMA[0],plot_pos_pixMA[1],plot_pos_pixMA[2]), vol_right.value(plot_pos_pixMA[0],plot_pos_pixMA[1],plot_pos_pixMA[2]))
            full_infoMAcomputed = []
         else:
            most_common,num_most_common = Counter(voxel_to_keep).most_common(1)[0]
            full_infoMA = Counter(voxel_to_keep).most_common()
            label = most_common
            full_infoMAcomputed = [(parcels_names[iLabel[0]],float(iLabel[1])/len(voxel_within_sphere_left)*100) for iLabel in full_infoMA]
            label_name = parcels_names[label]

         if not voxel_to_keep_FS:
             label_freesurfer_name = 'not in a freesurfer parcel'
             label_freesurfer = vol_freesurfer.value(plot_pos_pixFS[0],plot_pos_pixFS[1],plot_pos_pixFS[2])
         else:
             most_common,num_most_common = Counter(voxel_to_keep_FS).most_common(1)[0]
             #check if it's in the hippocampus
             if (most_common == 53 or most_common == 17) and vol_hippoanteropost != False:
               voxel_within_sphere_FS = [vol_hippoanteropost.value(plot_pos_pixFS[0]+vox_i,plot_pos_pixFS[1]+vox_j,plot_pos_pixFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereFS[2],nb_voxel_sphereFS[2]+1) for vox_j in range(-nb_voxel_sphereFS[1],nb_voxel_sphereFS[1]+1) for vox_i in range(-nb_voxel_sphereFS[0],nb_voxel_sphereFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
               voxel_to_keep_FS = [x for x in voxel_within_sphere_FS if x != 0 and x != 2 and x != 41]
               try:
                 most_common,num_most_common = Counter(voxel_to_keep_FS).most_common(1)[0]
               except:
                 pdb.set_trace()
               label_freesurfer = most_common
               if label_freesurfer == 3403:
                   pdb.set_trace()
               label_freesurfer_name = freesurfer_parcel_names[str(label_freesurfer)][0]
             else:
               label_freesurfer = most_common
               label_freesurfer_name = freesurfer_parcel_names[str(label_freesurfer)][0]

         if not voxel_to_keep_HippoFS:
             label_hippoFS_name = 'not in a hippocamp subfield'
             label_hippoFS = vol_hipposubfieldFS.value(plot_pos_pixHippoFS[0],plot_pos_pixHippoFS[1],plot_pos_pixHippoFS[2])
         else:
             most_common,num_most_common = Counter(voxel_to_keep_HippoFS).most_common(1)[0]
             label_hippoFS = most_common
             label_hippoFS_name = freesurfer_parcel_names[str(int(label_hippoFS))][0]

         if GWAtlas:
           if not voxelGW_to_keep:
              GW_label = max(volGW_left.value(plot_pos_pix_indi[0],plot_pos_pix_indi[1],plot_pos_pix_indi[2]), volGW_right.value(plot_pos_pix_indi[0],plot_pos_pix_indi[1],plot_pos_pix_indi[2]))
           else:
              most_common2,num_most_common2 = Counter(voxelGW_to_keep).most_common(1)[0]
              GW_label = most_common2

         if not voxel_to_keepAAL:
            label_AAL_name = "not in a AAL parcel" 
            label_AAL = round(vol_AAL.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepAAL).most_common(1)[0]
            label_AAL = most_common
            label_AAL_name = AAL_parcels_names[label_AAL]
             
         if not voxel_to_keepAALDilate:
            label_AALDilate_name = "not in a AALDilate parcel" 
            label_AALDilate = round(vol_AALDilate.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepAALDilate).most_common(1)[0]
            label_AALDilate = most_common
            label_AALDilate_name = AALDilate_parcels_names[label_AALDilate]
         
         if not voxel_to_keepBroadmann:
            label_Broadmann_name = "not in a Broadmann parcel" 
            label_Broadmann = round(vol_Broadmann.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepBroadmann).most_common(1)[0]
            label_Broadmann = most_common
            #label_Broadmann_name = unicode(label_Broadmann)
            if plot_pos_pix_MNI[0]>90:
                label_Broadmann_name = unicode(label_Broadmann+100)
            else:
                label_Broadmann_name = unicode(label_Broadmann)
                
         if not voxel_to_keepBroadmannDilate:
            label_BroadmannDilate_name = "not in a Broadmann parcel" 
            label_BroadmannDilate = round(vol_BroadmannDilate.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepBroadmannDilate).most_common(1)[0]
            label_BroadmannDilate = most_common
            #label_Broadmann_name = unicode(label_Broadmann)
            if plot_pos_pix_MNI[0]>90:
                label_BroadmannDilate_name = unicode(label_BroadmannDilate+100)
            else:
                label_BroadmannDilate_name = unicode(label_BroadmannDilate-48)       
             
         if not voxel_to_keepHammers:
            label_Hammers_name = "not in a Hammers parcel" 
            label_Hammers = round(vol_Hammers.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:    
            most_common,num_most_common = Counter(voxel_to_keepHammers).most_common(1)[0]
            label_Hammers = most_common
            label_Hammers_name = Hammers_parcels_names[label_Hammers]
                        
         if DoResection:
            most_common_res,num_most_common_res = Counter(voxel_resec).most_common(1)[0]
            Resec_label = max(voxel_resec) #most_common_res
         else:
            Resec_label = 255

         GW_label_name={0:'not in brain matter',100:'GreyMatter',200:'WhiteMatter',255:'Not Calculated'}[GW_label]
         Resec_label_name = {0:'not in resection',1:'in resection',255:'resection not calculated'}[Resec_label]

         plots_label[plot_sorted[pindex][0]]={'MarsAtlas':(label,label_name),'MarsAtlasFull':full_infoMAcomputed,'Freesurfer':(label_freesurfer,label_freesurfer_name),'Hippocampal Subfield':(label_hippoFS,label_hippoFS_name),'GreyWhite':(GW_label,GW_label_name),'AAL':(label_AAL,label_AAL_name),'AALDilate':(label_AALDilate,label_AALDilate_name),'Broadmann':(label_Broadmann,label_Broadmann_name), 'BroadmannDilate':(label_BroadmannDilate,label_BroadmannDilate_name),'Hammers':(label_Hammers,label_Hammers_name),'Resection':(Resec_label,Resec_label_name)}
         # add subacq_stat dictionnaries
         if len(subacq_existing)>0:
             for i_substat in range(len(subacq_stat)):
                 plots_label[plot_sorted[pindex][0]].update({di_stat[i_substat].attributes()['subacquisition']:float(format( voxel_substat[i_substat],'.3f'))})

      plot_name = [x[0] for x in plot_sorted]
      plots_by_label = dict([(Lab,[p for p in plot_name if plots_label[p]['MarsAtlas'][1]==Lab]) for Lab in parcels_names.values()])
      plots_by_label_FS = dict([(Lab,[p for p in plot_name if plots_label[p]['Freesurfer'][1]==Lab]) for Lab in [x[0] for x in freesurfer_parcel_names.values()]])
      plots_by_label_BM = dict([(Lab,[p for p in plot_name if plots_label[p]['Broadmann'][1]==Lab]) for Lab in [unicode("%1.1f"%x) for x in range(0,100)]])
      plots_by_label_HM = dict([(Lab,[p for p in plot_name if plots_label[p]['Hammers'][1]==Lab]) for Lab in Hammers_parcels_names.values()])
      plots_by_label_AAL = dict([(Lab,[p for p in plot_name if plots_label[p]['AAL'][1]==Lab]) for Lab in AAL_parcels_names.values()])
      plots_by_label_AALDilate = dict([(Lab,[p for p in plot_name if plots_label[p]['AALDilate'][1]==Lab]) for Lab in AALDilate_parcels_names.values()])

      print("start bipole estimation")
      #conversion x mm en nombre de voxel:
      sphere_size_bipole = 5
      nb_voxel_sphere = [int(round(sphere_size_bipole/info_image['voxel_size'][i])) for i in range(0,3)]
      #if useTemplateMarsAtlas or useTemplateFreeSurfer or useTemplateHippoSubFreesurfer:
      nb_voxel_sphere_MNI = [sphere_size_bipole, sphere_size_bipole, sphere_size_bipole] #[int(round(sphere_size_bipole/info_image['voxel_size'][i])) for i in range(0,3)]

      plots_label_bipolar = {}
      for pindex in range(len(info_plot_bipolaire)):
         plot_pos_pix_indi= [round(info_plot_bipolaire[pindex][1][i]/info_image['voxel_size'][i]) for i in range(3)]
         plot_pos_pix_MNI = [round(info_plot_bipolaire_MNI[info_plot_bipolaire[pindex][0]][i]) for i in range(3)]

         #on regarde si une sphère de x mm de rayon touche une parcel
         #mars atlas:
         if not useTemplateMarsAtlas:
             plot_pos_pixMA = plot_pos_pix_indi
             nb_voxel_sphereMA = nb_voxel_sphere
         elif useTemplateMarsAtlas:
             plot_pos_pixMA = plot_pos_pix_MNI #because MNI has a 1 mm istropic resolution #/info_image['voxel_size'][i]) for i in range(3)]
             nb_voxel_sphereMA = nb_voxel_sphere_MNI

         voxel_within_sphere_left = [vol_left.value(plot_pos_pixMA[0]+vox_i,plot_pos_pixMA[1]+vox_j,plot_pos_pixMA[2]+vox_k) for vox_k in range(-nb_voxel_sphereMA[2],nb_voxel_sphereMA[2]+1) for vox_j in range(-nb_voxel_sphereMA[1],nb_voxel_sphereMA[1]+1) for vox_i in range(-nb_voxel_sphereMA[0],nb_voxel_sphereMA[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_bipole]
         voxel_within_sphere_right = [vol_right.value(plot_pos_pixMA[0]+vox_i,plot_pos_pixMA[1]+vox_j,plot_pos_pixMA[2]+vox_k) for vox_k in range(-nb_voxel_sphereMA[2],nb_voxel_sphereMA[2]+1) for vox_j in range(-nb_voxel_sphereMA[1],nb_voxel_sphereMA[1]+1) for vox_i in range(-nb_voxel_sphereMA[0],nb_voxel_sphereMA[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_bipole]

         voxel_to_keep = [x for x in voxel_within_sphere_left+voxel_within_sphere_right if x != 0 and x !=255 and x != 100]

         if GWAtlas:
            voxelGW_within_sphere_left = [volGW_left.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_bipole]
            voxelGW_within_sphere_right = [volGW_right.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_bipole]

            voxelGW_to_keep = [x for x in voxelGW_within_sphere_left+voxelGW_within_sphere_right if x !=255 and x !=0]
         else:
            GW_label = 255

         #freesurfer:
         if not useTemplateFreeSurfer:
             plot_pos_pixFS = plot_pos_pix_indi
             nb_voxel_sphereFS = nb_voxel_sphere
         elif useTemplateFreeSurfer:
             plot_pos_pixFS = plot_pos_pix_MNI #because MNI has a 1 mm istropic resolution #/info_image['voxel_size'][i]) for i in range(3)]
             nb_voxel_sphereFS = nb_voxel_sphere_MNI

         voxel_within_sphere_FS = [vol_freesurfer.value(plot_pos_pixFS[0]+vox_i,plot_pos_pixFS[1]+vox_j,plot_pos_pixFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereFS[2],nb_voxel_sphereFS[2]+1) for vox_j in range(-nb_voxel_sphereFS[1],nb_voxel_sphereFS[1]+1) for vox_i in range(-nb_voxel_sphereFS[0],nb_voxel_sphereFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keep_FS = [x for x in voxel_within_sphere_FS if x != 0 and x != 2 and x != 41]

         #HippoSubfield
         if not useTemplateHippoSubFreesurfer:
             plot_pos_pixHippoFS = plot_pos_pix_indi
             nb_voxel_sphereHippoFS = nb_voxel_sphere
         elif useTemplateHippoSubFreesurfer:
             plot_pos_pixHippoFS = plot_pos_pix_MNI
             nb_voxel_sphereHippoFS = nb_voxel_sphere_MNI

         voxel_within_sphere_HippoFS = [vol_hipposubfieldFS.value(plot_pos_pixHippoFS[0]+vox_i,plot_pos_pixHippoFS[1]+vox_j,plot_pos_pixHippoFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereHippoFS[2],nb_voxel_sphereHippoFS[2]+1) for vox_j in range(-nb_voxel_sphereHippoFS[1],nb_voxel_sphereHippoFS[1]+1) for vox_i in range(-nb_voxel_sphereHippoFS[0],nb_voxel_sphereHippoFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keep_HippoFS = [x for x in voxel_within_sphere_HippoFS if x != 0 and x != 2 and x != 41] #et 2 et 41 ? left and right white cerebral matter

         #MNI Atlases
         #AAL
         voxel_within_sphere_AAL = [round(vol_AAL.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepAAL = [x for x in voxel_within_sphere_AAL if x != 0 and not math.isnan(x)]
         
         #AALDilate
         voxel_within_sphere_AALdilate = [round(vol_AALDilate.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepAALDilate = [x for x in voxel_within_sphere_AALdilate if x != 0 and not math.isnan(x)]
         
         #Broadmann
         voxel_within_sphere_Broadmann = [round(vol_Broadmann.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepBroadmann = [x for x in voxel_within_sphere_Broadmann if x != 0 and not math.isnan(x)]
         
         #Brodmann dilate
         voxel_within_sphere_Broadmanndilate = [round(vol_BroadmannDilate.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepBroadmannDilate = [x for x in voxel_within_sphere_Broadmanndilate if x != 0 and not math.isnan(x)]
         
         #Hammers
         voxel_within_sphere_Hammers = [round(vol_Hammers.value(plot_pos_pix_MNI[0]+vox_i,plot_pos_pix_MNI[1]+vox_j,plot_pos_pix_MNI[2]+vox_k)) for vox_k in range(-nb_voxel_sphere_MNI[2],nb_voxel_sphere_MNI[2]+1) for vox_j in range(-nb_voxel_sphere_MNI[1],nb_voxel_sphere_MNI[1]+1) for vox_i in range(-nb_voxel_sphere_MNI[0],nb_voxel_sphere_MNI[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
         voxel_to_keepHammers = [x for x in voxel_within_sphere_Hammers if x != 0 and not math.isnan(x)]

         
         if DoResection:
            voxel_resec = [vol_resec.value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_bipole]

         if len(subacq_existing)>0:
             voxel_substat=[]
             for i_substat in range(len(subacq_stat)):
                voxel_substat_all = [subacq_stat[i_substat].value(plot_pos_pix_indi[0]+vox_i,plot_pos_pix_indi[1]+vox_j,plot_pos_pix_indi[2]+vox_k) for vox_k in range(-nb_voxel_sphere_stat[2],nb_voxel_sphere_stat[2]+1) for vox_j in range(-nb_voxel_sphere_stat[1],nb_voxel_sphere_stat[1]+1) for vox_i in range(-nb_voxel_sphere_stat[0],nb_voxel_sphere_stat[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size_stat]
                voxel_substat.append(numpy.mean([x for x in voxel_substat_all if abs(x) >= 0.1]))

         #prendre le label qui revient le plus en dehors de zero (au cas où il y en ait plusieurs)
         from collections import Counter

         if not voxel_to_keep:
            label_name = 'not in a mars atlas parcel'
            label = max(vol_left.value(plot_pos_pixMA[0],plot_pos_pixMA[1],plot_pos_pixMA[2]), vol_right.value(plot_pos_pixMA[0],plot_pos_pixMA[1],plot_pos_pixMA[2]))
            full_infoMAcomputed = []
         else:
             most_common,num_most_common = Counter(voxel_to_keep).most_common(1)[0]
             label = most_common
             full_infoMA = Counter(voxel_to_keep).most_common()
             full_infoMAcomputed = [(parcels_names[iLabel[0]],float(iLabel[1])/len(voxel_within_sphere_left)*100) for iLabel in full_infoMA]
             label_name = parcels_names[label]

         if not voxel_to_keep_FS:
             label_freesurfer_name = 'not in a freesurfer parcel'
             label_freesurfer = vol_freesurfer.value(plot_pos_pixFS[0],plot_pos_pixFS[1],plot_pos_pixFS[2])
         else:
             most_common,num_most_common = Counter(voxel_to_keep_FS).most_common(1)[0]
             #check if it's in the hippocampus
             if (most_common == 53 or most_common == 17) and vol_hippoanteropost != False:
               voxel_within_sphere_FS = [vol_hippoanteropost.value(plot_pos_pixFS[0]+vox_i,plot_pos_pixFS[1]+vox_j,plot_pos_pixFS[2]+vox_k) for vox_k in range(-nb_voxel_sphereFS[2],nb_voxel_sphereFS[2]+1) for vox_j in range(-nb_voxel_sphereFS[1],nb_voxel_sphereFS[1]+1) for vox_i in range(-nb_voxel_sphereFS[0],nb_voxel_sphereFS[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]
               voxel_to_keep_FS = [x for x in voxel_within_sphere_FS if x != 0 and x != 2 and x != 41]
               most_common,num_most_common = Counter(voxel_to_keep_FS).most_common(1)[0]
               label_freesurfer = most_common
               label_freesurfer_name = freesurfer_parcel_names[str(label_freesurfer)][0]
             else:
               label_freesurfer = most_common
               label_freesurfer_name = freesurfer_parcel_names[str(label_freesurfer)][0]

         if not voxel_to_keep_HippoFS:
             label_hippoFS_name = 'not in a hippocamp subfield'
             label_hippoFS = vol_hipposubfieldFS.value(plot_pos_pixHippoFS[0],plot_pos_pixHippoFS[1],plot_pos_pixHippoFS[2])
         else:
             most_common,num_most_common = Counter(voxel_to_keep_HippoFS).most_common(1)[0]
             label_hippoFS = most_common
             label_hippoFS_name = freesurfer_parcel_names[str(int(label_hippoFS))][0]

         if GWAtlas:
           if not voxelGW_to_keep:
               GW_label = max(volGW_left.value(plot_pos_pix_indi[0],plot_pos_pix_indi[1],plot_pos_pix_indi[2]), volGW_right.value(plot_pos_pix_indi[0],plot_pos_pix_indi[1],plot_pos_pix_indi[2]))
           else:
               most_common2,num_most_common2 = Counter(voxelGW_to_keep).most_common(1)[0]
               GW_label = most_common2

         if not voxel_to_keepAAL:
            label_AAL_name = "not in a AAL parcel" 
            label_AAL = round(vol_AAL.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepAAL).most_common(1)[0]
            label_AAL = most_common
            label_AAL_name = AAL_parcels_names[label_AAL]
             
         if not voxel_to_keepAALDilate:
            label_AALDilate_name = "not in a AALDilate parcel" 
            label_AALDilate = round(vol_AALDilate.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepAALDilate).most_common(1)[0]
            label_AALDilate = most_common              
            label_AALDilate_name = AALDilate_parcels_names[label_AALDilate]
         
         if not voxel_to_keepBroadmann:
            label_Broadmann_name = "not in a Broadmann parcel" 
            label_Broadmann = round(vol_Broadmann.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepBroadmann).most_common(1)[0]
            label_Broadmann = most_common
            #label_Broadmann_name = unicode(label_Broadmann)
            if plot_pos_pix_MNI[0]>90:
                label_Broadmann_name = unicode(label_Broadmann+100)
            else:
                label_Broadmann_name = unicode(label_Broadmann)
             
         if not voxel_to_keepBroadmannDilate:
            label_BroadmannDilate_name = "not in a Broadmann parcel" 
            label_BroadmannDilate = round(vol_BroadmannDilate.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:
            most_common,num_most_common = Counter(voxel_to_keepBroadmannDilate).most_common(1)[0]
            label_BroadmannDilate = most_common
            #label_Broadmann_name = unicode(label_Broadmann)
            if plot_pos_pix_MNI[0]>90:
                label_BroadmannDilate_name = unicode(label_BroadmannDilate+100)
            else:
                label_BroadmannDilate_name = unicode(label_BroadmannDilate-48) 
         
         if not voxel_to_keepHammers:
            label_Hammers_name = "not in a Hammers parcel" 
            label_Hammers = round(vol_Hammers.value(plot_pos_pix_MNI[0],plot_pos_pix_MNI[1],plot_pos_pix_MNI[2]))
         else:    
            most_common,num_most_common = Counter(voxel_to_keepHammers).most_common(1)[0]
            label_Hammers = most_common
            label_Hammers_name = Hammers_parcels_names[label_Hammers]

         
         if DoResection:
            most_common_res,num_most_common_res = Counter(voxel_resec).most_common(1)[0]
            Resec_label = max(voxel_resec) #most_common_res
         else:
            Resec_label = 255

         GW_label_name={0:'not in brain matter',100:'GreyMatter',200:'WhiteMatter',255:'Not Calculated'}[GW_label]
         Resec_label_name = {0:'not in resection',1:'in resection',255:'resection not calculated'}[Resec_label]

         plots_label_bipolar[info_plot_bipolaire[pindex][0]]={'MarsAtlas':(label,label_name),'MarsAtlasFull':full_infoMAcomputed,'Freesurfer':(label_freesurfer,label_freesurfer_name),'Hippocampal Subfield':(label_hippoFS,label_hippoFS_name),'GreyWhite':(GW_label,GW_label_name),'AAL':(label_AAL,label_AAL_name),'AALDilate':(label_AALDilate,label_AALDilate_name),'Broadmann':(label_Broadmann,label_Broadmann_name),'BroadmannDilate':(label_BroadmannDilate,label_BroadmannDilate_name),'Hammers':(label_Hammers,label_Hammers_name),'Resection':(Resec_label,Resec_label_name)}
         #plots_label_bipolar.append((info_plot_bipolaire[pindex][0],label,label_name,GW_label))
         # add subacq_stat dictionnaries
         if len(subacq_existing)>0:
             for i_substat in range(len(subacq_stat)):
                 plots_label_bipolar[info_plot_bipolaire[pindex][0]].update({di_stat[i_substat].attributes()['subacquisition']:float(format( voxel_substat[i_substat],'.3f'))})

      plot_name_bip = [x[0] for x in info_plot_bipolaire]
      #plots_by_label = {Lab:[p for p in plot_name if plots_label[p]['MarsAtlas'][1]==Lab] for Lab in parcels_names.values()}
      plots_bipolar_by_label = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['MarsAtlas'][1]==Lab]) for Lab in parcels_names.values()])
      #do the same for freesurfer, broadmann, hammers, all and alldilate
      plots_bipolar_by_label_FS = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['Freesurfer'][1]==Lab]) for Lab in [x[0] for x in freesurfer_parcel_names.values()]])
      plots_bipolar_by_label_BM = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['Broadmann'][1]==Lab]) for Lab in [unicode("%1.1f"%x) for x in range(0,100)]])
      plots_bipolar_by_label_HM = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['Hammers'][1]==Lab]) for Lab in Hammers_parcels_names.values()])
      plots_bipolar_by_label_AAL = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['AAL'][1]==Lab]) for Lab in AAL_parcels_names.values()])
      plots_bipolar_by_label_AALDilate = dict([(Lab,[p for p in plot_name_bip if plots_label_bipolar[p]['AALDilate'][1]==Lab]) for Lab in AALDilate_parcels_names.values()])
      wdi = WriteDiskItem('Electrodes Labels','Electrode Label Format')
      di = wdi.findValue(self.diskItems['T1pre'])
      if di is None:
         print('Can t generate files')
         return

      UseTemplateOrPatient = {'MarsAtlas':useTemplateMarsAtlas,'Freesurfer':useTemplateFreeSurfer,'HippocampalSubfield Freesurfer':useTemplateHippoSubFreesurfer}

      fout = open(di.fullPath(),'w')
      fout.write(json.dumps({'Template':UseTemplateOrPatient,'plots_label':plots_label,'plots_by_label':plots_by_label, 'plots_by_label_FS':plots_by_label_FS, 'plots_by_label_BM':plots_by_label_BM, 'plots_by_label_HM':plots_by_label_HM, 'plots_by_label_AAL':plots_by_label_AAL, 'plots_by_label_AALDilate':plots_by_label_AALDilate, 'plots_label_bipolar':plots_label_bipolar, 'plots_bipolar_by_label':plots_bipolar_by_label, 'plots_bipolar_by_label_FS':plots_bipolar_by_label_FS, 'plots_bipolar_by_label_BM':plots_bipolar_by_label_BM, 'plots_bipolar_by_label_HM':plots_bipolar_by_label_HM, 'plots_bipolar_by_label_AAL':plots_bipolar_by_label_AAL, 'plots_bipolar_by_label_AALDilate':plots_bipolar_by_label_AALDilate}))
      fout.close()

      neuroHierarchy.databases.insertDiskItem(di, update=True )
      #fin = open(di.fullPath(),'r')
      #dictée = json.loads(fin.read())
      #dictée['plots_label']

      print "export electrode done"



  def getAllPlotsCentersT1preRef(self):
    """Return a dictionary {'ElectrodeName-$&_&$-PlotName':[x,y,z], ...} where x,y,z is in the T1pre native referential"""
    return dict((el['name']+'-$&_&$-'+plotName, el['transf'].transform(plotCoords)) for el in self.electrodes for plotName, plotCoords in getPlotsCenters(el['elecModel']).iteritems())

  def getAllPlotsCentersT1preScannerBasedRef(self):
    """Return a dictionary {'ElectrodeName-$&_&$-PlotName':[x,y,z], ...} where x,y,z is in the T1pre scanner-based referential"""
    transfo = self.t1pre2ScannerBased()
    return dict((key, transfo.transform(coords)) for key, coords in self.getAllPlotsCentersT1preRef().iteritems())

  def getAllPlotsCentersAnyReferential(self, referential):
    """Return a dictionary {'ElectrodeName-$&_&$-PlotName':[x,y,z], ...} where x,y,z is in the provided referential (must be available in referential converter self.refConv)"""
    if not self.refConv.isRefAvailable(referential):
      print "Trying to convert to unknown referential %s"%repr(referential)
      return None
    return dict((key, self.refConv.real2AnyRef(coords, referential)) for key, coords in self.getAllPlotsCentersT1preRef().iteritems())

  def getAllPlotsCentersMNIRef(self, plots=None):
    """Return a dictionary {'ElectrodeName-$&_&$-PlotName':[x,y,z], ...} where x,y,z is in the MNI referential"""
    if plots is None:
      plots = self.getAllPlotsCentersT1preScannerBasedRef()
    coords = [plots[k] for k in sorted(plots.keys())]
    newCoords = self.convertT1ScannerBasedToMni(coords)
    if newCoords is None:
      return None
    return dict(zip(sorted(plots.keys()), newCoords))

  def saveTXT(self, contacts=None, path=None, pathPos=None, pathName=None):
    """ Saves two txt files electrode_Name.txt and electrode_Pos.txt. Path should be supplied as /myPath/electrode.txt
       or as an alternative, both file path should be supplied as pathPos and pathName
       contacts should be a dictionary {'ElectrodeName-$&_&$-PlotName:[x,y,z],...} in the appropriate referential
     """
    # Get a valid path
    if not path and not (pathPos is not None and pathName is not None):
      path = str(QtGui.QFileDialog.getSaveFileName(self, "Save TXT files", "", "Electrode implantation TXT files (*.txt)"))
      if not path:
        return None
    if not contacts:
      return None
    if path is not None:
      path = os.path.splitext(path)[0]
      pathName = path+'_Name.txt'
      pathPos = path+"_Pos.txt"
    fileName = open(pathName, 'wb')
    filePos = open(pathPos, 'wb')
    # Iteration over electrodes
    for contactName in sorted(contacts.keys(), key=natural_keys):
      (elName, plotName) = contactName.split('-$&_&$-')
      # Name of each contact is name of electrode (p for prime) + number of the plot (electrode A' contact 5 is "Ap5")
      fileName.write(elName.replace('\'', 'p') + plotName.replace('Plot','') + "\n")
      filePos.write("%f %f %f\n"%tuple(contacts[contactName]))
    fileName.close()
    filePos.close()
    return pathPos

  def savePTS(self, contacts=None, path=None):
    """ Save a PTS file with all contacts coordinates.
       contacts parameter should be a dictionary {'ElectrodeName-$&_&$-PlotName:[x,y,z],...} in the appropriate referential
    """
    # Get a valid path
    if not path:
      path = str(QtGui.QFileDialog.getSaveFileName(self, "Save PTS file (%s referential)"%referential, "", "PTS Electrode implantation (*.pts)"))
      if not path:
        return None
    if contacts is None:
      return None

    plots=[]
    for contactName in sorted(contacts.keys(), key=natural_keys):
      coords = contacts[contactName]
      (elName, plotName) = contactName.split('-$&_&$-')
        # Name of each contact is name of electrode (prime ' replaced by the letter p) + number of the plot (electrode A' contact 5 is "Ap5")
      plots.append("%s\t%.2f\t%.2f\t%.2f\t0\t0\t0\t2\t2.0\n"%(elName.replace('\'', 'p') + plotName.replace('Plot',''), coords[0], coords[1], coords[2]))

    fileout = open(path, 'wb')
    fileout.write("ptsfile\n1\t1\t1\n%s\n"%str(len(plots)))
    for p in plots:
      fileout.write(p)

    fileout.close()
    return path

  # Open configuration dialog
  def configureColors(self):
    
    self.bipoleSEEGColors=bipoleSEEGColors(self)
    self.bipoleSEEGColors.show()
    
    

  def allWindowsUpdate(self):
    self.windowUpdate(0, self.windowCombo1.currentText())
    self.windowUpdate(1, self.windowCombo2.currentText())

  # Update a window content
  def windowUpdate(self, winId, key):
    key = str(key) # Can be a QString !
    if key not in self.windowContent:
        return #when QT interface but, there was a variable generated in "frame" and we were not able to delete it, then we were not able to load another patient, there was leftover from previous patient
    w = self.wins[winId]
    for obj in self.dispObj:
      if obj in self.windowContent[key]:
        #print "Adding %s"%obj
        self.a.addObjects(self.dispObj[obj], w)
      else:
        #print "Removing %s"%obj
        self.a.removeObjects([self.dispObj[obj],],w)#CURRENT

  def updateElectrodeView(self, checkStatus=None):
    """Sets the view to electrode referential or back to native T1 referential
    checkStatus is Qt.CheckState"""
    if checkStatus is None:
      checkStatus = self.electrodeRefCheck.checkState()
    if checkStatus == QtCore.Qt.Checked:
      if self.electrodeList.count() > 0 and self.electrodeList.currentRow() <= len(self.electrodes):
        #Change to electrode referential
        el = self.currentElectrode()
        if el is None:
          return
        if 'refRotation' in el:
          self.electrodeRefRotationSlider.setValue(el['refRotation'])
        if 'ref' in el:
          self.setWindowsReferential(el['ref'])
          self.electrodeGo(electrode = el)
    else:
      # Change back to T1pre native ref
      self.setWindowsReferential()

  def updateElectrodeViewRotation(self, degrees):
    """Sets the angle of an electrode referential, degrees is the angle in degrees"""
    if self.electrodeList.count() > 0 and self.electrodeList.currentRow() <= len(self.electrodes):
        #Change angle of electrode referential
        el = self.currentElectrode()
        if el:
          el['refRotation'] = degrees

  def clippingUpdate(self):

      if self.Clipping_checkbox.isChecked():
         print "clipping activated"
         self.a.execute('WindowConfig',windows = [self.wins[0]],clipping=2,clip_distance=5.)

      else:
         print "clipping not activated"
         self.a.execute('WindowConfig',windows = [self.wins[0]],clipping=0)


  
  def makeFusion(self):

      #get objects
      Text_win1 = self.windowCombo1.currentText()
      Text_win2 = self.windowCombo2.currentText()

      for obj in self.dispObj.keys():
        if obj in self.windowContent[str(Text_win1)][0]:
        #print "Adding %s"%obj
           obj1 = obj

      for obj in self.dispObj.keys():
        if obj in self.windowContent[str(Text_win2)][0]:
        #print "Adding %s"%obj
           obj2 = obj

      #pdb.set_trace()
      if 'obj1' in locals() and 'obj2' in locals():
          print "do the fusion"
          fusion_obj = self.a.fusionObjects((self.dispObj[obj1], self.dispObj[obj2]), method='Fusion2DMethod')
          self.a.addObjects(fusion_obj, self.wins[1])

          #add the fusion in the disObj and the windowCombo
          self.dispObj[obj1+'+'+obj2] = fusion_obj

          self.windowContent.update({obj1+'+'+obj2:[obj1+'+'+obj2,'electrodes']})
          self.windowCombo1.clear()
          self.windowCombo1.addItems(sorted(self.windowContent.keys()))
          self.windowCombo2.clear()
          self.windowCombo2.addItems(sorted(self.windowContent.keys()))

          self.windowCombo1.setCurrentIndex(max(self.windowCombo1.findText(Text_win1),0))
          self.windowCombo2.setCurrentIndex(max(self.windowCombo2.findText(obj1+'+'+obj2),0))

          #self.allWindowsUpdate()

      else:
         print "one of the image is not recognized"
         return


  def generateResection(self):

      #look for T1 pre and T1 postop
      T1 = ReadDiskItem('Raw T1 MRI', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
      diT1 = list(T1.findValues({}, None, False ))
      for t in diT1:
          if 'pre' in t.attributes()['acquisition']:
              T1pre = t.fullPath()
              try:
                SB_transf = t.attributes()['SB_Transform']
              except:
                SB_transf = t.attributes()['transformations'][0]

          elif 'postOp' in t.attributes()['acquisition']:
              T1postop = t.fullPath()

      if not T1pre:
          print('don t find the t1pre')
          return
      
      if 'T1postop' not in locals():
          print('don t find the t1postop')
          print('look for a CTpostop')
          CTs = ReadDiskItem('CT', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
          diCTs = list(CTs.findValues({},None,False))
          for t in diCTs:
              if 'postOp' in t.attributes()['acquisition']:
                  CTpostop = t.fullPath()
                  method = 'CT'
          if 'CTpostop' not in locals():
              print("can't find a CTpostop either")
              return
      else:
          method = 'T1'

      T1pre = None
      T1postop = None
      
      id_pre = [x for x in range(len(diT1)) if 'pre' in str(diT1[x])]
      T1pre = diT1[id_pre[0]]
      
      if method == 'T1':          
         id_postop = [x for x in range(len(diT1)) if 'postOp' in str(diT1[x])]
         T1postop = diT1[id_postop[0]]
      elif method == 'CT': 
         id_ctpostop = [x for x in range(len(diCTs)) if 'postOp' in str(diCTs[x])]
         CTpostop = diCTs[id_ctpostop[0]]


      if method == 'T1':

        try:
           self.refConv.loadACPC(T1pre)
        except Exception, e:
           print "Cannot load AC-PC referential from T1 pre MRI : "+repr(e)
           return

        Ac = self.refConv.Ac
        Pc = self.refConv.Pc
        Ih = self.refConv.Ih
        Ac.append(1)
        Pc.append(1)
        Ih.append(1)

        Ac_vector = numpy.array([Ac])
        Pc_vector = numpy.array([Pc])
        Ih_vector = numpy.array([Ih])

        wdiTransform = ReadDiskItem('Transform Raw T1 MRI to another image', 'Transformation matrix', exactType=True, requiredAttributes = {'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol }) #pourquoi la suite marche pas ?, requiredAttributes = {'modalityTarget':T1pre.attributes()['modality'], 'acquisitionTarget':T1pre.attributes()['acquisition']}
        diTransform = list(wdiTransform.findValues({}, None, False ))

        for t in diTransform:
          if t.attributes()['modality'] == 't1mri' and 'postOp' in t.attributes()['acquisition']:
             trmpostop_to_pre = t
             #transfo_postop_to_pre = aims.read(trmpostop_to_pre.fullPath()).toMatrix()
             #transfo_pre_to_postop = numpy.linalg.inv(transfo_postop_to_pre)

        wdiTransform2 = ReadDiskItem('Transformation to Scanner Based Referential', 'Transformation matrix', exactType=True, requiredAttributes = {'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diTransform2 = wdiTransform2.findValues({}, None, False )

        for t in diTransform2:
          if t.attributes()['modality'] == 't1mri' and 'postOp' in t.attributes()['acquisition']:
             trmpostop_to_SB = t
             #transfo_postop_to_SB = aims.read(trmpostop_to_SB.fullPath()).toMatrix()
          if t.attributes()['modality'] == 't1mri' and 'pre' in t.attributes()['acquisition']:
             trmpre_to_SB = t
             #transfo_pre_to_SB = aims.read(trmpre_to_SB.fullPath()).toMatrix()

      if method == 'CT':
          
        wdiTransform = ReadDiskItem('Transform CT to another image', 'Transformation matrix', exactType=True, requiredAttributes = {'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol }) #pourquoi la suite marche pas ?, requiredAttributes = {'modalityTarget':T1pre.attributes()['modality'], 'acquisitionTarget':T1pre.attributes()['acquisition']}
        diTransform = list(wdiTransform.findValues({}, None, False ))

        for t in diTransform:
          if t.attributes()['modality'] == 'ct' and 'postOp' in t.attributes()['acquisition']:
             trmpostop_to_pre = t
             
        wdiTransform2 = ReadDiskItem('Transformation to Scanner Based Referential', 'Transformation matrix', exactType=True, requiredAttributes = {'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diTransform2 = wdiTransform2.findValues({}, None, False )

        for t in diTransform2:
          if t.attributes()['modality'] == 'ct' and 'postOp' in t.attributes()['acquisition']:
             trmpostop_to_SB = t
             #transfo_postop_to_SB = aims.read(trmpostop_to_SB.fullPath()).toMatrix()
          if t.attributes()['modality'] == 't1mri' and 'pre' in t.attributes()['acquisition']:
             trmpre_to_SB = t
             #transfo_pre_to_SB = aims.read(trmpre_to_SB.fullPath()).toMatrix()
   



      import copy
      trmpre_to_SBinvpath = trmpre_to_SB.fullPath().split('/')
      trmpre_to_SBinvpath[-1] = 'inv'+trmpre_to_SBinvpath[-1]
      trmpostop_to_pre_path = copy.deepcopy(trmpre_to_SBinvpath)
      trmpostop_to_pre_path[-1] = 'postop_to_pre.trm'
      trmpre_to_postop_path = copy.deepcopy(trmpre_to_SBinvpath)
      trmpre_to_postop_path[-1] = 'pre_to_postop.trm'
      trmpre_to_SBinvpath = '/'.join(trmpre_to_SBinvpath)
      trmpostop_to_pre_path = '/'.join(trmpostop_to_pre_path)
      trmpre_to_postop_path = '/'.join(trmpre_to_postop_path)


      ret = subprocess.call(['AimsInvertTransformation','-i',trmpre_to_SB.fullPath(),'-o', trmpre_to_SBinvpath])
      ret = subprocess.call(['AimsComposeTransformation', '-o',trmpostop_to_pre_path, trmpre_to_SBinvpath, trmpostop_to_pre.fullPath(), trmpostop_to_SB.fullPath()])
      ret = subprocess.call(['AimsInvertTransformation','-i',trmpostop_to_pre_path,'-o',trmpre_to_postop_path])

      print 'select the center of the resection'
      center_seg = QtGui.QMessageBox(self)
      center_seg.setText("Enter the center (approximatly) of the resection")
      center_seg.setWindowTitle("resection center")
      center_seg.setWindowModality(QtCore.Qt.NonModal)
      center_seg.show()
      center_seg.exec_()

      #convert brainvisa voxel position into spm_voxel position
      ResecCenterCoord = list(self.a.linkCursorLastClickedPosition(self.preReferential()).items())
      print ResecCenterCoord

      if method == 'T1':
        transfo_pre_to_postop = aims.read(trmpre_to_postop_path).toMatrix()

        Ac_vector_postop = transfo_pre_to_postop.dot(Ac_vector.T)
        Pc_vector_postop = transfo_pre_to_postop.dot(Pc_vector.T)
        Ih_vector_postop = transfo_pre_to_postop.dot(Ih_vector.T)
        vect1 = numpy.array(Ac_vector_postop[0:3])-numpy.array(Pc_vector_postop[0:3])
        vect2 = numpy.array(Ih_vector_postop[0:3])-numpy.array(Pc_vector_postop[0:3])

        result_cross = cross(vect1.T.tolist(),vect2.T.tolist())/numpy.linalg.norm(cross(vect1.T.tolist(),vect2.T.tolist()))*40
        Lh_postop = numpy.array(Ac[0:3]) + result_cross

        brainvisaContext = defaultContext()
        morphologist = processes.getProcessInstance('morphologist')
        morphologist.executionNode().PrepareSubject.setSelected(True)
        morphologist.executionNode().BiasCorrection.setSelected(True)
        morphologist.executionNode().HistoAnalysis.setSelected(True)
        morphologist.executionNode().BrainSegmentation.setSelected(True)
        morphologist.executionNode().Renorm.setSelected(False)
        morphologist.executionNode().SplitBrain.setSelected(False)
        morphologist.executionNode().TalairachTransformation.setSelected(False)
        morphologist.executionNode().HeadMesh.setSelected(False)
        morphologist.executionNode().HemispheresProcessing.setSelected(False)
        morphologist.executionNode().SulcalMorphometry.setSelected(False)
        brainvisaContext.runInteractiveProcess(lambda x='',trm=trmpostop_to_pre_path,resec_coord=ResecCenterCoord,methodo=method:self.resectionStart(trm,resec_coord,methodo) , morphologist, t1mri = T1postop, perform_normalization = False, anterior_commissure = Ac_vector_postop[0:3].T.tolist()[0],\
	                           posterior_commissure = Pc_vector_postop[0:3].T.tolist()[0], interhemispheric_point = Ih_vector_postop[0:3].T.tolist()[0], left_hemisphere_point = Lh_postop.tolist()[0], perform_sulci_recognition = False)
        
      if method == 'CT':
          self.resectionStart(trmpostop_to_pre_path,ResecCenterCoord,method = 'CT')


  def generateFiberContactDistance(self):
      
      print("not finished")
      return
      pdb.set_trace()
  
  
  def resectionStart(self,trm_postop_to_pre,resec_coord,method = 'T1'):

      wdi_resec = WriteDiskItem('Resection', 'NIFTI-1 image')
      di_resec = wdi_resec.findValue({'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol, 'acquisition':'Resection'})

      #if the path doesn't exist, create it
      if not os.path.exists(os.path.dirname(di_resec.fullPath())):
          os.makedirs( os.path.dirname(di_resec.fullPath()))
      
      
      if method == 'T1':
        brainMask = ReadDiskItem('Brain Mask', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diBrain = list(brainMask.findValues({}, None, False ))

        id_pre = [x for x in range(len(diBrain)) if 'pre' in str(diBrain[x])]
        id_postop = [x for x in range(len(diBrain)) if 'postOp' in str(diBrain[x])]

        fullname_postop = diBrain[id_postop[0]].fullPath()
        fullpost_split = fullname_postop.split('/')
        fullpost_split[-1] = 'brainpostop_on_pre.nii'
        fullpost = '/'.join(fullpost_split)

        ret = subprocess.call(['AimsResample', '-i', str(diBrain[id_postop[0]].fullPath()), '-m', trm_postop_to_pre, '-o', fullpost, '-t', 'n', '-r',diBrain[id_pre[0]].fullPath()])
        ret = subprocess.call(['AimsLinearComb', '-i',str(diBrain[id_pre[0]].fullPath()),'-j',fullpost, '-c', '-1', '-o', di_resec.fullPath()])
        ret = subprocess.call(['AimsThreshold', '-i',di_resec.fullPath(),'-m', 'ge','-t','250' , '-o', di_resec.fullPath()])
        ret = subprocess.call(['AimsMorphoMath', '-m','ope', '-i', di_resec.fullPath(), '-o', di_resec.fullPath(), '-r', '2'])
        ret = subprocess.call(['AimsConnectComp', '-i',di_resec.fullPath(),'-o',di_resec.fullPath(),'-c','6',])

      
      if method == 'CT':
          
        brainMask = ReadDiskItem('Brain Mask', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diBrain = list(brainMask.findValues({}, None, False ))

        id_pre = [x for x in range(len(diBrain)) if 'pre' in str(diBrain[x])]
        
        CTs = ReadDiskItem('CT', 'aims readable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
        diCTs = list(CTs.findValues({},None,False))
        for t in diCTs:
          if 'postOp' in t.attributes()['acquisition']:
            CTpostop = t.fullPath()
            
        #id_ctpostop = [x for x in range(len(diCTs)) if 'postOp' in str(diCTs[x])]
        
        #fullname_postop = diCTs[id_ctpostop[0]].fullPath()
        #fullpost_split = fullname_postop.split('/')
        #fullpost_split[-1] = 'CTpostop_on_pre.nii'
        #fullpost = '/'.join(fullpost_split)

        ret = subprocess.call(['AimsResample', '-i', str(CTpostop), '-m', trm_postop_to_pre, '-o',  di_resec.fullPath(), '-t', 'c', '-r',diBrain[id_pre[0]].fullPath()])

        ret = subprocess.call(['AimsThreshold', '-i',di_resec.fullPath(),'-m', 'be','-t','-20','-u','25','-b','-o', di_resec.fullPath()])
        ret = subprocess.call(['cartoLinearComb.py','-i',str(di_resec.fullPath()),'-i',str(diBrain[id_pre[0]]),'-o',str(di_resec.fullPath()),'-f','(I1/32767*I2/255)'])

        #apply brainmask
#./cartoLinearComb.py -i /data/brainvisa_4.5/Epilepsy/Gre_2015_CETo/t1mri/T1pre_2015-1-1/default_analysis/segmentation/brain_Gre_2015_CETo.nii -i /data/brainvisa_4.5/Epilepsy/Gre_2015_CETo/Resection/Resection/Gre_2015_CETo-Resection.nii -f 'I2/32767*I1/255' -o /tmp/test.nii        
        ret = subprocess.call(['AimsMorphoMath', '-m','ope', '-i', di_resec.fullPath(), '-o', di_resec.fullPath(), '-r', '2'])

        #la faut passer par numpy ...

        ret = subprocess.call(['AimsConnectComp', '-i',di_resec.fullPath(),'-o',di_resec.fullPath(),'-c','6',])

      vol_connectcomp = aims.read(di_resec.fullPath())
      #a small sphere here as for plots:
      sphere_size = 4
      nb_voxel_sphere = [int(round(sphere_size/vol_connectcomp.getVoxelSize()[i])) for i in range(0,3)]
      voxel_within_sphere = [vol_connectcomp.value(resec_coord[0]/vol_connectcomp.getVoxelSize()[0]+vox_i,resec_coord[1]/vol_connectcomp.getVoxelSize()[1]+vox_j,resec_coord[2]/vol_connectcomp.getVoxelSize()[2]+vox_k) for vox_k in range(-nb_voxel_sphere[2],nb_voxel_sphere[2]+1) for vox_j in range(-nb_voxel_sphere[1],nb_voxel_sphere[1]+1) for vox_i in range(-nb_voxel_sphere[0],nb_voxel_sphere[0]+1) if math.sqrt(vox_i**2+vox_j**2+vox_k**2) < sphere_size]

      voxel_to_keep = [x for x in voxel_within_sphere if x != 0]
      from collections import Counter
      most_common,num_most_common = Counter(voxel_to_keep).most_common(1)[0]

      #value_connectcomp = vol_connectcomp.value(resec_coord[0],resec_coord[1],resec_coord[2])
      ret = subprocess.call(['AimsThreshold', '-i',di_resec.fullPath(),'-m', 'eq','-t',str(most_common) , '-o', di_resec.fullPath()])

      #resave as the resection Image and do the .minf
      #ret = subprocess.call(['AimsFileConvert', '-i', str(di_resec.fullPath()), '-o', str(di_resec.fullPath()), '-t', 'S16'])
      if ret < 0:
		print "Importation error"
		QtGui.QMessageBox.warning(self, "Error", "Brainvisa Importation error/ AimsFileConvert")
		return

      neuroHierarchy.databases.insertDiskItem( di_resec, update=True )
      self.transfoManager.setReferentialTo(di_resec, self.diskItems['T1pre'].attributes()['referential'] )

      wdi_resec_roi = WriteDiskItem( 'ROI IntrAnat', 'Graph' )
      di_resec_roi = wdi_resec_roi.findValue({'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol, 'acquisition':'Resection'})

      brainvisaContext = defaultContext()

      brainvisaContext.runInteractiveProcess(lambda x='',di_roi=di_resec_roi,di_res=di_resec:self.roiconversionDone(di_roi,di_resec),'Volume To ROI Graph Converter', read = di_resec, write = di_resec_roi)  #, sulcus_identification ='label')


  def roiconversionDone(self,di_resec_roi,di_resec):


      neuroHierarchy.databases.insertDiskItem( di_resec_roi, update=True )
      self.transfoManager.setReferentialTo(di_resec_roi, self.diskItems['T1pre'].attributes()['referential'] )
      obj_roi = self.a.loadObject(di_resec_roi)

      Text_win1 = self.windowCombo1.currentText()
      obj = self.a.loadObject(di_resec)
      self.diskItems['Resection'] = di_resec
      self.dispObj['Resection']=obj
      self.windowContent.update({'Resection':['Resection','electrodes']})
      self.windowCombo1.clear()
      self.windowCombo1.addItems(sorted(self.windowContent.keys()))
      self.windowCombo2.clear()
      self.windowCombo2.addItems(sorted(self.windowContent.keys()))

      self.windowCombo1.setCurrentIndex(max(self.windowCombo1.findText(Text_win1),0))
      self.windowCombo2.setCurrentIndex(max(self.windowCombo2.findText('Resection'),0))

      self.allWindowsUpdate()

  def ROIResectiontoNiftiResection(self):

      wdi_resec = WriteDiskItem('Resection', 'NIFTI-1 image')
      di_resec = wdi_resec.findValue({'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol, 'acquisition':'Resection'})

      wdi_resec_roi = ReadDiskItem( 'ROI Intranat', 'Graph', requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol} )
      di_resec_roi = list(wdi_resec_roi.findValues({}, None, False))

      brainvisaContext = defaultContext()
      brainvisaContext.runInteractiveProcess(lambda x='',di_roi=di_resec_roi[0],di_res=di_resec:self.roiconversionDone(di_roi,di_resec),'Graph To Volume Converter', read = di_resec_roi[0], write = di_resec) #removeSource, False, extract_contours, 'No'

  def  removeFromDB(self, file, db=None):
	  """
	  If the file is a directory, recursive call to remove all its content before removing the directory.
	  Corresponding diskitem is removed from the database if it exists.
	  Taken from brainvisa-4.3.0/python/brainvisa/data/qt4gui/hierarchyBrowser.py
	  """
	  if db is None:
		try:
		  db=neuroHierarchy.databases.database(neuroHierarchy.databases.getDiskItemFromFileName(file).get("_database"))
		except:
		  pass

	  if os.path.isdir(file):
		for f in os.listdir(file):
		  self.removeFromDB(os.path.join(file, f), db)
		os.rmdir(file)
	  else:
		os.remove(file)
	  if db:
		diskItem=db.getDiskItemFromFileName(file, None)
		if diskItem:
		  db.removeDiskItem(diskItem)

  def DeleteMarsAtlasFiles(self):

	  rep = QtGui.QMessageBox.warning(self, u'Confirmation', u"<font color='red'><b>ATTENTION</b><br/>You are gonna delete MarsAtlas files, are you sure?<br/><b>DELETE MARSATLAS ?</b></font>", QtGui.QMessageBox.Yes | QtGui.QMessageBox.No, QtGui.QMessageBox.No)
	  if rep == QtGui.QMessageBox.Yes:
	      atlas_di = ReadDiskItem('hemisphere marsAtlas parcellation texture', 'aims Texture formats', requiredAttributes={ 'regularized': 'false','subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
	      atlas_di_list = list(atlas_di._findValues({}, None, False ))
	      Mask_left = ReadDiskItem('Left Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
	      diMaskleft = list(Mask_left.findValues({}, None, False ))
	      Mask_right = ReadDiskItem('Right Gyri Volume', 'Aims writable volume formats',requiredAttributes={'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
	      diMaskright = list(Mask_right.findValues({}, None, False ))

	      if len(atlas_di_list)>0:
	          for i,infoi in enumerate(atlas_di_list):
	              self.removeFromDB(infoi.fullPath(), neuroHierarchy.databases.database(infoi.get("_database")))


	      if len(diMaskleft)>0:
	          self.removeFromDB(diMaskleft[0].fullPath(), neuroHierarchy.databases.database(diMaskleft[0].get("_database")))

	      if len(diMaskright)>0:
	          self.removeFromDB(diMaskright[0].fullPath(), neuroHierarchy.databases.database(diMaskright[0].get("_database")))

	  print("MarsAtlas resulting files suppressed")
	  
	  
  def importRosaImplantation(self):
       
       TheoElectrodes = ImportTheoreticalImplentation.importRosaImplantation(self)
       #wdiTransform2 = ReadDiskItem('Transformation to Scanner Based Referential', 'Transformation matrix', exactType=True, requiredAttributes = {'subject':self.brainvisaPatientAttributes['subject'], 'center':self.currentProtocol })
       #self.t1pre2ScannerBasedTransform
       print("not finished")
       return
       for cle, valeur in TheoElectrodes.items():

           pdb.set_trace()
           rdi = ReadDiskItem('Transformation to Scanner Based Referential', 'Transformation matrix', exactType=True,  requiredAttributes={'modality':'t1mri','subject':self.brainvisaPatientAttributes['subject'], 'center':self.brainvisaPatientAttributes['center']})
           newvalTarget=[float(x) for x in valeur["target"]]
           newvalTarget=tuple(newvalTarget)
           newvalEntry=[float(x) for x in valeur["entry"]]
           newvalEntry=tuple(newvalEntry)
           volume = aims.read(str(self.diskItems['T1pre']))
           size=volume.getVoxelSize()
           sizex=size[0]
           sizey=size[1]
           sizez=size[2]
           xt=float(newvalTarget[0])/sizex
           yt=float(newvalTarget[1])/sizey
           zt=float(newvalTarget[2])/sizez
           xe=float(newvalEntry[0])/sizex
           ye=float(newvalEntry[1])/sizey
           ze=float(newvalEntry[2])/sizez
           #correctTarget=(xt,yt,zt)
           #correctEntry=(xe,ye,ze)
           #di = rdi.findValue(self.diskItems['T1pre'])
           #matriceTransfo = aims.read(di.fullPath())
           #newXYZentry = matriceTransfo.transform(correctTarget)
           #newXYZtarget= matriceTransfo.transform(correctEntry)
           #matriceTransfoT = numpy.array([[-volume.getSizeX()/2, -volume.getSizeY()/2, -volume.getSizeZ()/2, 0]])
           #correctTarget=numpy.array(newXYZtarget)
           #correctTarget=numpy.insert(correctTarget,3,1) 
           #newXYZtarget=numpy.add(correctTarget,matriceTransfoT)
           #correctEntry=numpy.array(newXYZentry)
           #correctEntry=numpy.insert(correctEntry,3,1) 
           #newXYZentry=numpy.add(correctEntry,matriceTransfoT)
           #xtf=newXYZtarget[0][0]
           #ytf=newXYZtarget[0][1]
           #ztf=newXYZtarget[0][2]
           #xef=newXYZentry[0][0]
           #yef=newXYZentry[0][1]
           #zef=newXYZentry[0][2]
           #self.t1pre2ScannerBasedTransform.transform((xt,yt,zt))
           self.addElectrode(unicode(cle), "Dixi-D08-12AM",[xt,yt,zt],[xe,ye,ze])	  

  def approximateElectrode(self):      
      
      
      self.deetoMaison=DeetoMaison(self)
      self.deetoMaison.show()	  

  # Fonction principale qui lance l'interface
def main(noapp=0):
    app = None
    if noapp == 0:
      print "NO APP"
      app = QtGui.QApplication(sys.argv)
    print "CREATE DIALOG"
    window = LocateElectrodes(app = app)
    window.show()
    if noapp == 0:
      sys.exit(app.exec_())

if __name__ == "__main__":
  print "LAUNCHING ELECTRODE LOCATE"
  axon.initializeProcesses()
  from brainvisa.data.readdiskitem import ReadDiskItem
  from brainvisa.data.writediskitem import WriteDiskItem

  print "MAIN"
  # Allow pdb to work for debugging !
  QtCore.pyqtRemoveInputHook()
  main()

#if __name__ == "__main__":
#    from PyQt5.QtGui import qApp
#    app = qApp
#    main(0)


